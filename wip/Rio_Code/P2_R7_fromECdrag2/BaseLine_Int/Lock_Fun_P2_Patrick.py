
import pybamm as pb;import pandas as pd   ;import numpy as np;import os;
import matplotlib.pyplot as plt;import os;#import imageio
from scipy.io import savemat,loadmat;from pybamm import constants,exp,sqrt;
import matplotlib as mpl; 
from multiprocessing import Queue, Process, set_start_method
from queue import Empty
fs=17; 
font = {'family' : 'DejaVu Sans','size': fs}
mpl.rc('font', **font)
import openpyxl
import traceback
import random;import time, signal
###################################################################
#############    From Patrick from Github        ##################
###################################################################
class TimeoutFunc(object):
    def __init__(self, func, timeout=None, timeout_val=None):
        assert callable(func), 'Positional argument 1 must be a callable method'
        self.func = func
        self.timeout = timeout
        self.timeout_val = timeout_val

    def _queued_f(self, *args, queue, **kwargs):
        Result_List = self.func(*args, **kwargs)
        try:
            queue.put(Result_List)
        except BrokenPipeError as exc:
            pass

    def __call__(self, *args, **kwargs):
        q = Queue(1)
        p = Process(target=self._queued_f, args=args,
            kwargs={**kwargs, 'queue': q})
        p.daemon = True
        p.start()
        try:
            Result_List = q.get(timeout=self.timeout)
        except Empty as exc: 
            # todo: need to incooperate other cases such as pybamm.expression_tree.exceptions.ModelError,
            #      pybamm.expression_tree.exceptions.SolverError
            p.terminate()
            Call_ref = RioCallback() 
            Result_List = [
                self.timeout_val,
                self.timeout_val,
                Call_ref]
        return Result_List  
###################################################################
#############    New functions from P3 - 221113        ############
###################################################################
# DEFINE my callback:
class RioCallback(pb.callbacks.Callback):
    def __init__(self, logfile=None):
        self.logfile = logfile
        self.success  = True
        if logfile is None:
            # Use pybamm's logger, which prints to command line
            self.logger = pb.logger
        else:
            # Use a custom logger, this will have its own level so set it to the same
            # level as the pybamm logger (users can override this)
            self.logger = pb.get_new_logger(__name__, logfile)
            self.logger.setLevel(pb.logger.level)
    
    def on_experiment_error(self, logs):
        self.success  = False
    def on_experiment_infeasible(self, logs):
        self.success  = False
# define to kill too long runs - Jianbo Huang kindly writes this
class TimeoutError(Exception):
	pass
def handle_signal(signal_num, frame):
	raise TimeoutError


# Define a function to calculate concentration change, whether electrolyte being squeezed out or added in
def Cal_new_con_Update(Sol,Para):   # subscript r means the reservoir
    # Note: c_EC_r is the initial EC  concentraiton in the reservoir; 
    #       c_e_r  is the initial Li+ concentraiton in the reservoir;
    #############################################################################################################################  
    ###############################           Step-1 Prepare parameters:        #################################################
    #############################################################################################################################  
    L_p   =Para["Positive electrode thickness [m]"]
    L_n   =Para["Negative electrode thickness [m]"]
    L_s   =Para["Separator thickness [m]"]
    L_y   =Para["Electrode width [m]"]   # Also update to change A_cc and therefore Q
    L_z   =Para["Electrode height [m]"]
    c_EC_r_old=Para["Current solvent concentration in the reservoir [mol.m-3]"]   # add two new parameters for paper 2
    c_e_r_old =Para["Current electrolyte concentration in the reservoir [mol.m-3]"]
    # Initial EC concentration in JR
    c_EC_JR_old =Para["Bulk solvent concentration [mol.m-3]"]  # Be careful when multiplying with pore volume to get amount in mole. Because of electrolyte dry out, that will give more amount than real.   
    # LLI due to electrode,Ratio of EC and lithium is 1:1 -> EC amount consumed is LLINegSEI[-1]
    LLINegSEI = (
        Sol["Loss of lithium to SEI [mol]"].entries[-1] 
        - Sol["Loss of lithium to SEI [mol]"].entries[0] )
    LLINegSEIcr = (
        Sol["Loss of lithium to SEI on cracks [mol]"].entries[-1]
        - 
        Sol["Loss of lithium to SEI on cracks [mol]"].entries[0]
        )
    LLINegDeadLiP = (
        Sol["Loss of lithium to dead lithium plating [mol]"].entries[-1] 
        - Sol["Loss of lithium to dead lithium plating [mol]"].entries[0])
    LLINegLiP = (
        Sol["Loss of lithium to lithium plating [mol]"].entries[-1] 
        - Sol["Loss of lithium to lithium plating [mol]"].entries[0])
    cLi_Xavg  = Sol["X-averaged electrolyte concentration [mol.m-3]"].entries[-1] 
    # Pore volume change with time:
    PoreVolNeg_0 = Sol["X-averaged negative electrode porosity"].entries[0]*L_n*L_y*L_z;
    PoreVolSep_0 = Sol["X-averaged separator porosity"].entries[0]*L_s*L_y*L_z;
    PoreVolPos_0 = Sol["X-averaged positive electrode porosity"].entries[0]*L_p*L_y*L_z;
    PoreVolNeg_1 = Sol["X-averaged negative electrode porosity"].entries[-1]*L_n*L_y*L_z;
    PoreVolSep_1 = Sol["X-averaged separator porosity"].entries[-1]*L_s*L_y*L_z;
    PoreVolPos_1 = Sol["X-averaged positive electrode porosity"].entries[-1]*L_p*L_y*L_z;
    #############################################################################################################################  
    ##### Step-2 Determine How much electrolyte is added (from the reservoir to JR) or squeezed out (from JR to reservoir) ######
    #######################   and finish electrolyte mixing     #################################################################  
    Vol_Elely_Tot_old = Para["Current total electrolyte volume in whole cell [m3]"] 
    Vol_Elely_JR_old  = Para["Current total electrolyte volume in jelly roll [m3]"] 
    if Vol_Elely_Tot_old - Vol_Elely_JR_old < 0:
        print('Model error! Electrolyte in JR is larger than in the cell!')
    Vol_Pore_tot_old  = PoreVolNeg_0 + PoreVolSep_0 + PoreVolPos_0    # pore volume at start time of the run
    Vol_Pore_tot_new  = PoreVolNeg_1 + PoreVolSep_1 + PoreVolPos_1    # pore volume at end   time of the run, intrinsic variable 
    Vol_Pore_decrease = Vol_Elely_JR_old  - Vol_Pore_tot_new;
    # EC:lithium:SEI=2:1:1     for SEI=(CH2OCO2Li)2, but because of too many changes are required, change to 2:1:1 for now
    # Because inner and outer SEI partial molar volume is the same, just set one for whole SEI
    VmolSEI   = Para["Outer SEI partial molar volume [m3.mol-1]"] # 9.8e-5,
    VmolLiP   = Para["Lithium metal partial molar volume [m3.mol-1]"] # 1.3e-05
    VmolEC    = 6.667e-5  # Unit:m3.mol-1; According to Wiki, correct value: EC molar volume is :66.67 cm3.mol-1  = 6.667e-5, 
    Vol_EC_consumed  =  ( LLINegSEI + LLINegSEIcr + LLINegDeadLiP  ) * 2 * VmolEC    # Mark: Ruihe add LLINegDeadLiP, either with 2 or not, will decide how fast electrolyte being consumed!
    Vol_Elely_need   = Vol_EC_consumed - Vol_Pore_decrease
    Vol_SEILiP_increase = 1.0*(
        (LLINegSEI+LLINegSEIcr) * VmolSEI 
        + LLINegLiP * VmolLiP)    #  volume increase due to SEI+total LiP 
    Test_V = Vol_SEILiP_increase - Vol_Pore_decrease  #  This value should always be zero, but now not, which becomes the source of error!
    Test_V2= (Vol_Pore_tot_old - Vol_Elely_JR_old) / Vol_Elely_JR_old * 100; # Porosity errors due to first time step
    
    # Start from here, there are serveral variables to be determined:
    #   1) Vol_Elely_add, or Vol_Elely_squeezed; depends on conditions. easier for 'squeezed' condition
    #   2) Vol_Elely_Tot_new, should always equals to Vol_Elely_Tot_old -  Vol_EC_consumed;
    #   3) Vol_Elely_JR_new, for 'add' condition: see old code; for 'squeezed' condition, equals to pore volume in JR
    #   4) Ratio_Dryout, for 'add' condition: see old code; for 'squeezed' condition, equals to Vol_Elely_Tot_new/Vol_Pore_tot_new 
    #   5) Ratio_CeEC_JR and Ratio_CeLi_JR: for 'add' condition: see old code; for 'squeezed' condition, equals to 1 (unchanged)
    #   6) c_e_r_new and c_EC_r_new: for 'add' condition: equals to old ones (unchanged); for 'squeezed' condition, need to carefully calculate     
    #   7) Width_new: for 'add' condition: see old code; for 'squeezed' condition, equals to L_y (unchanged)
    Vol_Elely_Tot_new = Vol_Elely_Tot_old - Vol_EC_consumed;

    
    if Vol_Elely_need < 0:
        print('Electrolyte is being squeezed out, check plated lithium (reversible part)')
        Vol_Elely_squeezed = - Vol_Elely_need;   # Make Vol_Elely_squeezed>0 for simplicity 
        Vol_Elely_add = 0.0;
        Vol_Elely_JR_new = Vol_Pore_tot_new; 
        Ratio_Dryout = Vol_Elely_Tot_new / Vol_Elely_JR_new
        Ratio_CeEC_JR = 1.0;    # the concentrations in JR don't need to change
        Ratio_CeLi_JR = 1.0; 
        SqueezedLiMol   =  Vol_Elely_squeezed*cLi_Xavg;
        SqueezedECMol   =  Vol_Elely_squeezed*c_EC_JR_old;
        Vol_Elely_reservoir_old = Vol_Elely_Tot_old - Vol_Elely_JR_old; 
        LiMol_reservoir_new = Vol_Elely_reservoir_old*c_e_r_old  + SqueezedLiMol;
        ECMol_reservoir_new = Vol_Elely_reservoir_old*c_EC_r_old + SqueezedECMol;
        c_e_r_new= LiMol_reservoir_new / (Vol_Elely_reservoir_old+Vol_Elely_squeezed);
        c_EC_r_new= ECMol_reservoir_new / (Vol_Elely_reservoir_old+Vol_Elely_squeezed);
        Width_new = L_y; 
    else:   # this means Vol_Elely_need >= 0, therefore the folliwing script should works for Vol_Elely_need=0 as well!
        # Important: Calculate added electrolyte based on excessive electrolyte, can be: 1) added as required; 2) added some, but less than required; 3) added 0 
        Vol_Elely_squeezed = 0;  
        if Vol_Elely_Tot_old > Vol_Elely_JR_old:                             # This means Vol_Elely_JR_old = Vol_Pore_tot_old ()
            if Vol_Elely_Tot_old-Vol_Elely_JR_old >= Vol_Elely_need:         # 1) added as required
                Vol_Elely_add     = Vol_Elely_need;  
                Vol_Elely_JR_new  = Vol_Pore_tot_new;  # also equals to 'Vol_Pore_tot_new', or Vol_Pore_tot_old - Vol_Pore_decrease
                Ratio_Dryout = 1.0;
            else:                                                            # 2) added some, but less than required;                                                         
                Vol_Elely_add     = Vol_Elely_Tot_old - Vol_Elely_JR_old;   
                Vol_Elely_JR_new  = Vol_Elely_Tot_new;                       # This means Vol_Elely_JR_new <= Vol_Pore_tot_new
                Ratio_Dryout = Vol_Elely_JR_new/Vol_Pore_tot_new;
        else:                                                                # 3) added 0 
            Vol_Elely_add = 0;
            Vol_Elely_JR_new  = Vol_Elely_Tot_new; 
            Ratio_Dryout = Vol_Elely_JR_new/Vol_Pore_tot_new;

        # Next: start mix electrolyte based on previous defined equation
        # Lithium amount in liquid phase, at initial time point
        TotLi_Elely_JR_Old = Sol["Total lithium in electrolyte [mol]"].entries[-1]; # remember this is only in JR
        # Added lithium and EC amount in the added lithium electrolyte: - 
        AddLiMol   =  Vol_Elely_add*c_e_r_old
        AddECMol   =  Vol_Elely_add*c_EC_r_old
        # Total amount of Li and EC in current electrolyte:
        # Remember Li has two parts, initial and added; EC has three parts, initial, consumed and added
        TotLi_Elely_JR_New   = TotLi_Elely_JR_Old + AddLiMol
        TotECMol_JR   = Vol_Elely_JR_old*c_EC_JR_old - LLINegSEI + AddECMol  # EC:lithium:SEI=2:2:1     for SEI=(CH2OCO2Li)2
        Ratio_CeEC_JR  = TotECMol_JR    /   Vol_Elely_JR_new   / c_EC_JR_old
        Ratio_CeLi_JR  = TotLi_Elely_JR_New    /   TotLi_Elely_JR_Old   /  Ratio_Dryout # Mark, change on 21-11-19
        c_e_r_new  = c_e_r_old;
        c_EC_r_new = c_EC_r_old;
        Width_new   = Ratio_Dryout * L_y;
        
    # Collect the above parameter in Data_Pack to shorten the code  
    Data_Pack   = [
        Vol_EC_consumed, 
        Vol_Elely_need, 
        Test_V, 
        Test_V2, 
        Vol_Elely_add, 
        Vol_Elely_Tot_new, 
        Vol_Elely_JR_new, 
        Vol_Pore_tot_new, 
        Vol_Pore_decrease, 
        c_e_r_new, c_EC_r_new,
        Ratio_Dryout, Ratio_CeEC_JR, 
        Ratio_CeLi_JR,
        Width_new, 
        ]  # 16 in total
    #print('Loss of lithium to SEI', LLINegSEI, 'mol') 
    #print('Loss of lithium to dead lithium plating', LLINegDeadLiP, 'mol') 
    #############################################################################################################################  
    ###################       Step-4 Update parameters here        ##############################################################
    #############################################################################################################################
    Para.update(   {'Bulk solvent concentration [mol.m-3]':  c_EC_JR_old * Ratio_CeEC_JR  })
    Para.update(
        {'EC initial concentration in electrolyte [mol.m-3]':c_EC_JR_old * Ratio_CeEC_JR }, 
        check_already_exists=False) 
    Para.update(   
        {'Ratio of Li-ion concentration change in electrolyte consider solvent consumption':  
        Ratio_CeLi_JR }, check_already_exists=False) 
    Para.update(   {'Current total electrolyte volume in whole cell [m3]':  Vol_Elely_Tot_new  }, check_already_exists=False)
    Para.update(   {'Current total electrolyte volume in jelly roll [m3]':  Vol_Elely_JR_new  }, check_already_exists=False)
    Para.update(   {'Ratio of electrolyte dry out in jelly roll':Ratio_Dryout}, check_already_exists=False)
    Para.update(   {'Electrode width [m]':Width_new})    
    Para.update(   {'Current solvent concentration in the reservoir [mol.m-3]':c_EC_r_new}, check_already_exists=False)     
    Para.update(   {'Current electrolyte concentration in the reservoir [mol.m-3]':c_e_r_new}, check_already_exists=False)             
    return Data_Pack,Para

# Define a function to calculate based on previous solution
def Run_Model_Base_On_Last_Solution( 
    Model  , Sol , Para_update, ModelExperiment, 
    Update_Cycles,Temper_i ,mesh_par,submesh_strech):
    # Use Sulzer's method: inplace = false
    # Important line: define new model based on previous solution
    Ratio_CeLi = Para_update["Ratio of Li-ion concentration change in electrolyte consider solvent consumption"]
    dict_short = {}; 
    list_short = []
    # update 220808 to satisfy random model option:
    for var, equation in Model.initial_conditions.items():
        #print(var._name)
        list_short.append(var._name)
    # delete Porosity times concentration and Electrolyte potential then add back
    list_short.remove("Porosity times concentration");
    list_short.remove("Electrolyte potential");
    list_short.extend(
        ("Negative electrode porosity times concentration",
        "Separator porosity times concentration",
        "Positive electrode porosity times concentration",   

        "Negative electrolyte potential",
        "Separator electrolyte potential",
        "Positive electrolyte potential",))
    for list_short_i in list_short:
        dict_short.update( { list_short_i : Sol.last_state[list_short_i].data  }  )
    dict_short["Negative electrode porosity times concentration"] = (
        dict_short["Negative electrode porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    dict_short["Separator porosity times concentration"] = (
        dict_short["Separator porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    dict_short["Positive electrode porosity times concentration"] = (
        dict_short["Positive electrode porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    Model_new = Model.set_initial_conditions_from(dict_short, inplace=False)
    Para_update.update(   {'Ambient temperature [K]':Temper_i });   # run model at 45 degree C
    
    var = pb.standard_spatial_vars  
    var_pts = {
        var.x_n: 20,  
        var.x_s: 10,  
        var.x_p: 20,  
        var.r_n: int(mesh_par),  
        var.r_p: int(mesh_par),  
        }
    submesh_types = Model.default_submesh_types
    if submesh_strech == "nan":
            pass
    else:
        particle_mesh = pb.MeshGenerator(
            pb.Exponential1DSubMesh, 
            submesh_params={"side": "right", "stretch": int(submesh_strech)})
        submesh_types["negative particle"] = particle_mesh
        submesh_types["positive particle"] = particle_mesh 
    Call_Age = RioCallback()  # define callback
    Simnew = pb.Simulation(
        Model_new,
        experiment = ModelExperiment, 
        parameter_values=Para_update, 
        solver = pb.CasadiSolver(),
        var_pts = var_pts,
        submesh_types=submesh_types
    )
    try:
        Sol_new = Simnew.solve(
            calc_esoh=False,
            save_at_cycles = Update_Cycles,
            callbacks=Call_Age)
    except (
        pb.expression_tree.exceptions.ModelError,
        pb.expression_tree.exceptions.SolverError
        ) as e:
        Sol_new = "Model error or solver error"
    else:
        pass
    # print("Solved this model in {}".format(ModelTimer.time()))
    Result_list = [Model_new, Sol_new,Call_Age]
    return Result_list

def Run_Model_Base_On_Last_Solution_RPT( 
    Model  , Sol,  Para_update, 
    ModelExperiment ,Update_Cycles, Temper_i,mesh_par,submesh_strech):
    # Use Sulzer's method: inplace = false
    Ratio_CeLi = Para_update["Ratio of Li-ion concentration change in electrolyte consider solvent consumption"]
    # print("Model is now using average EC Concentration of:",Para_update['Bulk solvent concentration [mol.m-3]'])
    # print("Ratio of electrolyte dry out in jelly roll is:",Para_update['Ratio of electrolyte dry out in jelly roll'])
    # print("Model is now using an electrode width of:",Para_update['Electrode width [m]'])
    # Important line: define new model based on previous solution
    dict_short = {}; 
    list_short = []
    # update 220808 to satisfy random model option:
    for var, equation in Model.initial_conditions.items():
        #print(var._name)
        list_short.append(var._name)
    # delete Porosity times concentration and Electrolyte potential then add back
    list_short.remove("Porosity times concentration");
    list_short.remove("Electrolyte potential");
    list_short.extend(
        ("Negative electrode porosity times concentration",
        "Separator porosity times concentration",
        "Positive electrode porosity times concentration",   

        "Negative electrolyte potential",
        "Separator electrolyte potential",
        "Positive electrolyte potential",))
    for list_short_i in list_short:
        dict_short.update( { list_short_i : Sol.last_state[list_short_i].data  }  )
    dict_short["Negative electrode porosity times concentration"] = (
        dict_short["Negative electrode porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    dict_short["Separator porosity times concentration"] = (
        dict_short["Separator porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    dict_short["Positive electrode porosity times concentration"] = (
        dict_short["Positive electrode porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    Model_new = Model.set_initial_conditions_from(dict_short, inplace=False)
    Para_update.update(   {'Ambient temperature [K]':Temper_i });
    
    var = pb.standard_spatial_vars  
    var_pts = {var.x_n: 20,  var.x_s: 10,  var.x_p: 20,  var.r_n: int(mesh_par),  var.r_p: int(mesh_par),  }
    submesh_types = Model.default_submesh_types
    if submesh_strech == "nan":
        pass
    else:
        particle_mesh = pb.MeshGenerator(
            pb.Exponential1DSubMesh, 
            submesh_params={"side": "right", "stretch": int(submesh_strech)})
        submesh_types["negative particle"] = particle_mesh
        submesh_types["positive particle"] = particle_mesh 
    Simnew = pb.Simulation(
        Model_new,
        experiment = ModelExperiment, 
        parameter_values=Para_update, 
        solver = pb.CasadiSolver(),
        var_pts = var_pts,
        submesh_types=submesh_types
    )
    Call_RPT = RioCallback()  # define callback
    try:
        Sol_new = Simnew.solve(
            calc_esoh=False,
            save_at_cycles = Update_Cycles,
            callbacks=Call_RPT) 
    except (
        pb.expression_tree.exceptions.ModelError,
        pb.expression_tree.exceptions.SolverError
        ) as e:
        Sol_new = "Model error or solver error"
    else:
        pass
    # print("Solved this model in {}".format(ModelTimer.time()))
    Result_List_RPT = [Model_new, Sol_new,Call_RPT]
    return Result_List_RPT

def write_excel_xlsx(path, sheet_name, value):
    import numpy as np
    index = len(value)
    workbook = openpyxl.Workbook()  # 新建工作簿（默认有一个sheet？）
    sheet = workbook.active  # 获得当前活跃的工作页，默认为第一个工作页
    sheet.title = sheet_name  # 给sheet页的title赋值
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))  # 行，列，值 这里是从1开始计数的
    workbook.save(path)  # 一定要保存
    print("Successfully create a excel file")
# From 22-08-15, this function is replaced by recursive_scan
def GetScan(
    Ratio_excess,cs_Neg_Init,Diff_SEI,R_SEI,Bulk_Sol_Con,
    D_Li_inSEI,c_Li_inte_ref,
    Diff_EC,k_SEI,LAMcr_prop,Crack_rate,
    Couple_SEI_LiP,k_LiP,Temper,MESH_PAR):
    import numpy as np
    TotalScan =(
        len(Ratio_excess)*
        len(cs_Neg_Init)*
        len(Diff_SEI)*
        len(R_SEI)*
        len(Bulk_Sol_Con)*
        len(D_Li_inSEI)*
        len(c_Li_inte_ref)*
        len(Diff_EC)*          # 
        len(k_SEI)*            #
        len(LAMcr_prop)*       #
        len(Crack_rate)*       #
        len(Couple_SEI_LiP)*
        len(k_LiP)*
        len(Temper)*
        len(MESH_PAR) 
        ) ;
    DatePack_scan = np.full([TotalScan,16], 0.0);
    index = 0;
    for Ratio_excess_i in Ratio_excess:
        for cs_Neg_Init_i in cs_Neg_Init:
            for Diff_SEI_i in Diff_SEI:
                for R_SEI_i in R_SEI:
                    for Bulk_Sol_Con_i in Bulk_Sol_Con:
                        for D_Li_inSEI_i in D_Li_inSEI:
                            for c_Li_inte_ref_i in c_Li_inte_ref:
                                for Diff_EC_i in Diff_EC:
                                    for k_SEI_i in k_SEI:
                                        for LAMcr_prop_i in LAMcr_prop:
                                            for Crack_rate_i in Crack_rate:
                                                for Couple_SEI_LiP_i in Couple_SEI_LiP:
                                                    for k_LiP_i in k_LiP:
                                                        for Temper_i in Temper:
                                                            for mesh_par in MESH_PAR:
                                                                index +=  1;
                                                                DatePack_scan[index-1] = [
                                                                    index,
                                                                    Ratio_excess_i,
                                                                    cs_Neg_Init_i,Diff_SEI_i,
                                                                    R_SEI_i,Bulk_Sol_Con_i,
                                                                    D_Li_inSEI_i,c_Li_inte_ref_i,
                                                                    Diff_EC_i,k_SEI_i,           # 
                                                                    LAMcr_prop_i,Crack_rate_i,   #
                                                                    Couple_SEI_LiP_i,k_LiP_i,Temper_i,
                                                                    mesh_par];
    return TotalScan, DatePack_scan

def recursive_scan(mylist,kvs, key_list, acc):
    # 递归终止条件
    if len(key_list) == 0:
        mylist.append(acc.copy())   # copy.deepcopy(acc) 如果value是个list，就要deep copy了
        return mylist
    # 继续递归
    k = key_list[0]
    for v in kvs[k]:
        acc[k] = v
        # print(v)
        recursive_scan(mylist,kvs, key_list[1:], acc)

# this function is to initialize the para with a known dict
def Para_init(Para_dict):
    Para_dict_used = Para_dict.copy();
    Para_0=pb.ParameterValues(Para_dict_used["Para_Set"]  )
    Para_dict_used.pop("Para_Set")

    if Para_dict_used.__contains__("Total ageing cycles"):
        Total_Cycles = Para_dict_used["Total ageing cycles"]  
        Para_dict_used.pop("Total ageing cycles")
    if Para_dict_used.__contains__("Ageing cycles between RPT"):
        Cycle_bt_RPT = Para_dict_used["Ageing cycles between RPT"]  
        Para_dict_used.pop("Ageing cycles between RPT")
    if Para_dict_used.__contains__("Update cycles for ageing"):
        Update_Cycles = Para_dict_used["Update cycles for ageing"]  
        Para_dict_used.pop("Update cycles for ageing")
    if Para_dict_used.__contains__("Cycles within RPT"):
        RPT_Cycles = Para_dict_used["Cycles within RPT"]  
        Para_dict_used.pop("Cycles within RPT")
    if Para_dict_used.__contains__("Ageing temperature"):
        Temper_i = Para_dict_used["Ageing temperature"]  + 273.15 # update: change to K
        Para_dict_used.pop("Ageing temperature")
    if Para_dict_used.__contains__("RPT temperature"):
        Temper_RPT = Para_dict_used["RPT temperature"] + 273.15 # update: change to K 
        Para_dict_used.pop("RPT temperature")
    if Para_dict_used.__contains__("Particle mesh points"):
        mesh_par = Para_dict_used["Particle mesh points"]  
        Para_dict_used.pop("Particle mesh points")
    if Para_dict_used.__contains__("Exponential mesh stretch"):
        submesh_strech = Para_dict_used["Exponential mesh stretch"]  
        Para_dict_used.pop("Exponential mesh stretch")
    else:
        submesh_strech = "nan";
    if Para_dict_used.__contains__("Model option"):
        model_options = Para_dict_used["Model option"]  
        Para_dict_used.pop("Model option")
    # Mark Ruihe - updated 221113 - from P3
    if Para_dict_used.__contains__("Func Electrolyte conductivity [S.m-1]"):
        Para_0.update({
            "Electrolyte conductivity [S.m-1]": 
            eval(Para_dict_used["Func Electrolyte conductivity [S.m-1]"])})  
        Para_dict_used.pop("Func Electrolyte conductivity [S.m-1]")
    if Para_dict_used.__contains__("Func Electrolyte diffusivity [m2.s-1]"):
        Para_0.update({
            "Electrolyte diffusivity [m2.s-1]": 
            eval(Para_dict_used["Func Electrolyte diffusivity [m2.s-1]"])})
        Para_dict_used.pop("Func Electrolyte diffusivity [m2.s-1]")

    if Para_dict_used.__contains__("Initial Neg SOC"):
        c_Neg1SOC_in = (
            Para_0["Maximum concentration in negative electrode [mol.m-3]"]
            *Para_dict_used["Initial Neg SOC"]  )
        Para_0.update(
            {"Initial concentration in negative electrode [mol.m-3]":
            c_Neg1SOC_in})
        Para_dict_used.pop("Initial Neg SOC")
    if Para_dict_used.__contains__("Initial Pos SOC"):    
        c_Pos1SOC_in = (
            Para_0["Maximum concentration in positive electrode [mol.m-3]"]
            *Para_dict_used["Initial Pos SOC"]  )
        Para_0.update(
            {"Initial concentration in positive electrode [mol.m-3]":
            c_Pos1SOC_in})
        Para_dict_used.pop("Initial Pos SOC")

    CyclePack = [ 
        Total_Cycles,Cycle_bt_RPT,Update_Cycles,RPT_Cycles,
        Temper_i,Temper_RPT,mesh_par,submesh_strech,model_options];
    
    for key, value in Para_dict_used.items():
        # risk: will update parameter that doesn't exist, so need to make sure the name is right 
        Para_0.update({key: value},check_already_exists=False)
    return CyclePack,Para_0

# Add 220808 - to simplify the post-processing
def GetSol_dict (my_dict, keys_all, Sol, 
    cycle_no,step_CD, step_CC , step_RE, step_CV ):

    [keys_loc,keys_tim,keys_cyc]  = keys_all;
    # get time_based variables:
    if len(keys_tim): 
        for key in keys_tim:
            step_no = eval("step_{}".format(key[0:2]))
            if key[3:] == "Time [h]":
                my_dict[key].append  (
                    Sol.cycles[cycle_no].steps[step_no][key[3:]].entries
                    -
                    Sol.cycles[cycle_no].steps[step_no][key[3:]].entries[0])
            else:
                #print("Solve up to Step",step_no)
                my_dict[key].append  (
                    Sol.cycles[cycle_no].steps[step_no][key[3:]].entries)
    # get cycle_step_based variables:
    if len(keys_cyc): 
        for key in keys_cyc:
            if key in ["Discharge capacity [A.h]"]:
                step_no = step_CD;
                my_dict[key].append  (
                    Sol.cycles[cycle_no].steps[step_no][key].entries[-1]
                    - 
                    Sol.cycles[cycle_no].steps[step_no][key].entries[0])
            elif key[0:5] in ["CDend","CCend","CVend","REend",
                            "CDsta","CCsta","CVsta","REsta",]:
                step_no = eval("step_{}".format(key[0:2]))
                if key[2:5] == "sta":
                    my_dict[key].append  (
                        Sol.cycles[cycle_no].steps[step_no][key[6:]].entries[0])
                elif key[2:5] == "end":
                    my_dict[key].append  (
                        Sol.cycles[cycle_no].steps[step_no][key[6:]].entries[-1])
                
    # get location_based variables:
    if len(keys_loc): 
        for key in keys_loc:
            if key in ["x_n [m]","x [m]","x_s [m]","x_p [m]"]:
                #print("These variables only add once")
                if not len(my_dict[key]):   # special: add only once
                    my_dict[key] = Sol[key].entries[:,-1]
            elif key[0:5] in ["CDend","CCend","CVend","REend",
                            "CDsta","CCsta","CVsta","REsta",]:      
                #print("These variables add multiple times")
                step_no = eval("step_{}".format(key[0:2]))
                if key[2:5] == "sta":
                    my_dict[key].append  (
                        Sol.cycles[cycle_no].steps[step_no][key[6:]].entries[:,0])
                elif key[2:5] == "end":
                    my_dict[key].append  (
                        Sol.cycles[cycle_no].steps[step_no][key[6:]].entries[:,-1])
    return my_dict                              

############## Get initial cap ############## 
# Input: just parameter set
# Output: 0% and 100% SOC of neg/pos; cap at last RPT cycle
# Simply run RPT cycle for 3 times
def Get_initial_cap(Para_dict_i,Neg1SOC_in,Pos1SOC_in,):
    # Para_dict_i contains all global settings
    CyclePack,Para_0 = Para_init(Para_dict_i)
    [
        Total_Cycles,Cycle_bt_RPT,Update_Cycles,
        RPT_Cycles,Temper_i,Temper_RPT,
        mesh_par,submesh_strech,
        model_options] = CyclePack;

    # define experiment
    V_max = 4.2;        
    V_min = 2.5; 
    exp_RPT_text = [ (
        f"Discharge at 0.1C until {V_min} V",  
        "Rest for 1 hours",  
        f"Charge at 0.1C until {V_max} V" ) ] # (5 minute period)
    Experiment_RPT    = pb.Experiment( exp_RPT_text * RPT_Cycles     ) 

    Model_0 = pb.lithium_ion.DFN(options=model_options ) 

    var = pb.standard_spatial_vars  
    var_pts = {
        var.x_n: 20,  
        var.x_s: 10,  
        var.x_p: 20,  
        var.r_n: int(mesh_par),  
        var.r_p: int(mesh_par),  }       
    submesh_types = Model_0.default_submesh_types
    if submesh_strech == "nan":
        pass
    else:
        particle_mesh = pb.MeshGenerator(
            pb.Exponential1DSubMesh, 
            submesh_params={"side": "right", "stretch": submesh_strech})
        submesh_types["negative particle"] = particle_mesh
        submesh_types["positive particle"] = particle_mesh 
    
    c_Neg1SOC_in = (
        Para_0["Maximum concentration in negative electrode [mol.m-3]"]
        *Neg1SOC_in)
    c_Pos1SOC_in = (
        Para_0["Maximum concentration in positive electrode [mol.m-3]"]
        *Pos1SOC_in)
    Para_0.update(
        {"Initial concentration in negative electrode [mol.m-3]":
        c_Neg1SOC_in})
    Para_0.update(
        {"Initial concentration in positive electrode [mol.m-3]":
        c_Pos1SOC_in})
    Sim_0    = pb.Simulation(
            Model_0,        experiment = Experiment_RPT,
            parameter_values = Para_0,
            solver = pb.CasadiSolver(),
            var_pts=var_pts,
            submesh_types=submesh_types) #mode="safe"
    Sol_0    = Sim_0.solve(calc_esoh=False)

    Cap = [];  Neg1SOC=[]; Pos1SOC = [];

    for i in range(0,RPT_Cycles):
        Cap.append(
            Sol_0.cycles[i].steps[0]["Discharge capacity [A.h]"].entries[-1] - 
            Sol_0.cycles[i].steps[0]["Discharge capacity [A.h]"].entries[0]) 
        Neg1SOC.append(
            Sol_0.cycles[i].steps[0]["Negative electrode SOC"].entries[0]) 
        Pos1SOC.append(
            Sol_0.cycles[i].steps[0]["Positive electrode SOC"].entries[0]) 
        
    return Sol_0, Cap, Neg1SOC, Pos1SOC

def Get_initial_cap2(Para_dict_i):
    # Para_dict_i contains all global settings
    CyclePack,Para_0 = Para_init(Para_dict_i)
    [
        Total_Cycles,Cycle_bt_RPT,Update_Cycles,
        RPT_Cycles,Temper_i,Temper_RPT,
        mesh_par,submesh_strech,
        model_options] = CyclePack;

    # define experiment
    V_max = 4.2;        
    V_min = 2.5; 
    exp_RPT_text = [ (
        f"Discharge at 0.1C until {V_min} V (30 minute period)",  
        "Rest for 1 hours (5 minute period)",  
        f"Charge at 0.1C until {V_max} V (30 minute period)" ) ] # 
    Experiment_RPT    = pb.Experiment( exp_RPT_text * RPT_Cycles     ) 

    Model_0 = pb.lithium_ion.DFN(options=model_options ) 

    var = pb.standard_spatial_vars  
    var_pts = {
        var.x_n: 20,  
        var.x_s: 10,  
        var.x_p: 20,  
        var.r_n: int(mesh_par),  
        var.r_p: int(mesh_par),  }       
    submesh_types = Model_0.default_submesh_types
    if submesh_strech == "nan":
        pass
    else:
        particle_mesh = pb.MeshGenerator(
            pb.Exponential1DSubMesh, 
            submesh_params={"side": "right", "stretch": submesh_strech})
        submesh_types["negative particle"] = particle_mesh
        submesh_types["positive particle"] = particle_mesh 

    Sim_0    = pb.Simulation(
            Model_0,        experiment = Experiment_RPT,
            parameter_values = Para_0,
            solver = pb.CasadiSolver(),
            var_pts=var_pts,
            submesh_types=submesh_types) #mode="safe"
    Sol_0    = Sim_0.solve(calc_esoh=False)

    Cap = [];  Neg1SOC=[]; Pos1SOC = [];

    for i in range(0,RPT_Cycles):
        Cap.append(
            Sol_0.cycles[i].steps[0]["Discharge capacity [A.h]"].entries[-1] - 
            Sol_0.cycles[i].steps[0]["Discharge capacity [A.h]"].entries[0]) 
        Neg1SOC.append(
            Sol_0.cycles[i].steps[0]["Negative electrode SOC"].entries[0]) 
        Pos1SOC.append(
            Sol_0.cycles[i].steps[0]["Positive electrode SOC"].entries[0]) 
        
    return Sol_0,Cap, Neg1SOC, Pos1SOC

# define the model and run break-in cycle - 
# input parameter: model_options, Experiment_Breakin, Para_0, mesh_par, submesh_strech
# output: Sol_0 , Model_0, Call_Breakin
def Run_Breakin(
    model_options, Experiment_Breakin, 
    Para_0, mesh_par, submesh_strech):

    Model_0 = pb.lithium_ion.DFN(options=model_options ) #
    # update 220926 - add diffusivity and conductivity as variables:
    c_e = Model_0.variables["Electrolyte concentration [mol.m-3]"]
    T = Model_0.variables["Cell temperature [K]"]
    D_e = Para_0["Electrolyte diffusivity [m2.s-1]"]
    sigma_e = Para_0["Electrolyte conductivity [S.m-1]"]
    Model_0.variables["Electrolyte diffusivity [m2.s-1]"] = D_e(c_e, T)
    Model_0.variables["Electrolyte conductivity [S.m-1]"] = sigma_e(c_e, T)
    var = pb.standard_spatial_vars  
    var_pts = {
        var.x_n: 20,  
        var.x_s: 10,  
        var.x_p: 20,  
        var.r_n: int(mesh_par),  
        var.r_p: int(mesh_par),  }       
    submesh_types = Model_0.default_submesh_types
    if submesh_strech == "nan":
        pass
    else:
        particle_mesh = pb.MeshGenerator(
            pb.Exponential1DSubMesh, 
            submesh_params={"side": "right", "stretch": submesh_strech})
        submesh_types["negative particle"] = particle_mesh
        submesh_types["positive particle"] = particle_mesh 
    Sim_0    = pb.Simulation(
        Model_0,        experiment = Experiment_Breakin,
        parameter_values = Para_0,
        solver = pb.CasadiSolver(),
        var_pts=var_pts,
        submesh_types=submesh_types) #mode="safe"
    Call_Breakin = RioCallback()
    try:
        Sol_0    = Sim_0.solve(calc_esoh=False,callbacks=Call_Breakin)
    except (
        pb.expression_tree.exceptions.ModelError,
        pb.expression_tree.exceptions.SolverError
        ) as e:
        Sol_0 = "Model error or solver error"
    else:
        pass
    Result_list_breakin = [Model_0,Sol_0,Call_Breakin]

    return Result_list_breakin

# Input: Para_0
# Output: mdic_dry, Para_0
def Initialize_mdic_dry(Para_0,Int_ElelyExces_Ratio):
    mdic_dry = {
        "CeEC_All": [],
        "c_EC_r_new_All": [],
        "c_e_r_new_All": [],
        "Ratio_CeEC_All":[],
        "Ratio_CeLi_All":[],
        "Ratio_Dryout_All":[],

        "Vol_Elely_Tot_All": [],
        "Vol_Elely_JR_All":[],
        "Vol_Pore_tot_All": [],        
        "Vol_EC_consumed_All":[],
        "Vol_Elely_need_All":[],
        "Width_all":[],
        "Vol_Elely_add_All":[],
        "Vol_Pore_decrease_All":[],
        "Test_V_All":[],
        "Test_V2_All":[],
    }
    T_0                  =  Para_0['Initial temperature [K]']
    Porosity_Neg_0       =  Para_0['Negative electrode porosity']  
    Porosity_Pos_0       =  Para_0['Positive electrode porosity']  
    Porosity_Sep_0       =  Para_0['Separator porosity']  
    cs_Neg_Max           =  Para_0["Maximum concentration in negative electrode [mol.m-3]"];
    L_p                  =  Para_0["Positive electrode thickness [m]"]
    L_n                  =  Para_0["Negative electrode thickness [m]"]
    L_s                  =  Para_0["Separator thickness [m]"]
    L_y                  =  Para_0["Electrode width [m]"]
    Para_0.update({'Initial Electrode width [m]':L_y}, check_already_exists=False)
    L_y_0                =  Para_0["Initial Electrode width [m]"]
    L_z                  =  Para_0["Electrode height [m]"]
    Para_0.update({'Initial Electrode height [m]':L_z}, check_already_exists=False)
    L_z_0                =  Para_0["Initial Electrode height [m]"]
    
    Vol_Elely_Tot        = (
        ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  
        * L_y_0 * L_z_0 * Int_ElelyExces_Ratio
     ) # Set initial electrolyte amount [L] 
    Vol_Elely_JR         =  (
        ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  
        * L_y_0 * L_z_0 )
    Vol_Pore_tot         =  (
        ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  
        * L_y_0 * L_z_0 )
    Ratio_CeEC           =  1.0; Ratio_CeLi =  1.0  ;Ratio_Dryout         =  1.0
    Vol_EC_consumed      =  0;Vol_Elely_need       =  0;Vol_Elely_add        =  0
    Vol_Pore_decrease    =  0;Test_V2 = 0; Test_V = 0;
    print('Initial electrolyte amount is ', Vol_Elely_Tot*1e6, 'mL') 
    Para_0.update(
        {'Current total electrolyte volume in jelly roll [m3]':
        Vol_Elely_JR}, check_already_exists=False)
    Para_0.update(
        {'Current total electrolyte volume in whole cell [m3]':
        Vol_Elely_Tot}, check_already_exists=False)   
    
    mdic_dry["Vol_Elely_Tot_All"].append(Vol_Elely_Tot*1e6);            
    mdic_dry["Vol_Elely_JR_All"].append(Vol_Elely_JR*1e6);     
    mdic_dry["Vol_Pore_tot_All"].append(Vol_Pore_tot*1e6);           
    mdic_dry["Ratio_CeEC_All"].append(Ratio_CeEC);                      
    mdic_dry["Ratio_CeLi_All"].append(Ratio_CeLi);             
    mdic_dry["Ratio_Dryout_All"].append(Ratio_Dryout);
    mdic_dry["Vol_EC_consumed_All"].append(Vol_EC_consumed*1e6);        
    mdic_dry["Vol_Elely_need_All"].append(Vol_Elely_need*1e6);     
    mdic_dry["Width_all"].append(L_y_0);
    mdic_dry["Vol_Elely_add_All"].append(Vol_Elely_add*1e6);            
    mdic_dry["Vol_Pore_decrease_All"].append(Vol_Pore_decrease*1e6);
    mdic_dry["Test_V_All"].append(Test_V*1e6); 
    mdic_dry["Test_V2_All"].append(Test_V2*1e6); 
    mdic_dry["c_e_r_new_All"].append(Para_0["Initial concentration in electrolyte [mol.m-3]"]); 
    mdic_dry["c_EC_r_new_All"].append(Para_0['EC initial concentration in electrolyte [mol.m-3]'])

    return mdic_dry,Para_0
    
def Update_mdic_dry(Data_Pack,mdic_dry):
    [
        Vol_EC_consumed, 
        Vol_Elely_need, 
        Test_V, 
        Test_V2, 
        Vol_Elely_add, 
        Vol_Elely_Tot_new, 
        Vol_Elely_JR_new, 
        Vol_Pore_tot_new, 
        Vol_Pore_decrease, 
        c_e_r_new, c_EC_r_new,
        Ratio_Dryout, Ratio_CeEC_JR, 
        Ratio_CeLi_JR,
        Width_new, ]= Data_Pack;
    mdic_dry["Vol_Elely_Tot_All"].append(Vol_Elely_Tot_new*1e6);            
    mdic_dry["Vol_Elely_JR_All"].append(Vol_Elely_JR_new*1e6);     
    mdic_dry["Vol_Pore_tot_All"].append(Vol_Pore_tot_new*1e6);           
    mdic_dry["Ratio_CeEC_All"].append(Ratio_CeEC_JR);                      
    mdic_dry["Ratio_CeLi_All"].append(Ratio_CeLi_JR);             
    mdic_dry["Ratio_Dryout_All"].append(Ratio_Dryout);
    mdic_dry["Vol_EC_consumed_All"].append(Vol_EC_consumed*1e6);        
    mdic_dry["Vol_Elely_need_All"].append(Vol_Elely_need*1e6);     
    mdic_dry["Width_all"].append(Width_new);
    mdic_dry["Vol_Elely_add_All"].append(Vol_Elely_add*1e6);            
    mdic_dry["Vol_Pore_decrease_All"].append(Vol_Pore_decrease*1e6);
    mdic_dry["Test_V_All"].append(Test_V*1e6); 
    mdic_dry["Test_V2_All"].append(Test_V2*1e6); 
    mdic_dry["c_e_r_new_All"].append(c_e_r_new); 
    mdic_dry["c_EC_r_new_All"].append(c_EC_r_new)

    return mdic_dry

def Get_Values_Excel(
    model_options,my_dict_RPT,mdic_dry,
    DryOut,Scan_i,Para_dict_i,str_exp_AGE_text,
    str_exp_RPT_text,str_error_AGE_final,
    str_error_RPT):
    if model_options.__contains__("SEI on cracks"):
            LossCap_seioncrack = my_dict_RPT["CDend Loss of capacity to SEI on cracks [A.h]"][-1]
    else:
        LossCap_seioncrack = "nan"
    if model_options.__contains__("lithium plating"):
        LossCap_LiP = my_dict_RPT["CDend Loss of capacity to lithium plating [A.h]"][-1]
    else:
        LossCap_LiP = "nan"
    if model_options.__contains__("SEI"):
        LossCap_SEI = my_dict_RPT["CDend Loss of capacity to SEI [A.h]"][-1]
    else:
        LossCap_SEI = "nan"
    if DryOut == "On":
        Vol_Elely_Tot_All_final = mdic_dry["Vol_Elely_Tot_All"][-1];
        Vol_Elely_JR_All_final  = mdic_dry["Vol_Elely_JR_All"][-1];
        Width_all_final         = mdic_dry["Width_all"][-1];
    else:
        Vol_Elely_Tot_All_final = "nan"
        Vol_Elely_JR_All_final  = "nan"
        Width_all_final         = "nan"
    value_list_temp = list(Para_dict_i.values())
    values = []
    for value_list_temp_i in value_list_temp:
        values.append(str(value_list_temp_i))
    values.insert(0,str(Scan_i));
    values.insert(1,DryOut);
    values.extend([
        str_exp_AGE_text,
        str_exp_RPT_text,
        str(my_dict_RPT["Discharge capacity [A.h]"][0] 
        - 
        my_dict_RPT["Discharge capacity [A.h]"][-1]),

        str(LossCap_LiP),
        str(LossCap_SEI),
        str(LossCap_seioncrack),

        str(my_dict_RPT["CDend Negative electrode capacity [A.h]"][0] 
        - 
        my_dict_RPT["CDend Negative electrode capacity [A.h]"][-1]),

        str(my_dict_RPT["CDend Positive electrode capacity [A.h]"][0] 
        - 
        my_dict_RPT["CDend Positive electrode capacity [A.h]"][-1]),

        str(Vol_Elely_Tot_All_final), 
        str(Vol_Elely_JR_All_final),
        str(Width_all_final),
        str_error_AGE_final, # Mark Ruihe
        str_error_RPT   # Mark Ruihe
        ])
    return values        

def Plot_Cyc_RPT_4(my_dict_RPT,Niall_data,Scan_i,Temper_i,model_options,BasicPath, Target,fs,dpi):
    Num_subplot = 3;
    fig, axs = plt.subplots(1,Num_subplot, figsize=(18,4.8),tight_layout=True)
    axs[0].plot(
        my_dict_RPT["Cycle_RPT"], 
        my_dict_RPT["Discharge capacity [A.h]"],     
        '-o', label="Scan=" + str(Scan_i) )
    if Temper_i  == 40 + 273.15:       
        axs[0].plot(Niall_data['F_Cap_all'][:,2],Niall_data['F_Cap_all'][:,5]/1e3, '-^',  label='Cell F' )
        axs[0].plot(Niall_data['G_Cap_all'][:,2],Niall_data['G_Cap_all'][:,5]/1e3, '-^',  label='Cell G' )
        axs[0].plot(Niall_data['H_Cap_all'][:,2],Niall_data['H_Cap_all'][:,5]/1e3, '-^',  label='Cell H' )
    elif Temper_i  == 25 + 273.15: 
        axs[0].plot(Niall_data['D_Cap_all'][:,2],Niall_data['D_Cap_all'][:,5]/1e3, '-^',  label='Cell D' )
        axs[0].plot(Niall_data['E_Cap_all'][:,2],Niall_data['E_Cap_all'][:,5]/1e3, '-^',  label='Cell E' )
    elif Temper_i  == 10 + 273.15: 
        axs[0].plot(Niall_data['A_Cap_all'][:,2],Niall_data['A_Cap_all'][:,5]/1e3, '-^',  label='Cell A' )
        axs[0].plot(Niall_data['B_Cap_all'][:,2],Niall_data['B_Cap_all'][:,5]/1e3, '-^',  label='Cell B' )
        axs[0].plot(Niall_data['C_Cap_all'][:,2],Niall_data['C_Cap_all'][:,5]/1e3, '-^',  label='Cell C' )
    else:
        pass
    if model_options.__contains__("lithium plating"):
        axs[1].plot(my_dict_RPT["Cycle_RPT"], my_dict_RPT["CDend Loss of capacity to lithium plating [A.h]"],'-o', label="LiP - Scan=" + str(Scan_i) )
    if model_options.__contains__("SEI"):
        axs[1].plot(my_dict_RPT["Cycle_RPT"], my_dict_RPT["CDend Loss of capacity to SEI [A.h]"] ,'-o', label="SEI - Scan" + str(Scan_i) )
    if model_options.__contains__("SEI on cracks"):
        axs[1].plot(my_dict_RPT["Cycle_RPT"], my_dict_RPT["CDend Loss of capacity to SEI on cracks [A.h]"] ,'-o', label="sei-on-cracks - Scan" + str(Scan_i) )
    axs[2].plot(
        my_dict_RPT["CDend Throughput capacity [A.h]"], 
        my_dict_RPT["Discharge capacity [A.h]"],     
        '-o', label="Scan=" + str(Scan_i) )
    if Temper_i  == 40:       
        axs[2].plot(Niall_data['F_Cap_all'][:,3],Niall_data['F_Cap_all'][:,5]/1e3, '-^',  label='Cell F' )
        axs[2].plot(Niall_data['G_Cap_all'][:,3],Niall_data['G_Cap_all'][:,5]/1e3, '-^',  label='Cell G' )
        axs[2].plot(Niall_data['H_Cap_all'][:,3],Niall_data['H_Cap_all'][:,5]/1e3, '-^',  label='Cell H' )
    elif Temper_i  == 25: 
        axs[2].plot(Niall_data['D_Cap_all'][:,3],Niall_data['D_Cap_all'][:,5]/1e3, '-^',  label='Cell D' )
        axs[2].plot(Niall_data['E_Cap_all'][:,3],Niall_data['E_Cap_all'][:,5]/1e3, '-^',  label='Cell E' )
    elif Temper_i  == 10: 
        axs[2].plot(Niall_data['A_Cap_all'][:,3],Niall_data['A_Cap_all'][:,5]/1e3, '-^',  label='Cell A' )
        axs[2].plot(Niall_data['B_Cap_all'][:,3],Niall_data['B_Cap_all'][:,5]/1e3, '-^',  label='Cell B' )
        axs[2].plot(Niall_data['C_Cap_all'][:,3],Niall_data['C_Cap_all'][:,5]/1e3, '-^',  label='Cell C' )
    else:
        pass
    for i in range(0,Num_subplot):
        axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[i].set_ylabel("Capacity [A.h]",   fontdict={'family':'DejaVu Sans','size':fs})
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
        axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
        axs[i].set_title("Discharge capacity",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[2].set_xlabel("Throughput capacity [A.h]",   fontdict={'family':'DejaVu Sans','size':fs})   
    axs[1].set_title("Capacity loss to LiP and SEI",   fontdict={'family':'DejaVu Sans','size':fs+1})
    plt.savefig(BasicPath + Target+ str(Scan_i)+"/Cap-LLI.png", dpi=dpi)

    if model_options.__contains__("SEI on cracks"):
        fs=17;Num_subplot = 2;
        fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
        axs[0].plot(my_dict_RPT["Cycle_RPT"], my_dict_RPT["CDend X-averaged total SEI on cracks thickness [m]"],     '-o', label="Scan=" + str(Scan_i) )
        axs[1].plot(my_dict_RPT["Cycle_RPT"], my_dict_RPT["CDend X-averaged negative electrode roughness ratio"],'-o', label="Scan=" + str(Scan_i) )
        axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[1].set_ylabel("Roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs})
        for i in range(0,Num_subplot):
            axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
            labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
            axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
            axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
        axs[0].set_title("X-avg tot Neg SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
        axs[1].set_title("X-avg Neg roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
        plt.savefig(BasicPath + Target+ str(Scan_i)+"/Cracks related.png", dpi=dpi)

        Num_subplot = 2;
        fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
        axs[1].plot(my_dict_RPT["Cycle_RPT"], 
            my_dict_RPT["CDend Negative electrode capacity [A.h]"][0]
            -
            my_dict_RPT["CDend Negative electrode capacity [A.h]"],'-o',label="Neg Scan=" + str(Scan_i))
        axs[1].plot(my_dict_RPT["Cycle_RPT"], 
            my_dict_RPT["CDend Positive electrode capacity [A.h]"][0]
            -
            my_dict_RPT["CDend Positive electrode capacity [A.h]"],'-^',label="Pos Scan=" + str(Scan_i))
        axs[0].plot(my_dict_RPT["Cycle_RPT"], my_dict_RPT["CDend X-averaged total SEI on cracks thickness [m]"],                  '-o',label="Scan="+ str(Scan_i))
        for i in range(0,1):
            axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
            labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
            axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
            axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
        axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[0].set_title("CDend X-avg tot SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
        for i in range(1,2):
            axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[i].set_ylabel("Capacity [A.h]",   fontdict={'family':'DejaVu Sans','size':fs})
            labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
            axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
            axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
        axs[1].set_title("LAM of Neg and Pos",   fontdict={'family':'DejaVu Sans','size':fs+1})
        plt.savefig(BasicPath + Target+ str(Scan_i)+"/LAM-IR.png", dpi=dpi)

    Num_subplot = 2;
    fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
    axs[0].plot(my_dict_RPT["Cycle_RPT"], my_dict_RPT["CDsta Positive electrode SOC"] ,'-o',label="Start" )
    axs[0].plot(my_dict_RPT["Cycle_RPT"], my_dict_RPT["CDend Positive electrode SOC"] ,'-^',label="End" )
    axs[1].plot(my_dict_RPT["Cycle_RPT"], my_dict_RPT["CDsta Negative electrode SOC"],'-o',label="Start" )
    axs[1].plot(my_dict_RPT["Cycle_RPT"], my_dict_RPT["CDend Negative electrode SOC"],'-^',label="End" )
    for i in range(0,2):
        axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[i].set_ylabel("SOC",   fontdict={'family':'DejaVu Sans','size':fs})
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
        axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)     
    axs[0].set_title("Neg SOC range (Dis)",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[1].set_title("Pos SOC range (Dis)",   fontdict={'family':'DejaVu Sans','size':fs+1})
    plt.savefig(BasicPath + Target+ str(Scan_i)+"/SOC_RPT_dis.png", dpi=dpi)

    return

def Plot_Loc_AGE_4(my_dict_AGE,Scan_i,model_options,BasicPath, Target,fs,dpi):
    Num_subplot = 2;
    fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Porosity"][0],'-o',label="First")
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Porosity"][-1],'-^',label="Last"  )
    axs[1].plot(
        my_dict_AGE["x_n [m]"],
        my_dict_AGE["CDend Negative electrode reaction overpotential [V]"][0],'-o',label="First" )
    axs[1].plot(
        my_dict_AGE["x_n [m]"],
        my_dict_AGE["CDend Negative electrode reaction overpotential [V]"][-1],'-^',label="Last" )

    axs[0].set_xlabel("Dimensional Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[1].set_xlabel("Dimensional Neg thickness",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[0].set_title("Porosity",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[1].set_title("Neg electrode reaction overpotential",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[0].set_ylabel("Porosity",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[1].set_ylabel("Overpotential [V]",   fontdict={'family':'DejaVu Sans','size':fs})
    for i in range(0,2):
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
        axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
    plt.savefig(BasicPath + Target+ str(Scan_i)+"/Por Neg_S_eta.png", dpi=dpi)

    if model_options.__contains__("SEI on cracks"):
        Num_subplot = 2;
        fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
        axs[0].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Negative electrode roughness ratio"][0],'-o',label="First")
        axs[0].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Negative electrode roughness ratio"][-1],'-^',label="Last"  )
        axs[1].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Total SEI on cracks thickness [m]"][0],'-o',label="First" )
        axs[1].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Total SEI on cracks thickness [m]"][-1],'-^',label="Last" )
        axs[0].set_title("Tot Neg SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
        axs[1].set_title("Neg roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
        axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[1].set_ylabel("Roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs})
        for i in range(0,2):
            axs[i].set_xlabel("Dimensional Neg thickness",   fontdict={'family':'DejaVu Sans','size':fs})
            labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
            axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
            axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
        plt.savefig(BasicPath + Target+ str(Scan_i)+"/Cracks related spatial.png", dpi=dpi)

    Num_subplot = 2;
    fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte concentration [mol.m-3]"][0],'-o',label="First")
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte concentration [mol.m-3]"][-1],'-^',label="Last"  )
    axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte potential [V]"][0],'-o',label="First" )
    axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte potential [V]"][-1],'-^',label="Last" )
    axs[0].set_title("Electrolyte concentration",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[1].set_title("Electrolyte potential",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[0].set_ylabel("Concentration [mol.m-3]",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[1].set_ylabel("Potential [V]",   fontdict={'family':'DejaVu Sans','size':fs})
    for i in range(0,2):
        axs[i].set_xlabel("Dimensional Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
        axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
    plt.savefig(BasicPath + Target+ str(Scan_i)+"/Electrolyte concentration and potential.png", dpi=dpi)
    
    Num_subplot = 2;
    fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte diffusivity [m2.s-1]"][0],'-o',label="First")
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte diffusivity [m2.s-1]"][-1],'-^',label="Last"  )
    axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte conductivity [S.m-1]"][0],'-o',label="First" )
    axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte conductivity [S.m-1]"][-1],'-^',label="Last" )
    axs[0].set_title("Electrolyte diffusivity",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[1].set_title("Electrolyte conductivity",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[0].set_ylabel("Diffusivity [m2.s-1]",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[1].set_ylabel("Conductivity [S.m-1]",   fontdict={'family':'DejaVu Sans','size':fs})
    for i in range(0,2):
        axs[i].set_xlabel("Dimensional Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
        axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
    plt.savefig(BasicPath + Target+ str(Scan_i)+"/Electrolyte diffusivity and conductivity.png", dpi=dpi)

    return

def Plot_Dryout(
    Cyc_Update_Index,mdic_dry,ce_EC_0,
    Scan_i,BasicPath, Target,fs,dpi):
    
    CeEC_All =np.full(np.size(mdic_dry["Ratio_CeEC_All"]),ce_EC_0); 
    for i in range(1,np.size(mdic_dry["Ratio_CeEC_All"])):
        for k in range(0,i):
            CeEC_All[i] *= mdic_dry["Ratio_CeEC_All"][k]
    mdic_dry["CeEC_All"]= CeEC_All

    Num_subplot = 3;
    fig, axs = plt.subplots(1,Num_subplot, figsize=(18,4.8),tight_layout=True)
    axs[0].plot(Cyc_Update_Index, mdic_dry["Vol_EC_consumed_All"],'-o',label="EC consumed")
    axs[0].plot(Cyc_Update_Index, mdic_dry["Vol_Elely_need_All"],'-.',label="Elely needed")
    axs[0].plot(Cyc_Update_Index, mdic_dry["Vol_Elely_add_All"],'-s',label="Elely added")
    axs[0].plot(Cyc_Update_Index, mdic_dry["Vol_Pore_decrease_All"],'--',label="Pore decreased")
    axs[1].plot(Cyc_Update_Index, mdic_dry["Vol_Elely_Tot_All"],     '-o', label="Total electrolyte in cell" )
    axs[1].plot(Cyc_Update_Index, mdic_dry["Vol_Elely_JR_All"],     '--', label="Total electrolyte in JR" )
    axs[1].plot(Cyc_Update_Index, mdic_dry["Vol_Pore_tot_All"],     '-s', label="Total pore in JR" )
    axs[2].plot(Cyc_Update_Index, mdic_dry["Ratio_Dryout_All"],     '-s', label="Dry out ratio" )
    axs[2].set_ylabel("Ratio",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[2].set_xlabel("Cycle number",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[2].set_title("Dry out ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
    for i in range(0,2):
        axs[i].set_xlabel("Cycle number",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[i].set_ylabel("Volume [mL]",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[i].set_title("Volume",   fontdict={'family':'DejaVu Sans','size':fs+1})
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
        axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
    plt.savefig(BasicPath + Target+ str(Scan_i)+"/Volume_total.png", dpi=dpi)
    
    return


# Add run till fail functions from P3
def Run_P2_till_Fail(
    index_xlsx, Para_dict_i,   Path_pack , 
    keys_all,   exp_text_list, exp_index_pack , Niall_data ):

    ##########################################################
    ##############    Part-1: Initialization    ##############
    ##########################################################
    ModelTimer = pb.Timer()
    Scan_i = int(index_xlsx)
    print('Start Now! Scan %d.' % Scan_i)  
    # pb.set_logging_level('INFO') # show more information!
    # set_start_method('fork') # from Patrick

    # Un-pack data:
    [cycle_no,step_AGE_CD,step_AGE_CC,step_AGE_CV,
        step_RPT_CD,step_RPT_RE , step_RPT_CC ] = exp_index_pack;
    [exp_AGE_text, exp_RPT_text,] = exp_text_list;
    [BasicPath,Target,book_name_xlsx,sheet_name_xlsx,] = Path_pack
    CyclePack,Para_0 = Para_init(Para_dict_i)
    [Total_Cycles,Cycle_bt_RPT,Update_Cycles,RPT_Cycles,
        Temper_i,Temper_RPT,mesh_par,submesh_strech,model_options] = CyclePack;
    [keys_all_RPT,keys_all_AGE] = keys_all
    str_exp_AGE_text  = str(exp_AGE_text);
    str_exp_RPT_text  = str(exp_RPT_text);

    # define experiment
    Experiment_Long   = pb.Experiment( exp_AGE_text * Update_Cycles  )  
    Experiment_RPT    = pb.Experiment( exp_RPT_text * RPT_Cycles     ) 
    Experiment_Breakin= pb.Experiment( exp_RPT_text * RPT_Cycles     )

    #####  index definition ######################
    Small_Loop =  int(Cycle_bt_RPT/Update_Cycles);   
    SaveTimes = int(Total_Cycles/Cycle_bt_RPT);   

    # initialize my_dict for outputs
    my_dict_RPT = {}
    for keys in keys_all_RPT:
        for key in keys:
            my_dict_RPT[key]=[];
    my_dict_AGE = {}; 
    for keys in keys_all_AGE:
        for key in keys:
            my_dict_AGE[key]=[];
    my_dict_RPT["Cycle_RPT"] = []; my_dict_AGE["Cycle_AGE"] = []; 
    Cyc_Update_Index     =[]; 
            
    # update 220924: merge DryOut and Int_ElelyExces_Ratio
    temp_Int_ElelyExces_Ratio =  Para_0["Initial electrolyte excessive amount ratio"] 
    ce_EC_0 = Para_0['EC initial concentration in electrolyte [mol.m-3]'] # used to calculate ce_EC_All
    if temp_Int_ElelyExces_Ratio < 1:
        Int_ElelyExces_Ratio = -1;
        DryOut = "Off";
    else:
        Int_ElelyExces_Ratio = temp_Int_ElelyExces_Ratio;
        DryOut = "On";
    print(f"Scan {Scan_i}: DryOut = {DryOut}")
    if DryOut == "On":  
        mdic_dry,Para_0 = Initialize_mdic_dry(Para_0,Int_ElelyExces_Ratio)
    else:
        mdic_dry ={}

    ##########################################################
    ##############    Part-2: Run model         ##############
    ##########################################################
    ##########################################################
    Timeout_text = 'I timed out'
    ##########    2-1: Define model and run break-in cycle
    try:  
        Timelimit = int(3600*0.5)
        # the following turns on for HPC only!
        timeout_RPT = TimeoutFunc(
            Run_Breakin, 
            timeout=Timelimit, 
            timeout_val=Timeout_text)
        Result_list_breakin  = timeout_RPT(
            model_options, Experiment_Breakin, 
            Para_0, mesh_par, submesh_strech)
        [Model_0,Sol_0,Call_Breakin] = Result_list_breakin
        if Call_Breakin.success == False:
            print("Fail due to Experiment error or infeasible")
            1/0
        if Sol_0 == Timeout_text: # to do: distinguish different failure cases
            print("Fail due to Timeout")
            1/0
        if Sol_0 == "Model error or solver error":
            print("Fail due to Model error or solver error")
            1/0
    except ZeroDivisionError as e:
        str_error_Breakin = str(e)
        print(f"Scan {Scan_i}: Fail break-in cycle, need to exit the whole scan now due to {str_error_Breakin} but do not know how!")
        
        Flag_Breakin = False
    else:
        print(f"Scan {Scan_i}: Finish break-in cycle")
        # post-process for break-in cycle
        my_dict_RPT = GetSol_dict (my_dict_RPT,keys_all_RPT, Sol_0, 
            cycle_no, step_RPT_CD , step_RPT_CC , step_RPT_RE, step_AGE_CV   )
        cycle_count =0; 
        my_dict_RPT["Cycle_RPT"].append(cycle_count)
        Cyc_Update_Index.append(cycle_count);
        
        Flag_Breakin = True

    Flag_AGE = True; str_error_AGE_final = "Empty";   str_error_RPT = "Empty";
    #############################################################
    #######   2-2: Write a big loop to finish the long experiment    
    if Flag_Breakin == True: 
        k=0
        # Para_All.append(Para_0);Model_All.append(Model_0);Sol_All_i.append(Sol_0); 
        Para_0_Dry_old = Para_0;     Model_Dry_old = Model_0  ; Sol_Dry_old = Sol_0;   del Model_0,Sol_0
        while k < SaveTimes:    
            i=0    
            while i < Small_Loop:
                if DryOut == "On":
                    Data_Pack,Paraupdate   = Cal_new_con_Update (  Sol_Dry_old,   Para_0_Dry_old )
                if DryOut == "Off":
                    Paraupdate = Para_0
                # Run aging cycle:
                try:
                    Timelimit = int(3600*0.5)
                    timeout_AGE = TimeoutFunc(
                        Run_Model_Base_On_Last_Solution, 
                        timeout=Timelimit, 
                        timeout_val=Timeout_text)
                    Result_list_AGE = timeout_AGE( 
                        Model_Dry_old  , Sol_Dry_old , Paraupdate ,Experiment_Long, 
                        Update_Cycles,Temper_i,mesh_par,submesh_strech )
                    [Model_Dry_i, Sol_Dry_i , Call_Age ] = Result_list_AGE
                    #print(f"Temperature for ageing is now: {Temper_i}")  
                    if Call_Age.success == False:
                        print("Fail due to Experiment error or infeasible")
                        str_error_AGE = "Experiment error or infeasible"
                        1/0
                    if Sol_Dry_i == Timeout_text: # fail due to timeout
                        print("Fail due to Timeout")
                        str_error_AGE = "Timeout"
                        1/0
                    if Sol_Dry_i == "Model error or solver error":
                        print("Fail due to Model error or solver error")
                        str_error_AGE = "Model error or solver error"
                        1/0
                except ZeroDivisionError as e:
                    print(f"Scan {Scan_i}: Fail during No.{Cyc_Update_Index[-1]} ageing cycles due to {str_error_AGE}")
                    Flag_AGE = False
                    str_error_AGE_final = str_error_AGE
                    break
                else:
                    Para_0_Dry_old = Paraupdate;       Model_Dry_old = Model_Dry_i;      Sol_Dry_old = Sol_Dry_i;   
                    del Paraupdate,Model_Dry_i,Sol_Dry_i
                    # post-process for first ageing cycle and every -1 ageing cycle
                    if k==0 and i==0:    
                        my_dict_AGE = GetSol_dict (my_dict_AGE,keys_all_AGE, Sol_Dry_old, 
                            0, step_AGE_CD , step_AGE_CC , step_RPT_RE, step_AGE_CV   )     
                        my_dict_AGE["Cycle_AGE"].append(1)
                    my_dict_AGE = GetSol_dict (my_dict_AGE,keys_all_AGE, Sol_Dry_old, 
                        cycle_no, step_AGE_CD , step_AGE_CC , step_RPT_RE, step_AGE_CV   )    
                    cycle_count +=  Update_Cycles; 
                    my_dict_AGE["Cycle_AGE"].append(cycle_count)           
                    Cyc_Update_Index.append(cycle_count)
                    print(f"Scan {Scan_i}: Finish for No.{Cyc_Update_Index[-1]} ageing cycles")
                    if DryOut == "On":
                        mdic_dry = Update_mdic_dry(Data_Pack,mdic_dry)
                    i += 1;   
            # run RPT, and also update parameters (otherwise will have problems)
            if DryOut == "On":
                Data_Pack , Paraupdate  = Cal_new_con_Update (  Sol_Dry_old,   Para_0_Dry_old   )
            if DryOut == "Off":
                Paraupdate = Para_0     
            try:
                Timelimit = int(3600*0.5)
                timeout_RPT = TimeoutFunc(
                    Run_Model_Base_On_Last_Solution_RPT, 
                    timeout=Timelimit, 
                    timeout_val=Timeout_text)
                Result_list_RPT = timeout_RPT(
                    Model_Dry_old  , Sol_Dry_old ,   
                    Paraupdate,      Experiment_RPT, RPT_Cycles, 
                    Temper_RPT ,mesh_par ,submesh_strech
                )
                [Model_Dry_i, Sol_Dry_i,Call_RPT]  = Result_list_RPT
                #print(f"Temperature for RPT is now: {Temper_RPT}")  
                if Call_RPT.success == False:
                    print("Fail due to Experiment error or infeasible")
                    str_error_RPT = "Experiment error or infeasible"
                    1/0 
                if Sol_Dry_i == Timeout_text:
                    print("Fail due to Timeout")
                    str_error_RPT = "Timeout"
                    1/0
                if Sol_Dry_i == "Model error or solver error":
                    print("Fail due to Model error or solver error")
                    str_error_RPT = "Model error or solver error"
                    1/0
            except ZeroDivisionError as e:
                print(f"Scan {Scan_i}: Fail during No.{Cyc_Update_Index[-1]} RPT cycles, due to {str_error_RPT}")
                break
            else:
                my_dict_RPT = GetSol_dict (my_dict_RPT,keys_all_RPT, Sol_Dry_i, 
                    cycle_no, step_RPT_CD , step_RPT_CC , step_RPT_RE, step_AGE_CV   )
                my_dict_RPT["Cycle_RPT"].append(cycle_count)
                Cyc_Update_Index.append(cycle_count)
                print(f"Scan {Scan_i}: Finish for No.{Cyc_Update_Index[-1]} RPT cycles")
                if DryOut == "On":
                    mdic_dry = Update_mdic_dry(Data_Pack,mdic_dry)
                Para_0_Dry_old = Paraupdate;    Model_Dry_old = Model_Dry_i  ;     Sol_Dry_old = Sol_Dry_i    ;   
                del Paraupdate,Model_Dry_i,Sol_Dry_i
                if Flag_AGE == False:
                    break
            k += 1 
    ############################################################# 
    #########   An extremely bad case: cannot even finish breakin
    if Flag_Breakin == False: 
        value_list_temp = list(Para_dict_i.values())
        values = []
        for value_list_temp_i in value_list_temp:
            values.append(str(value_list_temp_i))
        values.insert(0,str(Scan_i));
        values.insert(1,DryOut);
        values.extend([
            str_exp_AGE_text,
            str_exp_RPT_text,
            "nan","nan",
            "nan","nan", 
            "nan","nan",
            "nan","nan",
            "nan",str_error_Breakin])
        values = [values,]
        print(str_error_Breakin)
        print("Fail in {}".format(ModelTimer.time())) 
        book_name_xlsx_seperate =   str(Scan_i)+ '_' + book_name_xlsx;
        sheet_name_xlsx =  str(Scan_i);
        write_excel_xlsx(
            BasicPath + Target+book_name_xlsx_seperate, 
            sheet_name_xlsx, values)
    ##########################################################
    ##############   Part-3: Post-prosessing    ##############
    ##########################################################
    # Newly add (220517): save plots, not just a single line in excel file:     
    # Newly add (221114): make plotting as functions
    # Niall_data = loadmat( 'Extracted_all_cell.mat'
    else:
        if not os.path.exists(BasicPath + Target + str(Scan_i)):
            os.mkdir(BasicPath + Target + str(Scan_i) );
        dpi= 100; fs = 17;
        ##########################################################
        #########      3-1: Plot cycle,location, Dryout related 
        Plot_Cyc_RPT_4(my_dict_RPT,Niall_data,Scan_i,Temper_i,model_options,BasicPath, Target,fs,dpi)
        if len(my_dict_AGE["CDend Porosity"])>1:
            Plot_Loc_AGE_4(my_dict_AGE,Scan_i,model_options,BasicPath, Target,fs,dpi)
        if DryOut == "On":
            Plot_Dryout(Cyc_Update_Index,mdic_dry,ce_EC_0,Scan_i,BasicPath, Target,fs,dpi)
        ##########################################################
        #########      3-2: Save data as .mat 
        my_dict_RPT["Cyc_Update_Index"] = Cyc_Update_Index
        my_dict_RPT["SaveTimes"]    = SaveTimes
        midc_merge = {**my_dict_RPT, **my_dict_AGE,**mdic_dry}
        savemat(BasicPath + Target+ str(Scan_i) + '/' + str(Scan_i)+ '-StructDara_for_Mat.mat',midc_merge)  
        ##########################################################
        #########      3-3: Save summary to excel 
        values=Get_Values_Excel(
            model_options,my_dict_RPT,mdic_dry,
            DryOut,Scan_i,Para_dict_i,str_exp_AGE_text,
            str_exp_RPT_text,
            str_error_AGE_final,
            str_error_RPT)
        values = [values,]
        book_name_xlsx_seperate =   str(Scan_i)+ '_' + book_name_xlsx;
        sheet_name_xlsx =  str(Scan_i);
        write_excel_xlsx(
            BasicPath + Target+book_name_xlsx_seperate, 
            sheet_name_xlsx, values)
        print("Succeed doing something in {}".format(ModelTimer.time()))
        print('This is the end of No.', Scan_i, ' scan')

def Run_P2_Opt_Timeout(
    index_xlsx, Para_dict_i,   Path_pack , 
    keys_all,   exp_text_list, exp_index_pack , Niall_data, Timeout,Return_Sol ):

    # add 221205: if Timeout=='True', use Patrick's version, disable pool
    #             else, use pool to accelerate 
    #  add Return_Sol, on HPC, always set to False, as it is useless, 

    ##########################################################
    ##############    Part-1: Initialization    ##############
    ##########################################################
    ModelTimer = pb.Timer()
    Scan_i = int(index_xlsx)
    print('Start Now! Scan %d.' % Scan_i)  
    Sol_RPT = [];  Sol_AGE = [];
    # pb.set_logging_level('INFO') # show more information!
    # set_start_method('fork') # from Patrick

    # Un-pack data:
    [cycle_no,step_AGE_CD,step_AGE_CC,step_AGE_CV,
        step_RPT_CD,step_RPT_RE , step_RPT_CC ] = exp_index_pack;
    [exp_AGE_text, exp_RPT_text,] = exp_text_list;
    [BasicPath,Target,book_name_xlsx,sheet_name_xlsx,] = Path_pack
    CyclePack,Para_0 = Para_init(Para_dict_i)
    [Total_Cycles,Cycle_bt_RPT,Update_Cycles,RPT_Cycles,
        Temper_i,Temper_RPT,mesh_par,submesh_strech,model_options] = CyclePack;
    [keys_all_RPT,keys_all_AGE] = keys_all
    str_exp_AGE_text  = str(exp_AGE_text);
    str_exp_RPT_text  = str(exp_RPT_text);

    # define experiment
    Experiment_Long   = pb.Experiment( exp_AGE_text * Update_Cycles  )  
    Experiment_RPT    = pb.Experiment( exp_RPT_text * RPT_Cycles     ) 
    Experiment_Breakin= pb.Experiment( exp_RPT_text * RPT_Cycles     )

    #####  index definition ######################
    Small_Loop =  int(Cycle_bt_RPT/Update_Cycles);   
    SaveTimes = int(Total_Cycles/Cycle_bt_RPT);   

    # initialize my_dict for outputs
    my_dict_RPT = {}
    for keys in keys_all_RPT:
        for key in keys:
            my_dict_RPT[key]=[];
    my_dict_AGE = {}; 
    for keys in keys_all_AGE:
        for key in keys:
            my_dict_AGE[key]=[];
    my_dict_RPT["Cycle_RPT"] = []; my_dict_AGE["Cycle_AGE"] = []; 
    Cyc_Update_Index     =[]; 
            
    # update 220924: merge DryOut and Int_ElelyExces_Ratio
    temp_Int_ElelyExces_Ratio =  Para_0["Initial electrolyte excessive amount ratio"] 
    ce_EC_0 = Para_0['EC initial concentration in electrolyte [mol.m-3]'] # used to calculate ce_EC_All
    if temp_Int_ElelyExces_Ratio < 1:
        Int_ElelyExces_Ratio = -1;
        DryOut = "Off";
    else:
        Int_ElelyExces_Ratio = temp_Int_ElelyExces_Ratio;
        DryOut = "On";
    print(f"Scan {Scan_i}: DryOut = {DryOut}")
    if DryOut == "On":  
        mdic_dry,Para_0 = Initialize_mdic_dry(Para_0,Int_ElelyExces_Ratio)
    else:
        mdic_dry ={}

    ##########################################################
    ##############    Part-2: Run model         ##############
    ##########################################################
    ##########################################################
    Timeout_text = 'I timed out'
    ##########    2-1: Define model and run break-in cycle
    try:  
        Timelimit = int(3600*0.5)
        # the following turns on for HPC only!
        if Timeout == True:
            timeout_RPT = TimeoutFunc(
                Run_Breakin, 
                timeout=Timelimit, 
                timeout_val=Timeout_text)
            Result_list_breakin  = timeout_RPT(
                model_options, Experiment_Breakin, 
                Para_0, mesh_par, submesh_strech)
        else:
            Result_list_breakin  = Run_Breakin(
                model_options, Experiment_Breakin, 
                Para_0, mesh_par, submesh_strech)
        [Model_0,Sol_0,Call_Breakin] = Result_list_breakin
        if Return_Sol == True:
            Sol_RPT.append(Sol_0)
        if Call_Breakin.success == False:
            print("Fail due to Experiment error or infeasible")
            1/0
        if Sol_0 == Timeout_text: # to do: distinguish different failure cases
            print("Fail due to Timeout")
            1/0
        if Sol_0 == "Model error or solver error":
            print("Fail due to Model error or solver error")
            1/0
    except ZeroDivisionError as e:
        str_error_Breakin = str(e)
        print(f"Scan {Scan_i}: Fail break-in cycle, need to exit the whole scan now due to {str_error_Breakin} but do not know how!")
        
        Flag_Breakin = False
    else:
        print(f"Scan {Scan_i}: Finish break-in cycle")
        # post-process for break-in cycle
        my_dict_RPT = GetSol_dict (my_dict_RPT,keys_all_RPT, Sol_0, 
            cycle_no, step_RPT_CD , step_RPT_CC , step_RPT_RE, step_AGE_CV   )
        cycle_count =0; 
        my_dict_RPT["Cycle_RPT"].append(cycle_count)
        Cyc_Update_Index.append(cycle_count);
        
        Flag_Breakin = True
        

    Flag_AGE = True; str_error_AGE_final = "Empty";   str_error_RPT = "Empty";
    #############################################################
    #######   2-2: Write a big loop to finish the long experiment    
    if Flag_Breakin == True: 
        k=0
        # Para_All.append(Para_0);Model_All.append(Model_0);Sol_All_i.append(Sol_0); 
        Para_0_Dry_old = Para_0;     Model_Dry_old = Model_0  ; Sol_Dry_old = Sol_0;   del Model_0,Sol_0
        while k < SaveTimes:    
            i=0    
            while i < Small_Loop:
                if DryOut == "On":
                    Data_Pack,Paraupdate   = Cal_new_con_Update (  Sol_Dry_old,   Para_0_Dry_old )
                if DryOut == "Off":
                    Paraupdate = Para_0
                # Run aging cycle:
                try:
                    Timelimit = int(3600*0.5)
                    if Timeout == True:
                        timeout_AGE = TimeoutFunc(
                            Run_Model_Base_On_Last_Solution, 
                            timeout=Timelimit, 
                            timeout_val=Timeout_text)
                        Result_list_AGE = timeout_AGE( 
                            Model_Dry_old  , Sol_Dry_old , Paraupdate ,Experiment_Long, 
                            Update_Cycles,Temper_i,mesh_par,submesh_strech )
                    else:
                        Result_list_AGE = Run_Model_Base_On_Last_Solution( 
                            Model_Dry_old  , Sol_Dry_old , Paraupdate ,Experiment_Long, 
                            Update_Cycles,Temper_i,mesh_par,submesh_strech )
                    [Model_Dry_i, Sol_Dry_i , Call_Age ] = Result_list_AGE
                    if Return_Sol == True:
                        Sol_AGE.append(Sol_Dry_i)
                    #print(f"Temperature for ageing is now: {Temper_i}")  
                    if Call_Age.success == False:
                        print("Fail due to Experiment error or infeasible")
                        str_error_AGE = "Experiment error or infeasible"
                        1/0
                    if Sol_Dry_i == Timeout_text: # fail due to timeout
                        print("Fail due to Timeout")
                        str_error_AGE = "Timeout"
                        1/0
                    if Sol_Dry_i == "Model error or solver error":
                        print("Fail due to Model error or solver error")
                        str_error_AGE = "Model error or solver error"
                        1/0
                except ZeroDivisionError as e:
                    print(f"Scan {Scan_i}: Fail during No.{Cyc_Update_Index[-1]} ageing cycles due to {str_error_AGE}")
                    Flag_AGE = False
                    str_error_AGE_final = str_error_AGE
                    break
                else:
                    Para_0_Dry_old = Paraupdate;       Model_Dry_old = Model_Dry_i;      Sol_Dry_old = Sol_Dry_i;   
                    del Paraupdate,Model_Dry_i,Sol_Dry_i
                    # post-process for first ageing cycle and every -1 ageing cycle
                    if k==0 and i==0:    
                        my_dict_AGE = GetSol_dict (my_dict_AGE,keys_all_AGE, Sol_Dry_old, 
                            0, step_AGE_CD , step_AGE_CC , step_RPT_RE, step_AGE_CV   )     
                        my_dict_AGE["Cycle_AGE"].append(1)
                    my_dict_AGE = GetSol_dict (my_dict_AGE,keys_all_AGE, Sol_Dry_old, 
                        cycle_no, step_AGE_CD , step_AGE_CC , step_RPT_RE, step_AGE_CV   )    
                    cycle_count +=  Update_Cycles; 
                    my_dict_AGE["Cycle_AGE"].append(cycle_count)           
                    Cyc_Update_Index.append(cycle_count)
                    print(f"Scan {Scan_i}: Finish for No.{Cyc_Update_Index[-1]} ageing cycles")
                    if DryOut == "On":
                        mdic_dry = Update_mdic_dry(Data_Pack,mdic_dry)
                    i += 1;   
            # run RPT, and also update parameters (otherwise will have problems)
            if DryOut == "On":
                Data_Pack , Paraupdate  = Cal_new_con_Update (  Sol_Dry_old,   Para_0_Dry_old   )
            if DryOut == "Off":
                Paraupdate = Para_0     
            try:
                Timelimit = int(3600*0.5)
                if Timeout == True:
                    timeout_RPT = TimeoutFunc(
                        Run_Model_Base_On_Last_Solution_RPT, 
                        timeout=Timelimit, 
                        timeout_val=Timeout_text)
                    Result_list_RPT = timeout_RPT(
                        Model_Dry_old  , Sol_Dry_old ,   
                        Paraupdate,      Experiment_RPT, RPT_Cycles, 
                        Temper_RPT ,mesh_par ,submesh_strech
                    )
                else:
                    Result_list_RPT = Run_Model_Base_On_Last_Solution_RPT(
                        Model_Dry_old  , Sol_Dry_old ,   
                        Paraupdate,      Experiment_RPT, RPT_Cycles, 
                        Temper_RPT ,mesh_par ,submesh_strech
                    )
                [Model_Dry_i, Sol_Dry_i,Call_RPT]  = Result_list_RPT
                if Return_Sol == True:
                    Sol_RPT.append(Sol_Dry_i)
                #print(f"Temperature for RPT is now: {Temper_RPT}")  
                if Call_RPT.success == False:
                    print("Fail due to Experiment error or infeasible")
                    str_error_RPT = "Experiment error or infeasible"
                    1/0 
                if Sol_Dry_i == Timeout_text:
                    print("Fail due to Timeout")
                    str_error_RPT = "Timeout"
                    1/0
                if Sol_Dry_i == "Model error or solver error":
                    print("Fail due to Model error or solver error")
                    str_error_RPT = "Model error or solver error"
                    1/0
            except ZeroDivisionError as e:
                print(f"Scan {Scan_i}: Fail during No.{Cyc_Update_Index[-1]} RPT cycles, due to {str_error_RPT}")
                break
            else:
                my_dict_RPT = GetSol_dict (my_dict_RPT,keys_all_RPT, Sol_Dry_i, 
                    cycle_no, step_RPT_CD , step_RPT_CC , step_RPT_RE, step_AGE_CV   )
                my_dict_RPT["Cycle_RPT"].append(cycle_count)
                Cyc_Update_Index.append(cycle_count)
                print(f"Scan {Scan_i}: Finish for No.{Cyc_Update_Index[-1]} RPT cycles")
                if DryOut == "On":
                    mdic_dry = Update_mdic_dry(Data_Pack,mdic_dry)
                Para_0_Dry_old = Paraupdate;    Model_Dry_old = Model_Dry_i  ;     Sol_Dry_old = Sol_Dry_i    ;   
                del Paraupdate,Model_Dry_i,Sol_Dry_i
                if Flag_AGE == False:
                    break
            k += 1 
    ############################################################# 
    #########   An extremely bad case: cannot even finish breakin
    if Flag_Breakin == False: 
        value_list_temp = list(Para_dict_i.values())
        values = []
        for value_list_temp_i in value_list_temp:
            values.append(str(value_list_temp_i))
        values.insert(0,str(Scan_i));
        values.insert(1,DryOut);
        values.extend([
            str_exp_AGE_text,
            str_exp_RPT_text,
            "nan","nan",
            "nan","nan", 
            "nan","nan",
            "nan","nan",
            "nan",str_error_Breakin])
        values = [values,]
        print(str_error_Breakin)
        print("Fail in {}".format(ModelTimer.time())) 
        book_name_xlsx_seperate =   str(Scan_i)+ '_' + book_name_xlsx;
        sheet_name_xlsx =  str(Scan_i);
        write_excel_xlsx(
            BasicPath + Target+book_name_xlsx_seperate, 
            sheet_name_xlsx, values)
        
        midc_merge = {**my_dict_RPT, **my_dict_AGE,**mdic_dry}

        return midc_merge,Sol_RPT,Sol_AGE
    ##########################################################
    ##############   Part-3: Post-prosessing    ##############
    ##########################################################
    # Newly add (220517): save plots, not just a single line in excel file:     
    # Newly add (221114): make plotting as functions
    # Niall_data = loadmat( 'Extracted_all_cell.mat'
    else:
        if not os.path.exists(BasicPath + Target + str(Scan_i)):
            os.mkdir(BasicPath + Target + str(Scan_i) );
        dpi= 100; fs = 17;
        ##########################################################
        #########      3-1: Plot cycle,location, Dryout related 
        Plot_Cyc_RPT_4(my_dict_RPT,Niall_data,Scan_i,Temper_i,model_options,BasicPath, Target,fs,dpi)
        if len(my_dict_AGE["CDend Porosity"])>1:
            Plot_Loc_AGE_4(my_dict_AGE,Scan_i,model_options,BasicPath, Target,fs,dpi)
        if DryOut == "On":
            Plot_Dryout(Cyc_Update_Index,mdic_dry,ce_EC_0,Scan_i,BasicPath, Target,fs,dpi)
        ##########################################################
        #########      3-2: Save data as .mat 
        my_dict_RPT["Cyc_Update_Index"] = Cyc_Update_Index
        my_dict_RPT["SaveTimes"]    = SaveTimes
        midc_merge = {**my_dict_RPT, **my_dict_AGE,**mdic_dry}
        savemat(BasicPath + Target+ str(Scan_i) + '/' + str(Scan_i)+ '-StructDara_for_Mat.mat',midc_merge)  
        ##########################################################
        #########      3-3: Save summary to excel 
        values=Get_Values_Excel(
            model_options,my_dict_RPT,mdic_dry,
            DryOut,Scan_i,Para_dict_i,str_exp_AGE_text,
            str_exp_RPT_text,
            str_error_AGE_final,
            str_error_RPT)
        values = [values,]
        book_name_xlsx_seperate =   str(Scan_i)+ '_' + book_name_xlsx;
        sheet_name_xlsx =  str(Scan_i);
        write_excel_xlsx(
            BasicPath + Target+book_name_xlsx_seperate, 
            sheet_name_xlsx, values)
        print("Succeed doing something in {}".format(ModelTimer.time()))
        print('This is the end of No.', Scan_i, ' scan')
        return midc_merge,Sol_RPT,Sol_AGE



# Run model with electrolyte dry out
def Run_model_wwo_dry_out(
      index_xlsx, Para_dict_i,   Path_pack , 
      keys_all,   exp_text_list, exp_index_pack , Niall_data ):

    ModelTimer = pb.Timer()

    # Un-pack data:
    [cycle_no,step_AGE_CD,step_AGE_CC,step_AGE_CV,
        step_RPT_CD,step_RPT_RE , step_RPT_CC ] = exp_index_pack;
    [exp_AGE_text, exp_RPT_text,] = exp_text_list;
    [BasicPath,Target,book_name_xlsx,sheet_name_xlsx,] = Path_pack
    CyclePack,Para_0 = Para_init(Para_dict_i)
    [Total_Cycles,Cycle_bt_RPT,Update_Cycles,RPT_Cycles,
        Temper_i,Temper_RPT,mesh_par,submesh_strech,model_options] = CyclePack;
    [keys_all_RPT,keys_all_AGE] = keys_all;

    Scan_i = int(index_xlsx);
    str_model_options = str(model_options);
    str_exp_AGE_text  = str(exp_AGE_text);
    str_exp_RPT_text  = str(exp_RPT_text);

    # define experiment
    Experiment_Long   = pb.Experiment( exp_AGE_text * Update_Cycles  )  
    Experiment_RPT    = pb.Experiment( exp_RPT_text * RPT_Cycles     ) 
    Experiment_Breakin= pb.Experiment( exp_RPT_text * RPT_Cycles     )

    ################ Important: index definition #################################
    Small_Loop =  int(Cycle_bt_RPT/Update_Cycles);   
    SaveTimes = int(Total_Cycles/Cycle_bt_RPT);   
    index = list(np.arange(1,SaveTimes+1)*(Small_Loop+RPT_Cycles)-1); #index.insert(0,0)  # get all last ageing cycle before RPT, plus the first RPT test
    index2= list(np.arange(0,SaveTimes+1)*(Small_Loop+RPT_Cycles));         # get all RPT test
    cycles = np.arange(0,SaveTimes+1)*Cycle_bt_RPT; 
    cycles3 = list(np.arange(1,SaveTimes+1)*(Cycle_bt_RPT));

    # initialize my_dict for outputs
    my_dict_RPT = {}; 
    for keys in keys_all_RPT:
        for key in keys:
            my_dict_RPT[key]=[];
    my_dict_AGE = {}; 
    for keys in keys_all_AGE:
        for key in keys:
            my_dict_AGE[key]=[];
            
    # update 2209-24: merge DryOut and Int_ElelyExces_Ratio
    temp_Int_ElelyExces_Ratio =  Para_0["Initial electrolyte excessive amount ratio"] 
    if temp_Int_ElelyExces_Ratio < 1:
        Int_ElelyExces_Ratio = -1;
        DryOut = "Off";
    else:
        Int_ElelyExces_Ratio = temp_Int_ElelyExces_Ratio;
        DryOut = "On";
    print(DryOut)
    try:             
        # define the model and run break-in cycle
        Model_0 = pb.lithium_ion.DFN(options=model_options ) #

        # update 220926 - add diffusivity and conductivity as variables:
        c_e = Model_0.variables["Electrolyte concentration [mol.m-3]"]
        T = Model_0.variables["Cell temperature [K]"]
        D_e = Para_0["Electrolyte diffusivity [m2.s-1]"]
        sigma_e = Para_0["Electrolyte conductivity [S.m-1]"]
        Model_0.variables["Electrolyte diffusivity [m2.s-1]"] = D_e(c_e, T)
        Model_0.variables["Electrolyte conductivity [S.m-1]"] = sigma_e(c_e, T)


        var = pb.standard_spatial_vars  
        var_pts = {
            var.x_n: 20,  
            var.x_s: 10,  
            var.x_p: 20,  
            var.r_n: int(mesh_par),  
            var.r_p: int(mesh_par),  }       
        submesh_types = Model_0.default_submesh_types
        if submesh_strech == "nan":
            pass
        else:
            particle_mesh = pb.MeshGenerator(
                pb.Exponential1DSubMesh, 
                submesh_params={"side": "right", "stretch": submesh_strech})
            submesh_types["negative particle"] = particle_mesh
            submesh_types["positive particle"] = particle_mesh 
        Sim_0    = pb.Simulation(
            Model_0,        experiment = Experiment_Breakin,
            parameter_values = Para_0,
            solver = pb.CasadiSolver(),
            var_pts=var_pts,
            submesh_types=submesh_types) #mode="safe"
        Sol_0    = Sim_0.solve(calc_esoh=False)
        print("Finish the break-in cycle")

        if DryOut == "On":    ######################   initialize for post-processing before running subsequent model   
            Vol_Elely_Tot_All  = [];       Vol_Elely_JR_All =[];         Vol_Pore_tot_All =[];
            Ratio_CeEC_All     = [];       Ratio_CeLi_All = [];          Ratio_Dryout_All =[];
            Vol_EC_consumed_All= [];       Vol_Elely_need_All = [];      Width_all        =[];
            Vol_Elely_add_All  = [];       Vol_Pore_decrease_All =[];    Test_V_All=[];                 
            Test_V2_All=[];                c_e_r_new_All=[];             c_EC_r_new_All=[]; 
            T_0                  =  Para_0['Initial temperature [K]']
            Porosity_Neg_0       =  Para_0['Negative electrode porosity']  
            Porosity_Pos_0       =  Para_0['Positive electrode porosity']  
            Porosity_Sep_0       =  Para_0['Separator porosity']  
            cs_Neg_Max           =  Para_0["Maximum concentration in negative electrode [mol.m-3]"];
            L_p                  =  Para_0["Positive electrode thickness [m]"]
            L_n                  =  Para_0["Negative electrode thickness [m]"]
            L_s                  =  Para_0["Separator thickness [m]"]
            L_y                  =  Para_0["Electrode width [m]"]
            Para_0.update({'Initial Electrode width [m]':L_y}, check_already_exists=False)
            L_y_0                =  Para_0["Initial Electrode width [m]"]
            L_z                  =  Para_0["Electrode height [m]"]
            Para_0.update({'Initial Electrode height [m]':L_z}, check_already_exists=False)
            L_z_0                =  Para_0["Initial Electrode height [m]"]
            
            Vol_Elely_Tot        =  ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  * L_y_0 * L_z_0 * Int_ElelyExces_Ratio # Set initial electrolyte amount [L] 
            Vol_Elely_JR         =  ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  * L_y_0 * L_z_0
            Vol_Pore_tot         =  ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  * L_y_0 * L_z_0
            Ratio_CeEC           =  1.0; Ratio_CeLi =  1.0  ;Ratio_Dryout         =  1.0
            Vol_EC_consumed      =  0;Vol_Elely_need       =  0;Vol_Elely_add        =  0
            Vol_Pore_decrease    =  0;Test_V2 = 0; Test_V = 0;
            print('Initial electrolyte amount is ', Vol_Elely_Tot*1e6, 'mL') 
            Para_0.update(
                {'Current total electrolyte volume in jelly roll [m3]':
                Vol_Elely_JR}, check_already_exists=False)
            Para_0.update(
                {'Current total electrolyte volume in whole cell [m3]':
                Vol_Elely_Tot}, check_already_exists=False)   
            
            Vol_Elely_Tot_All.append(Vol_Elely_Tot*1e6);            
            Vol_Elely_JR_All.append(Vol_Elely_JR*1e6);     
            Vol_Pore_tot_All.append(Vol_Pore_tot*1e6);           
            Ratio_CeEC_All.append(Ratio_CeEC);                      
            Ratio_CeLi_All.append(Ratio_CeLi);             
            Ratio_Dryout_All.append(Ratio_Dryout);
            Vol_EC_consumed_All.append(Vol_EC_consumed*1e6);        
            Vol_Elely_need_All.append(Vol_Elely_need*1e6);     
            Width_all.append(L_y_0);
            Vol_Elely_add_All.append(Vol_Elely_add*1e6);            
            Vol_Pore_decrease_All.append(Vol_Pore_decrease*1e6);
            Test_V_All.append(Test_V*1e6); 
            Test_V2_All.append(Test_V2*1e6); 
            c_e_r_new_All.append(1000.0); 
            c_EC_r_new_All.append(Para_0['EC initial concentration in electrolyte [mol.m-3]']); 
        
        # post-process for break-in cycle
        my_dict_RPT_old = my_dict_RPT; del my_dict_RPT
        my_dict_RPT = GetSol_dict (my_dict_RPT_old,keys_all_RPT, Sol_0, 
            cycle_no, step_RPT_CD , step_RPT_CC , step_RPT_RE, step_AGE_CV   )

        # Para_All.append(Para_0);                                Model_All.append(Model_0);    Sol_All_i.append(Sol_0); 
        Para_0_Dry_old = Para_0;     Model_Dry_old = Model_0  ; Sol_Dry_old = Sol_0;   del Model_0,Sol_0
        if DryOut == "On":
            del Para_0
        #Step3: Write a big loop to finish the long experiment,    
        Cyc_Update_Index     =[]; cycle_count =0; Cyc_Update_Index.append(cycle_count);
        k=0; 
        while k < SaveTimes:    # biggest loop, 
            i=0;     cap_i = 0;
            while i < Small_Loop:
                # run ageing cycles, and update parameters (have another version not update)
                if DryOut == "On":
                    Data_Pack,Paraupdate   = Cal_new_con_Update (  Sol_Dry_old,   Para_0_Dry_old )
                    [
                        Vol_EC_consumed, 
                        Vol_Elely_need, 
                        Test_V, 
                        Test_V2, 
                        Vol_Elely_add, 
                        Vol_Elely_Tot_new, 
                        Vol_Elely_JR_new, 
                        Vol_Pore_tot_new, 
                        Vol_Pore_decrease, 
                        c_e_r_new, c_EC_r_new,
                        Ratio_Dryout, Ratio_CeEC_JR, 
                        Ratio_CeLi_JR,
                        Width_new, ]= Data_Pack;
                    # Append single object to All object     
                    Vol_Elely_Tot_All.append(Vol_Elely_Tot_new*1e6);    Vol_Elely_JR_All.append(Vol_Elely_JR_new*1e6);     Vol_Pore_tot_All.append(Vol_Pore_tot_new*1e6);           
                    Ratio_CeEC_All.append(Ratio_CeEC_JR);               Ratio_CeLi_All.append(Ratio_CeLi_JR);                 Ratio_Dryout_All.append(Ratio_Dryout);
                    Vol_EC_consumed_All.append(Vol_EC_consumed*1e6);    Vol_Elely_need_All.append(Vol_Elely_need*1e6);     Width_all.append(Width_new);
                    Vol_Elely_add_All.append(Vol_Elely_add*1e6);        Vol_Pore_decrease_All.append(Vol_Pore_decrease*1e6);
                    Test_V_All.append(Test_V*1e6); Test_V2_All.append(Test_V2*1e6); 
                    c_e_r_new_All.append(c_e_r_new); c_EC_r_new_All.append(c_EC_r_new); 
                if DryOut == "Off":
                    Paraupdate = Para_0;

                # 3rd: run model based on new parameter and last updated solution Model  , Sol, Ratio_CeLi, Para_update, ModelExperiment, SaveAs_Cycles
                Model_Dry_i, Sol_Dry_i   = Run_Model_Base_On_Last_Solution( 
                    Model_Dry_old  , Sol_Dry_old ,  
                    Paraupdate ,Experiment_Long, Update_Cycles,Temper_i,mesh_par,submesh_strech )
                Para_0_Dry_old = Paraupdate;       Model_Dry_old = Model_Dry_i;      Sol_Dry_old = Sol_Dry_i;   
                
                if (k==0 and i==0) or (k==SaveTimes-1 and i==Small_Loop-1):
                    # post-process for first or last ageing cycle
                    my_dict_AGE_old = my_dict_AGE; del my_dict_AGE
                    my_dict_AGE = GetSol_dict (my_dict_AGE_old,keys_all_AGE, Sol_Dry_i, 
                        cycle_no, step_AGE_CD , step_AGE_CC , step_RPT_RE, step_AGE_CV   )

                del Paraupdate,Model_Dry_i,Sol_Dry_i
                i += 1;   cycle_count +=  Update_Cycles; 
                Cyc_Update_Index.append(cycle_count);

            # run RPT, and also update parameters (otherwise will have problems)
            if DryOut == "On":
                Data_Pack , Paraupdate  = Cal_new_con_Update (  Sol_Dry_old,   Para_0_Dry_old   )
                [
                        Vol_EC_consumed, 
                        Vol_Elely_need, 
                        Test_V, 
                        Test_V2, 
                        Vol_Elely_add, 
                        Vol_Elely_Tot_new, 
                        Vol_Elely_JR_new, 
                        Vol_Pore_tot_new, 
                        Vol_Pore_decrease, 
                        c_e_r_new, c_EC_r_new,
                        Ratio_Dryout, Ratio_CeEC_JR, 
                        Ratio_CeLi_JR,
                        Width_new, ]= Data_Pack;
                # Append solvent consumption model related variables     
                Vol_Elely_Tot_All.append(Vol_Elely_Tot_new*1e6);    Vol_Elely_JR_All.append(Vol_Elely_JR_new*1e6);     Vol_Pore_tot_All.append(Vol_Pore_tot_new*1e6);           
                Ratio_CeEC_All.append(Ratio_CeEC_JR);                  Ratio_CeLi_All.append(Ratio_CeLi_JR);                 Ratio_Dryout_All.append(Ratio_Dryout);
                Vol_EC_consumed_All.append(Vol_EC_consumed*1e6);    Vol_Elely_need_All.append(Vol_Elely_need*1e6);     Width_all.append(Width_new);
                Vol_Elely_add_All.append(Vol_Elely_add*1e6);        Vol_Pore_decrease_All.append(Vol_Pore_decrease*1e6); 
                Test_V_All.append(Test_V*1e6); Test_V2_All.append(Test_V2*1e6); 
                c_e_r_new_All.append(c_e_r_new); c_EC_r_new_All.append(c_EC_r_new); 
            if DryOut == "Off":
                    Paraupdate = Para_0;

            Cyc_Update_Index.append(cycle_count);

            Model_Dry_i, Sol_Dry_i  = Run_Model_Base_On_Last_Solution_RPT( 
                Model_Dry_old  , Sol_Dry_old ,   
                Paraupdate,      Experiment_RPT, RPT_Cycles, 
                Temper_RPT ,mesh_par ,submesh_strech )     
            #if k == SaveTimes-1:  # post-process for last RPT cycle
            my_dict_RPT_old = my_dict_RPT; del my_dict_RPT
            my_dict_RPT = GetSol_dict (my_dict_RPT_old,keys_all_RPT, Sol_Dry_i, 
                cycle_no, step_RPT_CD , step_RPT_CC , step_RPT_RE, step_AGE_CV   )

            Para_0_Dry_old = Paraupdate;    Model_Dry_old = Model_Dry_i  ;                 
            Sol_Dry_old = Sol_Dry_i    ;   del Paraupdate,Model_Dry_i,Sol_Dry_i
            k += 1;    Cyc_Update_Index =Cyc_Update_Index; 
    except:
        #data = openpyxl.load_workbook(BasicPath + Target + book_name_xlsx)   
        str_error = traceback.format_exc()      
        #table = data.get_sheet_by_name(sheet_name_xlsx)
        #nrows = table.max_row  # 获得行数
        #ncolumns = table.max_column  # 获得列数
        value_list_temp = list(Para_dict_i.values())
        values = []
        for value_list_temp_i in value_list_temp:
            values.append(str(value_list_temp_i))
        values.insert(0,str(Scan_i));
        values.insert(1,DryOut);
        values.extend([
            str_exp_AGE_text,
            str_exp_RPT_text,
            "nan","nan",
            "nan","nan", 
            "nan","nan",
            "nan","nan",
            "nan",str_error])
        values = [values,]     # add a list bracket to ensure proper write
        """ for i in range(1, len(values)+1):
            for j in range(1, len(values[i-1])+1):
                table.cell(nrows+i, j).value = values[i-1][j-1]     
        data.save(BasicPath + Target + book_name_xlsx)  """
        print(str_error)
        print("Fail in {}".format(ModelTimer.time())) 
        book_name_xlsx_seperate =   str(Scan_i)+ '_' + book_name_xlsx;
        sheet_name_xlsx =  str(Scan_i);
        write_excel_xlsx(
            BasicPath + Target+book_name_xlsx_seperate, 
            sheet_name_xlsx, values)
    else:
        # Finish everything, try to write into excel with real results or 
        """ data = openpyxl.load_workbook(BasicPath + Target+ book_name_xlsx)         
        table = data.get_sheet_by_name('Results')
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得列数 """
        if model_options.__contains__("SEI on cracks"):
            LossCap_seioncrack = my_dict_RPT["CDend Loss of capacity to SEI on cracks [A.h]"][-1]
        else:
            LossCap_seioncrack = "nan"
        if model_options.__contains__("lithium plating"):
            LossCap_LiP = my_dict_RPT["CDend Loss of capacity to lithium plating [A.h]"][-1]
        else:
            LossCap_LiP = "nan"
        if model_options.__contains__("SEI"):
            LossCap_SEI = my_dict_RPT["CDend Loss of capacity to SEI [A.h]"][-1]
        else:
            LossCap_SEI = "nan"
        if DryOut == "On":
            Vol_Elely_Tot_All_final = Vol_Elely_Tot_All[-1];
            Vol_Elely_JR_All_final  = Vol_Elely_JR_All[-1];
            Width_all_final         = Width_all[-1];
        else:
            Vol_Elely_Tot_All_final = "nan"
            Vol_Elely_JR_All_final  = "nan"
            Width_all_final         = "nan"
        value_list_temp = list(Para_dict_i.values())
        values = []
        for value_list_temp_i in value_list_temp:
            values.append(str(value_list_temp_i))
        values.insert(0,str(Scan_i));
        values.insert(1,DryOut);
        values.extend([
            str_exp_AGE_text,
            str_exp_RPT_text,
            str(my_dict_RPT["Discharge capacity [A.h]"][0] 
            - 
            my_dict_RPT["Discharge capacity [A.h]"][-1]),

            str(LossCap_LiP),
            str(LossCap_SEI),
            str(LossCap_seioncrack),

            str(my_dict_RPT["CDend Negative electrode capacity [A.h]"][0] 
            - 
            my_dict_RPT["CDend Negative electrode capacity [A.h]"][-1]),

            str(my_dict_RPT["CDend Positive electrode capacity [A.h]"][0] 
            - 
            my_dict_RPT["CDend Positive electrode capacity [A.h]"][-1]),

            str(Vol_Elely_Tot_All_final), 
            str(Vol_Elely_JR_All_final),
            str(Width_all_final),
            ])
        
        print("Succeed in {}".format(ModelTimer.time()))
        print('This is the ', Scan_i, ' scan')
        if not os.path.exists(BasicPath + Target + str(Scan_i)):
            os.mkdir(BasicPath + Target + str(Scan_i) );
        
        # Newly add (220517): save plots, not just a single line in excel file:     
        # Fig. 1 how much capacity is loss, how much is due to SEI and LiP?
        # Niall_data = loadmat( 'Extracted_all_cell.mat')
        for mm in range(0,1):
            fs=17;Num_subplot = 3;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(18,4.8),tight_layout=True)
            axs[0].plot(cycles, my_dict_RPT["Discharge capacity [A.h]"],     '-o', label="Scan=" + str(Scan_i) )
            if Temper_i  == 40:       
                axs[0].plot(Niall_data['F_Cap_all'][:,2],Niall_data['F_Cap_all'][:,5]/1e3, '-^',  label='Cell F' )
                axs[0].plot(Niall_data['G_Cap_all'][:,2],Niall_data['G_Cap_all'][:,5]/1e3, '-^',  label='Cell G' )
                axs[0].plot(Niall_data['H_Cap_all'][:,2],Niall_data['H_Cap_all'][:,5]/1e3, '-^',  label='Cell H' )
            elif Temper_i  == 25: 
                axs[0].plot(Niall_data['D_Cap_all'][:,2],Niall_data['D_Cap_all'][:,5]/1e3, '-^',  label='Cell D' )
                axs[0].plot(Niall_data['E_Cap_all'][:,2],Niall_data['E_Cap_all'][:,5]/1e3, '-^',  label='Cell E' )
            elif Temper_i  == 10: 
                axs[0].plot(Niall_data['A_Cap_all'][:,2],Niall_data['A_Cap_all'][:,5]/1e3, '-^',  label='Cell A' )
                axs[0].plot(Niall_data['B_Cap_all'][:,2],Niall_data['B_Cap_all'][:,5]/1e3, '-^',  label='Cell B' )
                axs[0].plot(Niall_data['C_Cap_all'][:,2],Niall_data['C_Cap_all'][:,5]/1e3, '-^',  label='Cell C' )
            else:
                pass
            if model_options.__contains__("lithium plating"):
                axs[1].plot(cycles, my_dict_RPT["CDend Loss of capacity to lithium plating [A.h]"],'-o', label="LiP - Scan=" + str(Scan_i) )
            if model_options.__contains__("SEI"):
                axs[1].plot(cycles, my_dict_RPT["CDend Loss of capacity to SEI [A.h]"] ,'-o', label="SEI - Scan" + str(Scan_i) )
            if model_options.__contains__("SEI on cracks"):
                axs[1].plot(cycles, my_dict_RPT["CDend Loss of capacity to SEI on cracks [A.h]"] ,'-o', label="sei-on-cracks - Scan" + str(Scan_i) )
            axs[2].plot(
                my_dict_RPT["CDend Throughput capacity [A.h]"], 
                my_dict_RPT["Discharge capacity [A.h]"],     
                '-o', label="Scan=" + str(Scan_i) )
            if Temper_i  == 40:       
                axs[2].plot(Niall_data['F_Cap_all'][:,3],Niall_data['F_Cap_all'][:,5]/1e3, '-^',  label='Cell F' )
                axs[2].plot(Niall_data['G_Cap_all'][:,3],Niall_data['G_Cap_all'][:,5]/1e3, '-^',  label='Cell G' )
                axs[2].plot(Niall_data['H_Cap_all'][:,3],Niall_data['H_Cap_all'][:,5]/1e3, '-^',  label='Cell H' )
            elif Temper_i  == 25: 
                axs[2].plot(Niall_data['D_Cap_all'][:,3],Niall_data['D_Cap_all'][:,5]/1e3, '-^',  label='Cell D' )
                axs[2].plot(Niall_data['E_Cap_all'][:,3],Niall_data['E_Cap_all'][:,5]/1e3, '-^',  label='Cell E' )
            elif Temper_i  == 10: 
                axs[2].plot(Niall_data['A_Cap_all'][:,3],Niall_data['A_Cap_all'][:,5]/1e3, '-^',  label='Cell A' )
                axs[2].plot(Niall_data['B_Cap_all'][:,3],Niall_data['B_Cap_all'][:,5]/1e3, '-^',  label='Cell B' )
                axs[2].plot(Niall_data['C_Cap_all'][:,3],Niall_data['C_Cap_all'][:,5]/1e3, '-^',  label='Cell C' )
            else:
                pass
            for i in range(0,Num_subplot):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[i].set_ylabel("Capacity [A.h]",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
                axs[i].set_title("Discharge capacity",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[2].set_xlabel("Throughput capacity [A.h]",   fontdict={'family':'DejaVu Sans','size':fs})   
            axs[1].set_title("Capacity loss to LiP and SEI",   fontdict={'family':'DejaVu Sans','size':fs+1})
            plt.savefig(BasicPath + Target+ str(Scan_i)+"/Cap-LLI.png", dpi=100)

            if model_options.__contains__("SEI on cracks"):
                fs=17;Num_subplot = 2;
                fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
                axs[0].plot(cycles, my_dict_RPT["CDend X-averaged total SEI on cracks thickness [m]"],     '-o', label="Scan=" + str(Scan_i) )
                axs[1].plot(cycles, my_dict_RPT["CDend X-averaged negative electrode roughness ratio"],'-o', label="Scan=" + str(Scan_i) )
                axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[1].set_ylabel("Roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs})
                for i in range(0,Num_subplot):
                    axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                    labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                    axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                    axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
                axs[0].set_title("X-avg tot Neg SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
                axs[1].set_title("X-avg Neg roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
                plt.savefig(BasicPath + Target+ str(Scan_i)+"/Cracks related.png", dpi=100)

                Num_subplot = 2;
                fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
                axs[1].plot(cycles, 
                    my_dict_RPT["CDend Negative electrode capacity [A.h]"][0]
                    -
                    my_dict_RPT["CDend Negative electrode capacity [A.h]"],'-o',label="Neg Scan=" + str(Scan_i))
                axs[1].plot(cycles, 
                    my_dict_RPT["CDend Positive electrode capacity [A.h]"][0]
                    -
                    my_dict_RPT["CDend Positive electrode capacity [A.h]"],'-^',label="Pos Scan=" + str(Scan_i))
                axs[0].plot(cycles, my_dict_RPT["CDend X-averaged total SEI on cracks thickness [m]"],                  '-o',label="Scan="+ str(Scan_i))
                for i in range(0,1):
                    axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                    labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                    axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                    axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
                axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[0].set_title("CDend X-avg tot SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
                for i in range(1,2):
                    axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                    axs[i].set_ylabel("Capacity [A.h]",   fontdict={'family':'DejaVu Sans','size':fs})
                    labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                    axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                    axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
                axs[1].set_title("LAM of Neg and Pos",   fontdict={'family':'DejaVu Sans','size':fs+1})
                plt.savefig(BasicPath + Target+ str(Scan_i)+"/LAM-IR.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(cycles, my_dict_RPT["CDsta Positive electrode SOC"] ,'-o',label="Start" )
            axs[0].plot(cycles, my_dict_RPT["CDend Positive electrode SOC"] ,'-^',label="End" )
            axs[1].plot(cycles, my_dict_RPT["CDsta Negative electrode SOC"],'-o',label="Start" )
            axs[1].plot(cycles, my_dict_RPT["CDend Negative electrode SOC"],'-^',label="End" )
            for i in range(0,2):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[i].set_ylabel("SOC",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)     
            axs[0].set_title("Neg SOC range (Dis)",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Pos SOC range (Dis)",   fontdict={'family':'DejaVu Sans','size':fs+1})
            plt.savefig(BasicPath + Target+ str(Scan_i)+"/SOC_RPT_dis.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Porosity"][0],'-o',label="First")
            axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Porosity"][-1],'-^',label="Last"  )
            axs[1].plot(
                my_dict_AGE["x_n [m]"],
                my_dict_AGE["CDend Negative electrode reaction overpotential [V]"][0],'-o',label="First" )
            axs[1].plot(
                my_dict_AGE["x_n [m]"],
                my_dict_AGE["CDend Negative electrode reaction overpotential [V]"][-1],'-^',label="Last" )

            axs[0].set_xlabel("Dimensional Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_xlabel("Dimensional Neg thickness",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[0].set_title("Porosity",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Neg electrode reaction overpotential",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[0].set_ylabel("Porosity",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_ylabel("Overpotential [V]",   fontdict={'family':'DejaVu Sans','size':fs})
            for i in range(0,2):
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
            plt.savefig(BasicPath + Target+ str(Scan_i)+"/Por Neg_S_eta.png", dpi=100)

            if model_options.__contains__("SEI on cracks"):
                Num_subplot = 2;
                fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
                axs[0].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Negative electrode roughness ratio"][0],'-o',label="First")
                axs[0].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Negative electrode roughness ratio"][-1],'-^',label="Last"  )
                axs[1].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Total SEI on cracks thickness [m]"][0],'-o',label="First" )
                axs[1].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Total SEI on cracks thickness [m]"][-1],'-^',label="Last" )
                axs[0].set_title("Tot Neg SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
                axs[1].set_title("Neg roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
                axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[1].set_ylabel("Roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs})
                for i in range(0,2):
                    axs[i].set_xlabel("Dimensional Neg thickness",   fontdict={'family':'DejaVu Sans','size':fs})
                    labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                    axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                    axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
                plt.savefig(BasicPath + Target+ str(Scan_i)+"/Cracks related spatial.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte concentration [mol.m-3]"][0],'-o',label="First")
            axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte concentration [mol.m-3]"][-1],'-^',label="Last"  )
            axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte potential [V]"][0],'-o',label="First" )
            axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte potential [V]"][-1],'-^',label="Last" )
            axs[0].set_title("Electrolyte concentration",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Electrolyte potential",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[0].set_ylabel("Concentration [mol.m-3]",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_ylabel("Potential [V]",   fontdict={'family':'DejaVu Sans','size':fs})
            for i in range(0,2):
                axs[i].set_xlabel("Dimensional Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
            plt.savefig(BasicPath + Target+ str(Scan_i)+"/Electrolyte concentration and potential.png", dpi=100)
            
            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte diffusivity [m2.s-1]"][0],'-o',label="First")
            axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte diffusivity [m2.s-1]"][-1],'-^',label="Last"  )
            axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte conductivity [S.m-1]"][0],'-o',label="First" )
            axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte conductivity [S.m-1]"][-1],'-^',label="Last" )
            axs[0].set_title("Electrolyte diffusivity",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Electrolyte conductivity",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[0].set_ylabel("Diffusivity [m2.s-1]",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_ylabel("Conductivity [S.m-1]",   fontdict={'family':'DejaVu Sans','size':fs})
            for i in range(0,2):
                axs[i].set_xlabel("Dimensional Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
            plt.savefig(BasicPath + Target+ str(Scan_i)+"/Electrolyte diffusivity and conductivity.png", dpi=100)
            

            if DryOut == "On":
                Num_subplot = 3;
                fig, axs = plt.subplots(1,Num_subplot, figsize=(18,4.8),tight_layout=True)
                axs[0].plot(Cyc_Update_Index, Vol_EC_consumed_All,'-o',label="EC consumed")
                axs[0].plot(Cyc_Update_Index, Vol_Elely_need_All,'-.',label="Elely needed")
                axs[0].plot(Cyc_Update_Index, Vol_Elely_add_All,'-s',label="Elely added")
                axs[0].plot(Cyc_Update_Index, Vol_Pore_decrease_All,'--',label="Pore decreased")
                axs[1].plot(Cyc_Update_Index, Vol_Elely_Tot_All,     '-o', label="Total electrolyte in cell" )
                axs[1].plot(Cyc_Update_Index, Vol_Elely_JR_All,     '--', label="Total electrolyte in JR" )
                axs[1].plot(Cyc_Update_Index, Vol_Pore_tot_All,     '-s', label="Total pore in JR" )
                axs[2].plot(Cyc_Update_Index, Ratio_Dryout_All,     '-s', label="Dry out ratio" )
                axs[2].set_ylabel("Ratio",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[2].set_xlabel("Cycle number",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[2].set_title("Dry out ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
                for i in range(0,2):
                    axs[i].set_xlabel("Cycle number",   fontdict={'family':'DejaVu Sans','size':fs})
                    axs[i].set_ylabel("Volume [mL]",   fontdict={'family':'DejaVu Sans','size':fs})
                    axs[i].set_title("Volume",   fontdict={'family':'DejaVu Sans','size':fs+1})
                    labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                    axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                    axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
                plt.savefig(BasicPath + Target+ str(Scan_i)+"/Volume_total.png", dpi=100)
        
        # Newly add (220706): save data, not just a single line in excel file:
        Bulk_Sol_Con_i = Para_0_Dry_old['EC initial concentration in electrolyte [mol.m-3]']
        mdic_cycles = {
            "cycles": cycles,
            "Cyc_Update_Index":Cyc_Update_Index,
            "SaveTimes": SaveTimes,
        }
        if DryOut == "On":
            for mm in range(0,1):
                CeEC_All =np.full(np.size(Ratio_CeEC_All),Bulk_Sol_Con_i); 
                for i in range(1,np.size(Ratio_CeEC_All)):
                    for k in range(0,i):
                        CeEC_All[i] *= Ratio_CeEC_All[k];        
            mdic_dry = {
                "CeEC_All": CeEC_All,
                "c_EC_r_new_All": c_EC_r_new_All,
                "c_e_r_new_All": c_e_r_new_All,
                "Ratio_CeEC_All":Ratio_CeEC_All ,  
                "Ratio_CeLi_All":Ratio_CeLi_All  ,       
                "Ratio_Dryout_All":Ratio_Dryout_All,

                "Vol_Elely_Tot_All": Vol_Elely_Tot_All,
                "Vol_Elely_JR_All": Vol_Elely_JR_All,
                "Vol_Pore_tot_All": Vol_Pore_tot_All,            
                "Vol_EC_consumed_All": Vol_EC_consumed_All,
                "Vol_Elely_need_All":Vol_Elely_need_All,
                "Width_all":Width_all,
                "Vol_Elely_add_All":Vol_Elely_add_All,
                "Vol_Pore_decrease_All":Vol_Pore_decrease_All,
                "Test_V_All":Test_V_All,
                "Test_V2_All":Test_V2_All,
            }
            midc_merge = {**my_dict_RPT, **mdic_cycles,**mdic_dry}
        else:
            midc_merge = {**my_dict_RPT, **mdic_cycles}

        savemat(BasicPath + Target+ str(Scan_i) + '/' + str(Scan_i)+ '-StructDara_for_Mat.mat',midc_merge)   
        savemat(BasicPath + Target+ str(Scan_i) + '/' + str(Scan_i)+ '-for_AGE_only.mat',my_dict_AGE)   

        values = [values,]
        """ 
        for i in range(1, len(values)+1):
            for j in range(1, len(values[i-1])+1):
                table.cell(nrows+i, j).value = values[i-1][j-1]     
        data.save(BasicPath + Target+ book_name_xlsx)     """
        book_name_xlsx_seperate =   str(Scan_i)+ '_' + book_name_xlsx;
        sheet_name_xlsx =  str(Scan_i);
        write_excel_xlsx(
            BasicPath + Target+book_name_xlsx_seperate, 
            sheet_name_xlsx, values)

        
        
        
        
  
        
        







































































