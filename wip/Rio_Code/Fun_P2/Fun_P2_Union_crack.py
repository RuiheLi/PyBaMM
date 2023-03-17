""" 
Variables to set as input:
Total_Cycles,Cycle_bt_RPT,Update_Cycles
cs_Neg_Init, Diff_SEI, R_SEI, Bulk_Sol_Con
BasicPath, Target, book_name_xlsx,
key scan parameters: 
D_Li_inSEI, c_Li_inte_ref, Couple_SEI_LiP, k_LiP, Ratio_excess

Key Output:
Cap_RPT_Dry_All,CapLoss_LiP_Dry_All[-1],
CapLoss_SEI_Dry_All[-1],CapLoss_SEIcr_Dry_All[-1],t_SEIcr_Xavgtot_Dry_All[-1], 
Rough_Neg_Xavg_Dry_All[-1],
Vol_Elely_Tot_All[-1], 
Vol_Elely_JR_All[-1],Width_all[-1]; 

also establish a folder to store data and .png

control parameter for different degradation mechanisms: 
for sei on cracks:
    Negative electrode cracking rate # function or 3.9e-20, Ai_2020 add temperature dependency
    Positive electrode cracking rate # function or 3.9e-20
    Outer SEI solvent diffusivity [m2.s-1] # 1.7e-20
    Bulk solvent concentration [mol.m-3] # 2326
for LAM due to cracks:
    Negative electrode LAM constant propotional term # default 1e-3
    Positive electrode LAM constant propotional term # default 1e-3
for solvent-diffusion limited sei:
    Outer SEI solvent diffusivity [m2.s-1] # 1.7e-20
    Bulk solvent concentration [mol.m-3] # 2326
for ec reaction limited sei:
    EC diffusivity [m2.s-1]	# default 2e-18
    EC initial concentration in electrolyte [mol.m-3] # default 4541, set as default
    SEI kinetic rate constant [m.s-1] # default 1e-12
for interstitial-diffusion limited:
    Inner SEI lithium interstitial diffusivity [m2.s-1] # default 1e-12 
    Lithium interstitial reference concentration [mol.m-3] # default 15 (doesn't matter)


further add-on: more data plot?
"""
import pybamm as pb;import pandas as pd   ;import numpy as np;import os;import matplotlib.pyplot as plt;import os;#import imageio
from scipy.io import savemat,loadmat;from pybamm import constants,exp;import matplotlib as mpl; fs=17; # or we can set import matplotlib.pyplot as plt then say 'mpl.rc...'
for k in range(0,1):
    mpl.rcParams["axes.labelsize"] = fs
    mpl.rcParams["axes.titlesize"] = fs
    mpl.rcParams["xtick.labelsize"] =  fs
    mpl.rcParams["ytick.labelsize"] =  fs
    mpl.rcParams["legend.fontsize"] =  fs
    mpl.rcParams['font.sans-serif'] = ['DejaVu Sans']
    mpl.rcParams['font.serif'] = ['DejaVu Sans']
    mpl.rcParams['axes.unicode_minus'] = False

import openpyxl
import traceback
###################################################################
#############           Step-2 Define Functions        ############
###################################################################

from pybamm import exp, constants, Parameter
def nmc_LGM50_electrolyte_exchange_current_density_ORegan2021_adjust(c_e, c_s_surf, T):
    i_ref = 5.028  # (A/m2)
    alpha = 0.43
    E_r = 2.401e4
    arrhenius = exp(E_r / constants.R * (1 / 298.15 - 1 / T))
    c_p_max = Parameter("Maximum concentration in positive electrode [mol.m-3]")
    c_e_ref = Parameter("Typical electrolyte concentration [mol.m-3]")
    return (
        i_ref
        * arrhenius
        * (c_e / c_e_ref) ** (1 - alpha)
        * (c_s_surf / c_p_max) ** alpha
        * (1 - c_s_surf / c_p_max) ** (1 - alpha))

def nmc_LGM50_diffusivity_ORegan2021_adjust(sto, T):
    a1 = -0.9231
    a2 = -0.4066
    a3 = -0.993
    b1 = 0.3216
    b2 = 0.4532
    b3 = 0.8098
    c0 = -13.96
    c1 = 0.002534
    c2 = 0.003926
    c3 = 0.09924
    d = 1449

    D_ref = (
        10
        ** (
            c0
            + a1 * exp(-((sto - b1) ** 2) / c1)
            + a2 * exp(-((sto - b2) ** 2) / c2)
            + a3 * exp(-((sto - b3) ** 2) / c3)
        )
        * 2.7  # correcting factor (see O'Regan et al 2021)
    )

    E_D_s = d * constants.R
    arrhenius = exp(E_D_s / constants.R * (1 / 298.15 - 1 / T))

    return D_ref * arrhenius

def graphite_LGM50_diffusivity_ORegan2021_adjust(sto, T):
    a0 = 11.17
    a1 = -1.553
    a2 = -6.136
    a3 = -9.725
    a4 = 1.85
    b1 = 0.2031
    b2 = 0.5375
    b3 = 0.9144
    b4 = 0.5953
    c0 = -15.11
    c1 = 0.0006091
    c2 = 0.06438
    c3 = 0.0578
    c4 = 0.001356
    d = 2092

    D_ref = (
        10
        ** (
            a0 * sto
            + c0
            + a1 * exp(-((sto - b1) ** 2) / c1)
            + a2 * exp(-((sto - b2) ** 2) / c2)
            + a3 * exp(-((sto - b3) ** 2) / c3)
            + a4 * exp(-((sto - b4) ** 2) / c4)
        )
        * 20  # correcting factor (see O'Regan et al 2021) mark: Ruihe change from 3.0321 
    )

    E_D_s = d * constants.R
    arrhenius = exp(E_D_s / constants.R * (1 / 298.15 - 1 / T))

    return D_ref * arrhenius

def graphite_LGM50_electrolyte_exchange_current_density_ORegan2021_adjust(c_e, c_s_surf, T):
    i_ref = 12.4  # (A/m2) replaced with value from Schmalstieg et al. (2018), from 2.668 
    alpha = 0.792
    E_r = 4e4
    arrhenius = exp(E_r / constants.R * (1 / 298.15 - 1 / T))

    c_n_max = Parameter("Maximum concentration in negative electrode [mol.m-3]")
    c_e_ref = Parameter("Typical electrolyte concentration [mol.m-3]")

    return (
        i_ref
        * arrhenius
        * (c_e / c_e_ref) ** (1 - alpha)
        * (c_s_surf / c_n_max) ** alpha
        * (1 - c_s_surf / c_n_max) ** (1 - alpha))

def nmc_LGM50_electronic_conductivity_ORegan2021_adjust(T):
    E_r = 3.5e3
    arrhenius = exp(E_r / constants.R * (1 / 298.15 - 1 / T))

    sigma = 0.8473 * arrhenius

    return sigma



# Define a function to calculate concentration change, whether electrolyte being squeezed out or added in
def Cal_new_con_Update(Sol,Para):   # subscript r means the reservoir
    # Note: c_EC_r is the initial EC  concentraiton in the reservoir; 
    #       c_e_r  is the initial Li+ concentraiton in the reservoir;
    #############################################################################################################################  
    ###############################           Step-1 Prepare parameters:        #################################################
    #############################################################################################################################  
    L_p   =Para["Positive electrode thickness [m]"]
    L_n   =Para["Negative electrode thickness [m]"]
    L_s   =Para["Separator thickness [m]"]
    L_y   =Para["Electrode width [m]"]   # Also update to change A_cc and therefore Q
    L_z   =Para["Electrode height [m]"]
    c_EC_r_old=Para["Current solvent concentration in the reservoir [mol.m-3]"]   # add two new parameters for paper 2
    c_e_r_old =Para["Current electrolyte concentration in the reservoir [mol.m-3]"]
    # Initial EC concentration in JR
    c_EC_JR_old =Para["Bulk solvent concentration [mol.m-3]"]  # Be careful when multiplying with pore volume to get amount in mole. Because of electrolyte dry out, that will give more amount than real.   
    # LLI due to electrode,Ratio of EC and lithium is 1:1 -> EC amount consumed is LLINegSEI[-1]
    LLINegSEI = (
        Sol["Loss of lithium to SEI [mol]"].entries[-1] 
        - Sol["Loss of lithium to SEI [mol]"].entries[0] )
    LLINegSEIcr = (
        Sol["Loss of lithium to SEI on cracks [mol]"].entries[-1]
        - 
        Sol["Loss of lithium to SEI on cracks [mol]"].entries[0]
        )
    LLINegDeadLiP = (
        Sol["Loss of lithium to dead lithium plating [mol]"].entries[-1] 
        - Sol["Loss of lithium to dead lithium plating [mol]"].entries[0])
    LLINegLiP = (
        Sol["Loss of lithium to lithium plating [mol]"].entries[-1] 
        - Sol["Loss of lithium to lithium plating [mol]"].entries[0])
    cLi_Xavg  = Sol["X-averaged electrolyte concentration [mol.m-3]"].entries[-1] 
    # Pore volume change with time:
    PoreVolNeg_0 = Sol["X-averaged negative electrode porosity"].entries[0]*L_n*L_y*L_z;
    PoreVolSep_0 = Sol["X-averaged separator porosity"].entries[0]*L_s*L_y*L_z;
    PoreVolPos_0 = Sol["X-averaged positive electrode porosity"].entries[0]*L_p*L_y*L_z;
    PoreVolNeg_1 = Sol["X-averaged negative electrode porosity"].entries[-1]*L_n*L_y*L_z;
    PoreVolSep_1 = Sol["X-averaged separator porosity"].entries[-1]*L_s*L_y*L_z;
    PoreVolPos_1 = Sol["X-averaged positive electrode porosity"].entries[-1]*L_p*L_y*L_z;
    #############################################################################################################################  
    ##### Step-2 Determine How much electrolyte is added (from the reservoir to JR) or squeezed out (from JR to reservoir) ######
    #######################   and finish electrolyte mixing     #################################################################  
    Vol_Elely_Tot_old = Para["Current total electrolyte volume in whole cell [m3]"] 
    Vol_Elely_JR_old  = Para["Current total electrolyte volume in jelly roll [m3]"] 
    if Vol_Elely_Tot_old - Vol_Elely_JR_old < 0:
        print('Model error! Electrolyte in JR is larger than in the cell!')
    Vol_Pore_tot_old  = PoreVolNeg_0 + PoreVolSep_0 + PoreVolPos_0    # pore volume at start time of the run
    Vol_Pore_tot_new  = PoreVolNeg_1 + PoreVolSep_1 + PoreVolPos_1    # pore volume at end   time of the run, intrinsic variable 
    Vol_Pore_decrease = Vol_Elely_JR_old  - Vol_Pore_tot_new;
    # EC:lithium:SEI=2:2:1     for SEI=(CH2OCO2Li)2, but because of too many changes are required, change to 2:1:1 for now
    # Because inner and outer SEI partial molar volume is the same, just set one for whole SEI
    VmolSEI   = Para["Outer SEI partial molar volume [m3.mol-1]"] # 9.8e-5,
    VmolLiP   = Para["Lithium metal partial molar volume [m3.mol-1]"] # 1.3e-05
    VmolEC    = 6.667e-5  # Unit:m3.mol-1; According to Wiki, correct value: EC molar volume is :66.67 cm3.mol-1  = 6.667e-5, 
    Vol_EC_consumed  =  ( LLINegSEI + LLINegSEIcr + LLINegDeadLiP  ) * 2 * VmolEC    # Mark: Ruihe add LLINegDeadLiP, either with 2 or not, will decide how fast electrolyte being consumed!
    Vol_Elely_need   = Vol_EC_consumed - Vol_Pore_decrease
    Vol_SEILiP_increase = 1.0*(
        (LLINegSEI+LLINegSEIcr) * VmolSEI 
        + LLINegLiP * VmolLiP)    #  volume increase due to SEI+total LiP 
    Test_V = Vol_SEILiP_increase - Vol_Pore_decrease  #  This value should always be zero, but now not, which becomes the source of error!
    Test_V2= (Vol_Pore_tot_old - Vol_Elely_JR_old) / Vol_Elely_JR_old * 100; # Porosity errors due to first time step
    
    # Start from here, there are serveral variables to be determined:
    #   1) Vol_Elely_add, or Vol_Elely_squeezed; depends on conditions. easier for 'squeezed' condition
    #   2) Vol_Elely_Tot_new, should always equals to Vol_Elely_Tot_old -  Vol_EC_consumed;
    #   3) Vol_Elely_JR_new, for 'add' condition: see old code; for 'squeezed' condition, equals to pore volume in JR
    #   4) Ratio_Dryout, for 'add' condition: see old code; for 'squeezed' condition, equals to Vol_Elely_Tot_new/Vol_Pore_tot_new 
    #   5) Ratio_CeEC_JR and Ratio_CeLi_JR: for 'add' condition: see old code; for 'squeezed' condition, equals to 1 (unchanged)
    #   6) c_e_r_new and c_EC_r_new: for 'add' condition: equals to old ones (unchanged); for 'squeezed' condition, need to carefully calculate     
    #   7) Width_new: for 'add' condition: see old code; for 'squeezed' condition, equals to L_y (unchanged)
    Vol_Elely_Tot_new = Vol_Elely_Tot_old - Vol_EC_consumed;

    
    if Vol_Elely_need < 0:
        print('Electrolyte is being squeezed out, check plated lithium (reversible part)')
        Vol_Elely_squeezed = - Vol_Elely_need;   # Make Vol_Elely_squeezed>0 for simplicity 
        Vol_Elely_add = 0.0;
        Vol_Elely_JR_new = Vol_Pore_tot_new; 
        Ratio_Dryout = Vol_Elely_Tot_new / Vol_Elely_JR_new
        Ratio_CeEC_JR = 1.0;    # the concentrations in JR don't need to change
        Ratio_CeLi_JR = 1.0; 
        SqueezedLiMol   =  Vol_Elely_squeezed*cLi_Xavg;
        SqueezedECMol   =  Vol_Elely_squeezed*c_EC_JR_old;
        Vol_Elely_reservoir_old = Vol_Elely_Tot_old - Vol_Elely_JR_old; 
        LiMol_reservoir_new = Vol_Elely_reservoir_old*c_e_r_old  + SqueezedLiMol;
        ECMol_reservoir_new = Vol_Elely_reservoir_old*c_EC_r_old + SqueezedECMol;
        c_e_r_new= LiMol_reservoir_new / (Vol_Elely_reservoir_old+Vol_Elely_squeezed);
        c_EC_r_new= ECMol_reservoir_new / (Vol_Elely_reservoir_old+Vol_Elely_squeezed);
        Width_new = L_y; 
    else:   # this means Vol_Elely_need >= 0, therefore the folliwing script should works for Vol_Elely_need=0 as well!
        # Important: Calculate added electrolyte based on excessive electrolyte, can be: 1) added as required; 2) added some, but less than required; 3) added 0 
        Vol_Elely_squeezed = 0;  
        if Vol_Elely_Tot_old > Vol_Elely_JR_old:                             # This means Vol_Elely_JR_old = Vol_Pore_tot_old ()
            if Vol_Elely_Tot_old-Vol_Elely_JR_old >= Vol_Elely_need:         # 1) added as required
                Vol_Elely_add     = Vol_Elely_need;  
                Vol_Elely_JR_new  = Vol_Pore_tot_new;  # also equals to 'Vol_Pore_tot_new', or Vol_Pore_tot_old - Vol_Pore_decrease
                Ratio_Dryout = 1.0;
            else:                                                            # 2) added some, but less than required;                                                         
                Vol_Elely_add     = Vol_Elely_Tot_old - Vol_Elely_JR_old;   
                Vol_Elely_JR_new  = Vol_Elely_Tot_new;                       # This means Vol_Elely_JR_new <= Vol_Pore_tot_new
                Ratio_Dryout = Vol_Elely_JR_new/Vol_Pore_tot_new;
        else:                                                                # 3) added 0 
            Vol_Elely_add = 0;
            Vol_Elely_JR_new  = Vol_Elely_Tot_new; 
            Ratio_Dryout = Vol_Elely_JR_new/Vol_Pore_tot_new;

        # Next: start mix electrolyte based on previous defined equation
        # Lithium amount in liquid phase, at initial time point
        TotLi_Elely_JR_Old = Sol["Total lithium in electrolyte [mol]"].entries[-1]; # remember this is only in JR
        # Added lithium and EC amount in the added lithium electrolyte: - 
        AddLiMol   =  Vol_Elely_add*c_e_r_old
        AddECMol   =  Vol_Elely_add*c_EC_r_old
        # Total amount of Li and EC in current electrolyte:
        # Remember Li has two parts, initial and added; EC has three parts, initial, consumed and added
        TotLi_Elely_JR_New   = TotLi_Elely_JR_Old + AddLiMol
        TotECMol_JR   = Vol_Elely_JR_old*c_EC_JR_old - LLINegSEI + AddECMol  # EC:lithium:SEI=2:2:1     for SEI=(CH2OCO2Li)2
        Ratio_CeEC_JR  = TotECMol_JR    /   Vol_Elely_JR_new   / c_EC_JR_old
        Ratio_CeLi_JR  = TotLi_Elely_JR_New    /   TotLi_Elely_JR_Old   /  Ratio_Dryout # Mark, change on 21-11-19
        c_e_r_new  = c_e_r_old;
        c_EC_r_new = c_EC_r_old;
        Width_new   = Ratio_Dryout * L_y;
        
    # Collect the above parameter in Data_Pack to shorten the code  
    Data_Pack   = [
        Vol_EC_consumed, 
        Vol_Elely_need, 
        Test_V, 
        Test_V2, 
        Vol_Elely_add, 
        Vol_Elely_Tot_new, 
        Vol_Elely_JR_new, 
        Vol_Pore_tot_new, 
        Vol_Pore_decrease, 
        c_e_r_new, c_EC_r_new,
        Ratio_Dryout, Ratio_CeEC_JR, 
        Ratio_CeLi_JR,
        Width_new, 
        ]  # 16 in total
    #print('Loss of lithium to SEI', LLINegSEI, 'mol') 
    #print('Loss of lithium to dead lithium plating', LLINegDeadLiP, 'mol') 
    #############################################################################################################################  
    ###################       Step-4 Update parameters here        ##############################################################
    #############################################################################################################################
    Para.update(   {'Bulk solvent concentration [mol.m-3]':  c_EC_JR_old * Ratio_CeEC_JR  })
    Para.update(
        {'EC initial concentration in electrolyte [mol.m-3]':c_EC_JR_old * Ratio_CeEC_JR }, 
        check_already_exists=False) 
    Para.update(   
        {'Ratio of Li-ion concentration change in electrolyte consider solvent consumption':  
        Ratio_CeLi_JR }, check_already_exists=False) 
    Para.update(   {'Current total electrolyte volume in whole cell [m3]':  Vol_Elely_Tot_new  }, check_already_exists=False)
    Para.update(   {'Current total electrolyte volume in jelly roll [m3]':  Vol_Elely_JR_new  }, check_already_exists=False)
    Para.update(   {'Ratio of electrolyte dry out in jelly roll':Ratio_Dryout}, check_already_exists=False)
    Para.update(   {'Electrode width [m]':Width_new})    
    Para.update(   {'Current solvent concentration in the reservoir [mol.m-3]':c_EC_r_new}, check_already_exists=False)     
    Para.update(   {'Current electrolyte concentration in the reservoir [mol.m-3]':c_e_r_new}, check_already_exists=False)             
    return Data_Pack,Para

# Define a function to calculate based on previous solution
def Run_Model_Base_On_Last_Solution( 
    Model  , Sol , Para_update, ModelExperiment, 
    Update_Cycles,Temper_i ,mesh_par):
    # Use Sulzer's method: inplace = false
    # Important line: define new model based on previous solution
    Ratio_CeLi = Para_update["Ratio of Li-ion concentration change in electrolyte consider solvent consumption"]
    dict_short = {}; 
    list_short = [  "Discharge capacity [A.h]", # mark Ruihe Li change for FullyCoupled as this is an old version of PyBaMM
        "Negative particle crack length",                     # add on 220728 for cracks option
        "Positive particle crack length",  
        "Negative electrode active material volume fraction", # add on 220728 for cracks option
        "Positive electrode active material volume fraction", # add on 220728 for cracks option
        "Total negative electrode interfacial current density variable" ,
        "Total positive electrode interfacial current density variable",
        "Negative particle concentration",
        "Positive particle concentration",
        "Negative electrode potential",
        "Positive electrode potential",

        "Negative electrode porosity times concentration",
        "Separator porosity times concentration",
        "Positive electrode porosity times concentration",   

        "Negative electrolyte potential",
        "Separator electrolyte potential",
        "Positive electrolyte potential",

        "Inner SEI thickness",
        "Outer SEI thickness",
        "Inner SEI on cracks thickness",   # add on 220728 for cracks option
        "Outer SEI on cracks thickness",   # add on 220728 for cracks option
        "Lithium plating concentration",
        "Dead lithium concentration", ]
    for list_short_i in list_short:
        dict_short.update( { list_short_i : Sol.last_state[list_short_i].data  }  )
    dict_short["Negative electrode porosity times concentration"] = (
        dict_short["Negative electrode porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    dict_short["Separator porosity times concentration"] = (
        dict_short["Separator porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    dict_short["Positive electrode porosity times concentration"] = (
        dict_short["Positive electrode porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    Model_new = Model.set_initial_conditions_from(dict_short, inplace=False)
    Para_update.update(   {'Ambient temperature [K]':273.15+Temper_i });   # run model at 45 degree C
    
    var = pb.standard_spatial_vars  
    var_pts = {var.x_n: 20,  var.x_s: 10,  var.x_p: 20,  var.r_n: int(mesh_par),  var.r_p: int(mesh_par),  }
    
    Simnew = pb.Simulation(
        Model_new,
        experiment = ModelExperiment, 
        parameter_values=Para_update, 
        solver = pb.CasadiSolver(),
        var_pts = var_pts,
    )
    Sol_new = Simnew.solve(calc_esoh=False,save_at_cycles = Update_Cycles) # save every several cycles, can save RAM greatly
    # print("Solved this model in {}".format(ModelTimer.time()))
    return Model_new, Sol_new
def Run_Model_Base_On_Last_Solution_RPT( 
    Model  , Sol,  Para_update, 
    ModelExperiment ,Update_Cycles, Temper_i,mesh_par):
    # Use Sulzer's method: inplace = false
    Ratio_CeLi = Para_update["Ratio of Li-ion concentration change in electrolyte consider solvent consumption"]
    # print("Model is now using average EC Concentration of:",Para_update['Bulk solvent concentration [mol.m-3]'])
    # print("Ratio of electrolyte dry out in jelly roll is:",Para_update['Ratio of electrolyte dry out in jelly roll'])
    # print("Model is now using an electrode width of:",Para_update['Electrode width [m]'])
    # Important line: define new model based on previous solution
    dict_short = {}; 
    list_short = [  "Discharge capacity [A.h]", # mark Ruihe Li change for FullyCoupled as this is an old version of PyBaMM
        "Negative particle crack length",                     # add on 220728 for cracks option
        "Positive particle crack length",  
        "Negative electrode active material volume fraction", # add on 220728 for cracks option
        "Positive electrode active material volume fraction", # add on 220728 for cracks option
        "Total negative electrode interfacial current density variable" ,
        "Total positive electrode interfacial current density variable",
        "Negative particle concentration",
        "Positive particle concentration",
        "Negative electrode potential",
        "Positive electrode potential",

        "Negative electrode porosity times concentration",
        "Separator porosity times concentration",
        "Positive electrode porosity times concentration",   

        "Negative electrolyte potential",
        "Separator electrolyte potential",
        "Positive electrolyte potential",

        "Inner SEI thickness",
        "Outer SEI thickness",
        "Inner SEI on cracks thickness",   # add on 220728 for cracks option
        "Outer SEI on cracks thickness",   # add on 220728 for cracks option
        "Lithium plating concentration",
        "Dead lithium concentration", ]
    for list_short_i in list_short:
        dict_short.update( { list_short_i : Sol.last_state[list_short_i].data  }  )
    dict_short["Negative electrode porosity times concentration"] = (
        dict_short["Negative electrode porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    dict_short["Separator porosity times concentration"] = (
        dict_short["Separator porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    dict_short["Positive electrode porosity times concentration"] = (
        dict_short["Positive electrode porosity times concentration"] * Ratio_CeLi )# important: update sol here!
    Model_new = Model.set_initial_conditions_from(dict_short, inplace=False)
    Para_update.update(   {'Ambient temperature [K]':273.15+Temper_i });
    
    var = pb.standard_spatial_vars  
    var_pts = {var.x_n: 20,  var.x_s: 10,  var.x_p: 20,  var.r_n: int(mesh_par),  var.r_p: int(mesh_par),  }
    
    Simnew = pb.Simulation(
        Model_new,
        experiment = ModelExperiment, 
        parameter_values=Para_update, 
        solver = pb.CasadiSolver(),
        var_pts = var_pts,
    )
    Sol_new = Simnew.solve(calc_esoh=False,save_at_cycles = Update_Cycles) # save every several cycles, can save RAM greatly
    # print("Solved this model in {}".format(ModelTimer.time()))
    return Model_new, Sol_new

def write_excel_xlsx(path, sheet_name, value):
    import numpy as np
    index = len(value)
    workbook = openpyxl.Workbook()  # 新建工作簿（默认有一个sheet？）
    sheet = workbook.active  # 获得当前活跃的工作页，默认为第一个工作页
    sheet.title = sheet_name  # 给sheet页的title赋值
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))  # 行，列，值 这里是从1开始计数的
    workbook.save(path)  # 一定要保存
    print("Successfully create a excel file")

def GetScan(
    Ratio_excess,cs_Neg_Init,Diff_SEI,R_SEI,Bulk_Sol_Con,
    D_Li_inSEI,c_Li_inte_ref,
    Diff_EC,k_SEI,LAMcr_prop,Crack_rate,
    Couple_SEI_LiP,k_LiP,Temper,MESH_PAR):
    import numpy as np
    TotalScan =(
        len(Ratio_excess)*
        len(cs_Neg_Init)*
        len(Diff_SEI)*
        len(R_SEI)*
        len(Bulk_Sol_Con)*
        len(D_Li_inSEI)*
        len(c_Li_inte_ref)*
        len(Diff_EC)*          # 
        len(k_SEI)*            #
        len(LAMcr_prop)*       #
        len(Crack_rate)*       #
        len(Couple_SEI_LiP)*
        len(k_LiP)*
        len(Temper)*
        len(MESH_PAR) 
        ) ;
    DatePack_scan = np.full([TotalScan,16], 0.0);
    index = 0;
    for Ratio_excess_i in Ratio_excess:
        for cs_Neg_Init_i in cs_Neg_Init:
            for Diff_SEI_i in Diff_SEI:
                for R_SEI_i in R_SEI:
                    for Bulk_Sol_Con_i in Bulk_Sol_Con:
                        for D_Li_inSEI_i in D_Li_inSEI:
                            for c_Li_inte_ref_i in c_Li_inte_ref:
                                for Diff_EC_i in Diff_EC:
                                    for k_SEI_i in k_SEI:
                                        for LAMcr_prop_i in LAMcr_prop:
                                            for Crack_rate_i in Crack_rate:
                                                for Couple_SEI_LiP_i in Couple_SEI_LiP:
                                                    for k_LiP_i in k_LiP:
                                                        for Temper_i in Temper:
                                                            for mesh_par in MESH_PAR:
                                                                index +=  1;
                                                                DatePack_scan[index-1] = [
                                                                    index,
                                                                    Ratio_excess_i,
                                                                    cs_Neg_Init_i,Diff_SEI_i,
                                                                    R_SEI_i,Bulk_Sol_Con_i,
                                                                    D_Li_inSEI_i,c_Li_inte_ref_i,
                                                                    Diff_EC_i,k_SEI_i,           # 
                                                                    LAMcr_prop_i,Crack_rate_i,   #
                                                                    Couple_SEI_LiP_i,k_LiP_i,Temper_i,
                                                                    mesh_par];
    return TotalScan, DatePack_scan

# Run model with electrolyte dry out
def Run_model_w_dry_out(CyclePack, DatePack_scan_i,  BasicPath, Target, book_name_xlsx ):
    V_max = 4.2;        
    V_min = 2.5;   
    Temper_RPT = 25; 
    (Total_Cycles, Cycle_bt_RPT, Update_Cycles ) = CyclePack;
    Small_Loop =  int(Cycle_bt_RPT/Update_Cycles);   
    SaveTimes = int(Total_Cycles/Cycle_bt_RPT);   
    RPT_Cycles = 2; 
    #index 
    ModelTimer = pb.Timer()
    # unpack data:
    [index_xlsx,
    Ratio_excess_i,
    cs_Neg_Init_i,
    Diff_SEI_i,
    R_SEI_i,
    Bulk_Sol_Con_i,
    D_Li_inSEI_i,
    c_Li_inte_ref_i,
    Diff_EC_i,      # 
    k_SEI_i,        #
    LAMcr_prop_i,   #
    Crack_rate_i,   #
    Couple_SEI_LiP_i,
    k_LiP_i,
    Temper_i,
    mesh_par,
    ] = DatePack_scan_i;
    count_i = int(index_xlsx);
    #
    Experiment_Long   = pb.Experiment(  
        [(f"Discharge at 1 C until {V_min} V", 
        f"Charge at 0.3 C until {V_max} V", 
        f"Hold at {V_max} V until C/100"),  ] * Update_Cycles  )  
    Experiment_RPT    = pb.Experiment( 
        [ (f"Discharge at 0.1C until {V_min} V",  
        "Rest for 1 hours",  
        f"Charge at 0.1C until {V_max} V" ) ] * RPT_Cycles ) 
    Experiment_Breakin= pb.Experiment( 
        [ ( f"Discharge at 0.1C until {V_min} V",  
        "Rest for 1 hours",  
        f"Charge at 0.1C until {V_max} V" )  ] *2 )
    
    sheet_name_xlsx = 'Results';
    ################ Important: index definition #################################
    index = list(np.arange(1,SaveTimes+1)*(Small_Loop+RPT_Cycles)-1); #index.insert(0,0)  # get all last ageing cycle before RPT, plus the first RPT test
    index2= list(np.arange(0,SaveTimes+1)*(Small_Loop+RPT_Cycles));         # get all RPT test
    cycles = np.arange(0,SaveTimes+1)*Cycle_bt_RPT; 
    cycles3 = list(np.arange(1,SaveTimes+1)*(Cycle_bt_RPT));
    try:             
        for i in range(0,1):    ######################   update parameter and run first RPT
            ChemistryChen=pb.parameter_sets.OKane2022
            ChemistryChen["electrolyte"] = "lipf6_Valoen2005"   
            Para_0=pb.ParameterValues(chemistry=ChemistryChen)
            # DFN parameter
            Para_0.update({"Upper voltage cut-off [V]": 4.21})
            Para_0.update({"Lower voltage cut-off [V]": 2.49})
            Para_0.update({"Negative electrode diffusivity [m2.s-1]": 2e-13})
            Para_0.update({"Positive electrode diffusivity [m2.s-1]": 1e-13})
            Para_0.update({
                'Initial concentration in negative electrode [mol.m-3]':
                cs_Neg_Init_i })
            # Solvent consumption sub-model
            Para_0.update(
                {'Initial electrolyte excessive amount ratio':Ratio_excess_i}, 
                check_already_exists=False)   # 
            Para_0.update(
                {'Current solvent concentration in the reservoir [mol.m-3]':
                Bulk_Sol_Con_i}, 
                check_already_exists=False)     
            Para_0.update(
                {'Current electrolyte concentration in the reservoir [mol.m-3]':
                1000.0}, 
                check_already_exists=False)           
            Para_0.update(
                {'Ratio of Li-ion concentration change in electrolyte consider solvent consumption':  
                1.0 }, check_already_exists=False)
            # general SEI
            Para_0.update({'SEI resistivity [Ohm.m]':R_SEI_i}) ;
            # solvent-diffusion limited
            Para_0.update({'Outer SEI solvent diffusivity [m2.s-1]':Diff_SEI_i});
            Para_0.update({'Bulk solvent concentration [mol.m-3]':Bulk_Sol_Con_i});
            # interstitial-diffusion limited
            Para_0.update({
                'Inner SEI lithium interstitial diffusivity [m2.s-1]':
                D_Li_inSEI_i})       
            Para_0.update({
                'Lithium interstitial reference concentration [mol.m-3]':
                c_Li_inte_ref_i}) 
            # ec-reaction limited
            Para_0.update({'EC diffusivity [m2.s-1]':Diff_EC_i}) 
            Para_0.update({'SEI kinetic rate constant [m.s-1]':k_SEI_i}) 
            Para_0.update({
                'EC initial concentration in electrolyte [mol.m-3]':
                Bulk_Sol_Con_i}) 
            # LiP and coupling with SEI:
            Para_0.update({"Dead lithium decay constant [s-1]": Couple_SEI_LiP_i})   
            Para_0.update({'Lithium plating kinetic rate constant [m.s-1]':k_LiP_i})  
            # Crack model
            Para_0.update({
                "Negative electrode LAM constant proportional term [s-1]": 
                LAMcr_prop_i})
            Para_0.update({
                "Positive electrode LAM constant proportional term [s-1]": 
                LAMcr_prop_i})  
            # make it simple for now, but may want to have T dependency in the future
            Para_0.update({"Negative electrode cracking rate": Crack_rate_i})
            Para_0.update({"Positive electrode cracking rate": Crack_rate_i})
            Para_0.update({"Negative electrode volume change": 0.0})
            Para_0.update({"Positive electrode volume change": 0.0})

            Model_0 = pb.lithium_ion.DFN(     
                options={
                    "particle": "Fickian diffusion",          
                    "SEI":"solvent-diffusion limited",   
                    "SEI on cracks":"true",  
                    "SEI film resistance":"distributed",          
                    "SEI porosity change":"true",      
                    "particle mechanics":"swelling and cracking",  
                    "loss of active material":"stress-driven", 
                    "lithium plating":"partially reversible"      } ) #
            
            var = pb.standard_spatial_vars  
            var_pts = {
                var.x_n: 20,  
                var.x_s: 10,  
                var.x_p: 20,  
                var.r_n: int(mesh_par),  
                var.r_p: int(mesh_par),  }
            
            Sim_0    = pb.Simulation(
                Model_0,        experiment = Experiment_Breakin,
                parameter_values = Para_0,
                solver = pb.CasadiSolver(),
                var_pts=var_pts,) #mode="safe"
            Sol_0    = Sim_0.solve(calc_esoh=False)

        for i in range(0,1):    ######################   initialize for post-processing before running subsequent model   
            Vol_Elely_Tot_All  = [];       Vol_Elely_JR_All =[];         Vol_Pore_tot_All =[];
            Ratio_CeEC_All     = [];       Ratio_CeLi_All = [];          Ratio_Dryout_All =[];
            Vol_EC_consumed_All= [];       Vol_Elely_need_All = [];      Width_all        =[];
            Vol_Elely_add_All  = [];       Vol_Pore_decrease_All =[];
            Para_All           = [];       Model_All = [];               Sol_All_i =[]  ;    
            Test_V_All=[];    Test_V2_All=[]; c_e_r_new_All=[]; c_EC_r_new_All=[]; 
            # for RPT:
            Cap_RPT_Dry_All=[]; CapLoss_LiP_Dry_All=[]; R_Local_ECM_Dry_All =[]; 
            NegSOC_CDsta_Dry_All =[];  NegSOC_CDend_Dry_All =[];  PosSOC_CDsta_Dry_All =[]; PosSOC_CDend_Dry_All =[];
            CapLoss_SEI_Dry_All=[];   CapLoss_SEIcr_Dry_All=[];      Cap_Pos_Dry_All=[];          Cap_Neg_Dry_All=[]; 
            t_SEIcr_Xavgtot_Dry_All = []; Rough_Neg_Xavg_Dry_All =[];

            T_0                  =  Para_0['Initial temperature [K]']
            Porosity_Neg_0       =  Para_0['Negative electrode porosity']  
            Porosity_Pos_0       =  Para_0['Positive electrode porosity']  
            Porosity_Sep_0       =  Para_0['Separator porosity']  
            cs_Neg_Max           =  Para_0["Maximum concentration in negative electrode [mol.m-3]"];
            L_p                  =  Para_0["Positive electrode thickness [m]"]
            L_n                  =  Para_0["Negative electrode thickness [m]"]
            L_s                  =  Para_0["Separator thickness [m]"]
            L_y                  =  Para_0["Electrode width [m]"]
            Para_0.update({'Initial Electrode width [m]':L_y}, check_already_exists=False)
            L_y_0                =  Para_0["Initial Electrode width [m]"]
            L_z                  =  Para_0["Electrode height [m]"]
            Para_0.update({'Initial Electrode height [m]':L_z}, check_already_exists=False)
            L_z_0                =  Para_0["Initial Electrode height [m]"]
            Int_ElelyExces_Ratio =  Para_0["Initial electrolyte excessive amount ratio"] 
            Vol_Elely_Tot        =  ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  * L_y_0 * L_z_0 * Int_ElelyExces_Ratio # Set initial electrolyte amount [L] 
            Vol_Elely_JR         =  ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  * L_y_0 * L_z_0
            Vol_Pore_tot         =  ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  * L_y_0 * L_z_0
            Ratio_CeEC           =  1.0; Ratio_CeLi =  1.0  ;Ratio_Dryout         =  1.0
            Vol_EC_consumed      =  0;Vol_Elely_need       =  0;Vol_Elely_add        =  0
            Vol_Pore_decrease    =  0;Test_V2 = 0; Test_V = 0;
            print('Initial electrolyte amount is ', Vol_Elely_Tot*1e6, 'mL') 
            Para_0.update({'Current total electrolyte volume in jelly roll [m3]':Vol_Elely_JR}, check_already_exists=False)
            Para_0.update({'Current total electrolyte volume in whole cell [m3]':Vol_Elely_Tot}, check_already_exists=False)   
            Vol_Elely_Tot_All.append(Vol_Elely_Tot*1e6);            Vol_Elely_JR_All.append(Vol_Elely_JR*1e6);     Vol_Pore_tot_All.append(Vol_Pore_tot*1e6);           
            Ratio_CeEC_All.append(Ratio_CeEC);                      Ratio_CeLi_All.append(Ratio_CeLi);             Ratio_Dryout_All.append(Ratio_Dryout);
            Vol_EC_consumed_All.append(Vol_EC_consumed*1e6);        Vol_Elely_need_All.append(Vol_Elely_need*1e6);     Width_all.append(L_y_0);
            Vol_Elely_add_All.append(Vol_Elely_add*1e6);            Vol_Pore_decrease_All.append(Vol_Pore_decrease*1e6);
            Test_V_All.append(Test_V*1e6); Test_V2_All.append(Test_V2*1e6); 
            c_e_r_new_All.append(1000.0); c_EC_r_new_All.append(Bulk_Sol_Con_i); 

        for i in range(0,1):    #################### post process for the first RPT cycle    ############################
            Cap_RPT_Dry_All.append    (
                Sol_0.cycles[-1].steps[0]["Discharge capacity [A.h]"].entries[-1] 
                - Sol_0.cycles[-1].steps[0]["Discharge capacity [A.h]"].entries[0])
                
            CapLoss_LiP_Dry_All.append(
                Sol_0.cycles[-1].steps[0]["Loss of capacity to lithium plating [A.h]"].entries[-1])
            CapLoss_SEI_Dry_All.append(Sol_0.cycles[-1].steps[0]["Loss of capacity to SEI [A.h]"].entries[-1])
            CapLoss_SEIcr_Dry_All.append(Sol_0.cycles[-1].steps[0]["Loss of capacity to SEI on cracks [A.h]"].entries[-1])
            t_SEIcr_Xavgtot_Dry_All.append(Sol_0.cycles[-1].steps[0]["X-averaged total SEI on cracks thickness [m]"].entries[-1])
            Rough_Neg_Xavg_Dry_All.append(
                Sol_0.cycles[-1].steps[0]["X-averaged negative electrode roughness ratio"].entries[-1])
            R_Local_ECM_Dry_All.append(Sol_0.cycles[-1].steps[0]["Local ECM resistance [Ohm]"].entries[-1])

            NegSOC_CDsta_Dry_All .append(Sol_0.cycles[-1].steps[0]["Negative electrode SOC"].entries[0]);  
            NegSOC_CDend_Dry_All .append(Sol_0.cycles[-1].steps[0]["Negative electrode SOC"].entries[-1]);  
            PosSOC_CDsta_Dry_All .append(Sol_0.cycles[-1].steps[0]["Positive electrode SOC"].entries[0]); 
            PosSOC_CDend_Dry_All .append(Sol_0.cycles[-1].steps[0]["Positive electrode SOC"].entries[-1]);  

            Cap_Pos_Dry_All.append    (Sol_0.cycles[-1].steps[0]["Positive electrode capacity [A.h]"].entries[-1])
            Cap_Neg_Dry_All.append    (Sol_0.cycles[-1].steps[0]["Negative electrode capacity [A.h]"].entries[-1])    

            Epsilon_CCend_Dry_First = Sol_0.cycles[-1].steps[-1]["Porosity"].entries[:,-1]
            j_Neg_Int_CCend_Dry_First = Sol_0.cycles[-1].steps[-1]["Negative electrode interfacial current density [A.m-2]"].entries[:,-1] 
            Eta_Elely_CCend_Dry_First = Sol_0.cycles[-1].steps[-1]["Electrolyte potential [V]"].entries[:,-1] 
            c_Elely_CCend_Dry_First   = Sol_0.cycles[-1].steps[-1]["Electrolyte concentration [mol.m-3]"].entries[:,-1] 
            Eta_Neg_rec_CCend_Dry_First = Sol_0.cycles[-1].steps[-1]["Negative electrode reaction overpotential [V]"].entries[:,-1] 
            c_s_Neg_Surf_CCend_Dry_First = Sol_0.cycles[-1].steps[-1]["Negative particle surface concentration [mol.m-3]"].entries[:,-1] 
            Rough_Neg_CCend_Dry_First = Sol_0.cycles[-1].steps[-1]["Negative electrode roughness ratio"].entries[:,-1] 
            t_SEIcr_Neg_CCend_Dry_First = Sol_0.cycles[-1].steps[-1]["Total SEI on cracks thickness [m]"].entries[:,-1] 
            
            Epsilon_CDend_Dry_First = Sol_0.cycles[-1].steps[0]["Porosity"].entries[:,-1]
            j_Neg_Int_CDend_Dry_First = Sol_0.cycles[-1].steps[0]["Negative electrode interfacial current density [A.m-2]"].entries[:,-1] 
            Eta_Elely_CDend_Dry_First = Sol_0.cycles[-1].steps[0]["Electrolyte potential [V]"].entries[:,-1] 
            c_Elely_CDend_Dry_First   = Sol_0.cycles[-1].steps[0]["Electrolyte concentration [mol.m-3]"].entries[:,-1] 
            Eta_Neg_rec_CDend_Dry_First = Sol_0.cycles[-1].steps[0]["Negative electrode reaction overpotential [V]"].entries[:,-1] 
            c_s_Neg_Surf_CDend_Dry_First = Sol_0.cycles[-1].steps[0]["Negative particle surface concentration [mol.m-3]"].entries[:,-1] 
            Rough_Neg_CDend_Dry_First = Sol_0.cycles[-1].steps[0]["Negative electrode roughness ratio"].entries[:,-1] 
            t_SEIcr_Neg_CDend_Dry_First = Sol_0.cycles[-1].steps[0]["Total SEI on cracks thickness [m]"].entries[:,-1] 
        
        x_n = Sol_0.cycles[0].steps[0]["x_n [m]"].entries[:,-1]; x = Sol_0.cycles[0].steps[0]["x [m]"].entries[:,-1];   # Get location variable, can be any single solution

        # Para_All.append(Para_0);                                Model_All.append(Model_0);    Sol_All_i.append(Sol_0); 
        Para_0_Dry_old = Para_0;     Model_Dry_old = Model_0  ; Sol_Dry_old = Sol_0    ;   del Model_0,Sol_0,Para_0
        #Step3: Write a big loop to finish the long experiment,    
        Index_update_all     =[]; cycle_count =0; Index_update_all.append(cycle_count);
        k=0; 
        while k < SaveTimes:    # biggest loop, 
            i=0;     cap_i = 0;
            while i < Small_Loop:
                # run ageing cycles, and update parameters (have another version not update)
                Data_Pack,Paraupdate   = Cal_new_con_Update (  Sol_Dry_old,   Para_0_Dry_old )
                [
                    Vol_EC_consumed, 
                    Vol_Elely_need, 
                    Test_V, 
                    Test_V2, 
                    Vol_Elely_add, 
                    Vol_Elely_Tot_new, 
                    Vol_Elely_JR_new, 
                    Vol_Pore_tot_new, 
                    Vol_Pore_decrease, 
                    c_e_r_new, c_EC_r_new,
                    Ratio_Dryout, Ratio_CeEC_JR, 
                    Ratio_CeLi_JR,
                    Width_new, ]= Data_Pack;
                # Append single object to All object     
                Vol_Elely_Tot_All.append(Vol_Elely_Tot_new*1e6);    Vol_Elely_JR_All.append(Vol_Elely_JR_new*1e6);     Vol_Pore_tot_All.append(Vol_Pore_tot_new*1e6);           
                Ratio_CeEC_All.append(Ratio_CeEC_JR);               Ratio_CeLi_All.append(Ratio_CeLi_JR);                 Ratio_Dryout_All.append(Ratio_Dryout);
                Vol_EC_consumed_All.append(Vol_EC_consumed*1e6);    Vol_Elely_need_All.append(Vol_Elely_need*1e6);     Width_all.append(Width_new);
                Vol_Elely_add_All.append(Vol_Elely_add*1e6);        Vol_Pore_decrease_All.append(Vol_Pore_decrease*1e6);
                Test_V_All.append(Test_V*1e6); Test_V2_All.append(Test_V2*1e6); 
                c_e_r_new_All.append(c_e_r_new); c_EC_r_new_All.append(c_EC_r_new); 

                # 3rd: run model based on new parameter and last updated solution Model  , Sol, Ratio_CeLi, Para_update, ModelExperiment, SaveAs_Cycles
                Model_Dry_i, Sol_Dry_i   = Run_Model_Base_On_Last_Solution( 
                    Model_Dry_old  , Sol_Dry_old ,  
                    Paraupdate ,Experiment_Long, Update_Cycles,Temper_i,mesh_par )
                Para_0_Dry_old = Paraupdate;       Model_Dry_old = Model_Dry_i;      Sol_Dry_old = Sol_Dry_i;   
                del Paraupdate,Model_Dry_i,Sol_Dry_i
                i += 1;   cycle_count +=  Update_Cycles; 
                Index_update_all.append(cycle_count);

            # run RPT, and also update parameters (otherwise will have problems)
            Data_Pack , Paraupdate  = Cal_new_con_Update (  Sol_Dry_old,   Para_0_Dry_old   )
            [
                    Vol_EC_consumed, 
                    Vol_Elely_need, 
                    Test_V, 
                    Test_V2, 
                    Vol_Elely_add, 
                    Vol_Elely_Tot_new, 
                    Vol_Elely_JR_new, 
                    Vol_Pore_tot_new, 
                    Vol_Pore_decrease, 
                    c_e_r_new, c_EC_r_new,
                    Ratio_Dryout, Ratio_CeEC_JR, 
                    Ratio_CeLi_JR,
                    Width_new, ]= Data_Pack;
            # Append single object to All object     
            Vol_Elely_Tot_All.append(Vol_Elely_Tot_new*1e6);    Vol_Elely_JR_All.append(Vol_Elely_JR_new*1e6);     Vol_Pore_tot_All.append(Vol_Pore_tot_new*1e6);           
            Ratio_CeEC_All.append(Ratio_CeEC_JR);                  Ratio_CeLi_All.append(Ratio_CeLi_JR);                 Ratio_Dryout_All.append(Ratio_Dryout);
            Vol_EC_consumed_All.append(Vol_EC_consumed*1e6);    Vol_Elely_need_All.append(Vol_Elely_need*1e6);     Width_all.append(Width_new);
            Vol_Elely_add_All.append(Vol_Elely_add*1e6);        Vol_Pore_decrease_All.append(Vol_Pore_decrease*1e6); 
            Test_V_All.append(Test_V*1e6); Test_V2_All.append(Test_V2*1e6); 
            c_e_r_new_All.append(c_e_r_new); c_EC_r_new_All.append(c_EC_r_new); 

            Index_update_all.append(cycle_count);

            Model_Dry_i, Sol_Dry_i  = Run_Model_Base_On_Last_Solution_RPT( 
                Model_Dry_old  , Sol_Dry_old ,   
                Paraupdate,      Experiment_RPT, RPT_Cycles, Temper_RPT ,mesh_par )     
            for ii in range(0,1):     # add later RPT cycle 
                Cap_RPT_Dry_All.append    (
                    Sol_Dry_i.cycles[-1].steps[0]["Discharge capacity [A.h]"].entries[-1] 
                    - 
                    Sol_Dry_i.cycles[-1].steps[0]["Discharge capacity [A.h]"].entries[0])
                CapLoss_LiP_Dry_All.append(Sol_Dry_i.cycles[-1].steps[0]["Loss of capacity to lithium plating [A.h]"].entries[-1])
                CapLoss_SEI_Dry_All.append(Sol_Dry_i.cycles[-1].steps[0]["Loss of capacity to SEI [A.h]"].entries[-1])
                CapLoss_SEIcr_Dry_All.append(Sol_Dry_i.cycles[-1].steps[0]["Loss of capacity to SEI on cracks [A.h]"].entries[-1])
                t_SEIcr_Xavgtot_Dry_All.append(Sol_Dry_i.cycles[-1].steps[0]["X-averaged total SEI on cracks thickness [m]"].entries[-1])
                Rough_Neg_Xavg_Dry_All.append(Sol_Dry_i.cycles[-1].steps[0]["X-averaged negative electrode roughness ratio"].entries[-1])
                R_Local_ECM_Dry_All.append(Sol_Dry_i.cycles[-1].steps[0]["Local ECM resistance [Ohm]"].entries[-1])
                NegSOC_CDsta_Dry_All .append(Sol_Dry_i.cycles[-1].steps[0]["Negative electrode SOC"].entries[0]);  
                NegSOC_CDend_Dry_All .append(Sol_Dry_i.cycles[-1].steps[0]["Negative electrode SOC"].entries[-1]);  
                PosSOC_CDsta_Dry_All .append(Sol_Dry_i.cycles[-1].steps[0]["Positive electrode SOC"].entries[0]); 
                PosSOC_CDend_Dry_All .append(Sol_Dry_i.cycles[-1].steps[0]["Positive electrode SOC"].entries[-1]);  
                Cap_Pos_Dry_All.append    (Sol_Dry_i.cycles[-1].steps[0]["Positive electrode capacity [A.h]"].entries[-1])
                Cap_Neg_Dry_All.append    (Sol_Dry_i.cycles[-1].steps[0]["Negative electrode capacity [A.h]"].entries[-1])      
            if k == SaveTimes-1:  # Last RPT cycle
                Epsilon_CCend_Dry_Last =  Sol_Dry_i.cycles[-1].steps[-1]["Porosity"].entries[:,-1]
                j_Neg_Int_CCend_Dry_Last = Sol_Dry_i.cycles[-1].steps[-1]["Negative electrode interfacial current density [A.m-2]"].entries[:,-1] 
                Eta_Elely_CCend_Dry_Last = Sol_Dry_i.cycles[-1].steps[-1]["Electrolyte potential [V]"].entries[:,-1] 
                c_Elely_CCend_Dry_Last   = Sol_Dry_i.cycles[-1].steps[-1]["Electrolyte concentration [mol.m-3]"].entries[:,-1] 
                Eta_Neg_rec_CCend_Dry_Last = Sol_Dry_i.cycles[-1].steps[-1]["Negative electrode reaction overpotential [V]"].entries[:,-1] 
                c_s_Neg_Surf_CCend_Dry_Last = Sol_Dry_i.cycles[-1].steps[-1]["Negative particle surface concentration [mol.m-3]"].entries[:,-1] 
                Rough_Neg_CCend_Dry_Last = Sol_Dry_i.cycles[-1].steps[-1]["Negative electrode roughness ratio"].entries[:,-1] 
                t_SEIcr_Neg_CCend_Dry_Last = Sol_Dry_i.cycles[-1].steps[-1]["Total SEI on cracks thickness [m]"].entries[:,-1] 
                
                Epsilon_CDend_Dry_Last =  Sol_Dry_i.cycles[-1].steps[0]["Porosity"].entries[:,-1]
                j_Neg_Int_CDend_Dry_Last = Sol_Dry_i.cycles[-1].steps[0]["Negative electrode interfacial current density [A.m-2]"].entries[:,-1] 
                Eta_Elely_CDend_Dry_Last = Sol_Dry_i.cycles[-1].steps[0]["Electrolyte potential [V]"].entries[:,-1] 
                c_Elely_CDend_Dry_Last   = Sol_Dry_i.cycles[-1].steps[0]["Electrolyte concentration [mol.m-3]"].entries[:,-1] 
                Eta_Neg_rec_CDend_Dry_Last = Sol_Dry_i.cycles[-1].steps[0]["Negative electrode reaction overpotential [V]"].entries[:,-1] 
                c_s_Neg_Surf_CDend_Dry_Last = Sol_Dry_i.cycles[-1].steps[0]["Negative particle surface concentration [mol.m-3]"].entries[:,-1] 
                Rough_Neg_CDend_Dry_Last = Sol_Dry_i.cycles[-1].steps[0]["Negative electrode roughness ratio"].entries[:,-1] 
                t_SEIcr_Neg_CDend_Dry_Last = Sol_Dry_i.cycles[-1].steps[0]["Total SEI on cracks thickness [m]"].entries[:,-1] 

            Para_0_Dry_old = Paraupdate;    Model_Dry_old = Model_Dry_i  ;                 Sol_Dry_old = Sol_Dry_i    ;   del Paraupdate,Model_Dry_i,Sol_Dry_i
            k += 1;    cycles2 =Index_update_all; 
    except:
        data = openpyxl.load_workbook(BasicPath + Target + book_name_xlsx)   
        str_error = traceback.format_exc()      
        table = data.get_sheet_by_name('Results')
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得列数
        values = [
            [count_i, Ratio_excess_i,cs_Neg_Init_i, Diff_SEI_i, R_SEI_i, 
            Bulk_Sol_Con_i,D_Li_inSEI_i, 
            c_Li_inte_ref_i,
            Diff_EC_i,      # 
            k_SEI_i,        #
            LAMcr_prop_i,   #
            Crack_rate_i,   #
            Couple_SEI_LiP_i,
            k_LiP_i,Temper_i,mesh_par,
            "nan","nan",
            "nan","nan", 
            "nan","nan",
            "nan","nan",
            "nan",str_error],
            ]
        for i in range(1, len(values)+1):
            for j in range(1, len(values[i-1])+1):
                table.cell(nrows+i, j).value = values[i-1][j-1]     
        data.save(BasicPath + Target + book_name_xlsx)
        print(str_error)
        print("Fail in {}".format(ModelTimer.time())) 
    else:
        # Finish everything, try to write into excel with real results or 
        data = openpyxl.load_workbook(BasicPath + Target+ book_name_xlsx)         
        table = data.get_sheet_by_name('Results')
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得列数
        Cap_RPT_Dry_Loss = Cap_RPT_Dry_All[0] - Cap_RPT_Dry_All[-1]
        Cap_Pos_Dry_All_loss = Cap_Pos_Dry_All[0] - Cap_Pos_Dry_All[-1]
        Cap_Neg_Dry_All_loss = Cap_Neg_Dry_All[0] - Cap_Neg_Dry_All[-1]
        values = [
            [count_i, Ratio_excess_i,
            cs_Neg_Init_i, Diff_SEI_i, R_SEI_i, 
            Bulk_Sol_Con_i,D_Li_inSEI_i, 
            c_Li_inte_ref_i,
            Diff_EC_i,      # 
            k_SEI_i,        #
            LAMcr_prop_i,   #
            Crack_rate_i,   #
            Couple_SEI_LiP_i,
            k_LiP_i,Temper_i,mesh_par,
            Cap_RPT_Dry_Loss,
            CapLoss_LiP_Dry_All[-1],
            CapLoss_SEI_Dry_All[-1],
            CapLoss_SEIcr_Dry_All[-1],
            Cap_Neg_Dry_All_loss,
            Cap_Pos_Dry_All_loss,
            Vol_Elely_Tot_All[-1], 
            Vol_Elely_JR_All[-1],
            Width_all[-1]],
            ]
        for i in range(1, len(values)+1):
            for j in range(1, len(values[i-1])+1):
                table.cell(nrows+i, j).value = values[i-1][j-1]     
        data.save(BasicPath + Target+ book_name_xlsx)
        print("Succeed in {}".format(ModelTimer.time()))
        print('This is the ', count_i, ' scan')
        if not os.path.exists(BasicPath + Target + str(count_i)):
            os.mkdir(BasicPath + Target + str(count_i) );
        
        # Newly add (220517): save plots, not just a single line in excel file:     
        # Fig. 1 how much capacity is loss, how much is due to SEI and LiP?
        for mm in range(0,1):
            fs=17;Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(cycles, Cap_RPT_Dry_All,     '-o', label="Scan=" + str(count_i) )
            axs[1].plot(cycles, CapLoss_LiP_Dry_All,'-o', label="LiP - Scan=" + str(count_i) )
            axs[1].plot(cycles, CapLoss_SEI_Dry_All ,'-o', label="SEI - Scan" + str(count_i) )
            axs[1].plot(cycles, CapLoss_SEIcr_Dry_All ,'-o', label="sei-on-cracks - Scan" + str(count_i) )
            for i in range(0,Num_subplot):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[i].set_ylabel("Capacity [A.h]",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
            axs[0].set_title("Discharge capacity",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Capacity loss to LiP and SEI",   fontdict={'family':'DejaVu Sans','size':fs+1})
            plt.savefig(BasicPath + Target+ str(count_i)+"/Cap-LLI.png", dpi=100)

            fs=17;Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(cycles, t_SEIcr_Xavgtot_Dry_All,     '-o', label="Scan=" + str(count_i) )
            axs[1].plot(cycles, Rough_Neg_Xavg_Dry_All,'-o', label="Scan=" + str(count_i) )
            axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_ylabel("Roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs})
            for i in range(0,Num_subplot):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
            axs[0].set_title("X-avg tot Neg SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("X-avg Neg roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
            plt.savefig(BasicPath + Target+ str(count_i)+"/Cracks related.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[1].plot(cycles, Cap_Neg_Dry_All[0]-Cap_Neg_Dry_All,'-o',label="Scan=" + str(count_i))
            axs[1].plot(cycles, Cap_Pos_Dry_All[0]-Cap_Pos_Dry_All   ,'-^',label="Scan=" + str(count_i))
            axs[0].plot(cycles, R_Local_ECM_Dry_All,                  '-o',label="Scan="+ str(count_i))
            for i in range(0,1):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
            axs[0].set_ylabel("ECM resistance [Ohm]",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[0].set_title("Local ECM resistance",   fontdict={'family':'DejaVu Sans','size':fs+1})
            for i in range(1,2):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[i].set_ylabel("Capacity [A.h]",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
            axs[1].set_title("LAM of Neg and Pos",   fontdict={'family':'DejaVu Sans','size':fs+1})
            plt.savefig(BasicPath + Target+ str(count_i)+"/LAM-IR.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(cycles, PosSOC_CDsta_Dry_All,'-o',label="Start" )
            axs[0].plot(cycles, PosSOC_CDend_Dry_All,'-^',label="End" )
            axs[1].plot(cycles, NegSOC_CDsta_Dry_All,'-o',label="Start" )
            axs[1].plot(cycles, NegSOC_CDend_Dry_All,'-^',label="End" )
            for i in range(0,2):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[i].set_ylabel("SOC",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)     
            axs[0].set_title("Neg SOC range (Dis)",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Pos SOC range (Dis)",   fontdict={'family':'DejaVu Sans','size':fs+1})
            plt.savefig(BasicPath + Target+ str(count_i)+"/SOC_RPT_dis.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(x/np.max(x), Epsilon_CDend_Dry_First,'-o',label="First")
            axs[0].plot(x/np.max(x), Epsilon_CDend_Dry_Last,'-^',label="Last"  )
            axs[1].plot(x_n/np.max(x), Eta_Neg_rec_CDend_Dry_First,'-o',label="First" )
            axs[1].plot(x_n/np.max(x), Eta_Neg_rec_CDend_Dry_Last,'-^',label="Last" )
            axs[0].set_xlabel("Dimensionless Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_xlabel("Dimensionless Neg thickness",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[0].set_title("Porosity",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Neg electrode reaction overpotential",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[0].set_ylabel("Porosity",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_ylabel("Overpotential [V]",   fontdict={'family':'DejaVu Sans','size':fs})
            for i in range(0,2):
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
            plt.savefig(BasicPath + Target+ str(count_i)+"/Por Neg_S_eta.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(x_n/np.max(x), t_SEIcr_Neg_CDend_Dry_First,'-o',label="First")
            axs[0].plot(x_n/np.max(x), t_SEIcr_Neg_CDend_Dry_Last,'-^',label="Last"  )
            axs[1].plot(x_n/np.max(x), Rough_Neg_CDend_Dry_First,'-o',label="First" )
            axs[1].plot(x_n/np.max(x), Rough_Neg_CDend_Dry_Last,'-^',label="Last" )
            axs[0].set_title("Tot Neg SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Neg roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_ylabel("Roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs})
            for i in range(0,2):
                axs[i].set_xlabel("Dimensionless Neg thickness",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
            plt.savefig(BasicPath + Target+ str(count_i)+"/Cracks related spatial.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(x/np.max(x), c_Elely_CDend_Dry_First,'-o',label="First")
            axs[0].plot(x/np.max(x), c_Elely_CDend_Dry_Last,'-^',label="Last"  )
            axs[1].plot(x/np.max(x), Eta_Elely_CDend_Dry_First,'-o',label="First" )
            axs[1].plot(x/np.max(x), Eta_Elely_CDend_Dry_Last,'-^',label="Last" )
            axs[0].set_title("Electrolyte concentration",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Electrolyte potential",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[0].set_ylabel("Concentration [mol.m-3]",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_ylabel("Potential [V]",   fontdict={'family':'DejaVu Sans','size':fs})
            for i in range(0,2):
                axs[i].set_xlabel("Dimensionless Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
            plt.savefig(BasicPath + Target+ str(count_i)+"/Electrolyte concentration and potential.png", dpi=100)
        
            Num_subplot = 3;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(18,4.8),tight_layout=True)
            axs[0].plot(cycles2, Vol_EC_consumed_All,'-o',label="EC consumed")
            axs[0].plot(cycles2, Vol_Elely_need_All,'-.',label="Elely needed")
            axs[0].plot(cycles2, Vol_Elely_add_All,'-s',label="Elely added")
            axs[0].plot(cycles2, Vol_Pore_decrease_All,'--',label="Pore decreased")
            axs[1].plot(cycles2, Vol_Elely_Tot_All,     '-o', label="Total electrolyte in cell" )
            axs[1].plot(cycles2, Vol_Elely_JR_All,     '--', label="Total electrolyte in JR" )
            axs[1].plot(cycles2, Vol_Pore_tot_All,     '-s', label="Total pore in JR" )
            axs[2].plot(cycles2, Ratio_Dryout_All,     '-s', label="Dry out ratio" )
            axs[2].set_ylabel("Ratio",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[2].set_xlabel("Cycle number",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[2].set_title("Dry out ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
            for i in range(0,2):
                axs[i].set_xlabel("Cycle number",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[i].set_title("Volume",   fontdict={'family':'DejaVu Sans','size':fs+1})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
            plt.savefig(BasicPath + Target+ str(count_i)+"/Volume_total.png", dpi=100)
        
 
        # Newly add (220706): save data, not just a single line in excel file:
        
        for mm in range(0,1):
            CeEC_All =np.full(np.size(Ratio_CeEC_All),Bulk_Sol_Con_i); 
            for i in range(1,np.size(Ratio_CeEC_All)):
                for k in range(0,i):
                    CeEC_All[i] *= Ratio_CeEC_All[k];        
        mdic = {
            "x":x,
            "x_n":x_n,
            "cycles": cycles,
            "cycles2":cycles2,
            "Cap_RPT_Dry_All": Cap_RPT_Dry_All,
            "Real_SaveAs_Cycles": CapLoss_LiP_Dry_All,
            "CapLoss_SEI_Dry_All": CapLoss_SEI_Dry_All,
            "CapLoss_SEIcr_Dry_All": CapLoss_SEIcr_Dry_All,# ,   
            "t_SEIcr_Xavgtot_Dry_All":t_SEIcr_Xavgtot_Dry_All,
            "Rough_Neg_Xavg_Dry_All":Rough_Neg_Xavg_Dry_All,
            "SaveTimes": SaveTimes,
            "Cap_Neg_Dry_All": Cap_Neg_Dry_All,
            "Cap_Pos_Dry_All": Cap_Pos_Dry_All,
            "R_Local_ECM_Dry_All": R_Local_ECM_Dry_All,
            "NegSOC_CDsta_Dry_All"  : NegSOC_CDsta_Dry_All,
            "NegSOC_CDend_Dry_All"  : NegSOC_CDend_Dry_All,
            "PosSOC_CDsta_Dry_All"  : PosSOC_CDsta_Dry_All,
            "PosSOC_CDend_Dry_All"  : PosSOC_CDend_Dry_All,
            
            "CeEC_All": CeEC_All,
            "c_EC_r_new_All": c_EC_r_new_All,
            "c_e_r_new_All": c_e_r_new_All,
            "Ratio_CeEC_All":Ratio_CeEC_All ,  
            "Ratio_CeLi_All":Ratio_CeLi_All  ,       
            "Ratio_Dryout_All":Ratio_Dryout_All,

            "Vol_Elely_Tot_All": Vol_Elely_Tot_All,
            "Vol_Elely_JR_All": Vol_Elely_JR_All,
            "Vol_Pore_tot_All": Vol_Pore_tot_All,            
            "Vol_EC_consumed_All": Vol_EC_consumed_All,
            "Vol_Elely_need_All":Vol_Elely_need_All,
            "Width_all":Width_all,
            "Vol_Elely_add_All":Vol_Elely_add_All,
            "Vol_Pore_decrease_All":Vol_Pore_decrease_All,
            "Test_V_All":Test_V_All,
            "Test_V2_All":Test_V2_All,

            "Epsilon_CCend_Dry_First":Epsilon_CCend_Dry_First,
            "j_Neg_Int_CCend_Dry_First":j_Neg_Int_CCend_Dry_First,
            "Eta_Elely_CCend_Dry_First":Eta_Elely_CCend_Dry_First,
            "c_Elely_CCend_Dry_First":c_Elely_CCend_Dry_First,
            "Eta_Neg_rec_CCend_Dry_First":Eta_Neg_rec_CCend_Dry_First,
            "c_s_Neg_Surf_CCend_Dry_First":c_s_Neg_Surf_CCend_Dry_First,
            "Rough_Neg_CCend_Dry_First":Rough_Neg_CCend_Dry_First,  
            "t_SEIcr_Neg_CCend_Dry_First":t_SEIcr_Neg_CCend_Dry_First,


            "Epsilon_CDend_Dry_First":Epsilon_CDend_Dry_First,
            "j_Neg_Int_CDend_Dry_First":j_Neg_Int_CDend_Dry_First,
            "Eta_Elely_CDend_Dry_First":Eta_Elely_CDend_Dry_First,
            "c_Elely_CDend_Dry_First":c_Elely_CDend_Dry_First,
            "Eta_Neg_rec_CDend_Dry_First":Eta_Neg_rec_CDend_Dry_First,
            "c_s_Neg_Surf_CDend_Dry_First":c_s_Neg_Surf_CDend_Dry_First,
            "Rough_Neg_CDend_Dry_First":Rough_Neg_CDend_Dry_First,
            "t_SEIcr_Neg_CDend_Dry_First":t_SEIcr_Neg_CDend_Dry_First,

            "Epsilon_CCend_Dry_Last":Epsilon_CCend_Dry_Last,
            "j_Neg_Int_CCend_Dry_Last":j_Neg_Int_CCend_Dry_Last,
            "Eta_Elely_CCend_Dry_Last":Eta_Elely_CCend_Dry_Last,
            "c_Elely_CCend_Dry_Last":c_Elely_CCend_Dry_Last,
            "Eta_Neg_rec_CCend_Dry_Last":Eta_Neg_rec_CCend_Dry_Last,
            "c_s_Neg_Surf_CCend_Dry_Last":c_s_Neg_Surf_CCend_Dry_Last,
            "Rough_Neg_CCend_Dry_Last":Rough_Neg_CCend_Dry_Last,
            "t_SEIcr_Neg_CCend_Dry_Last":t_SEIcr_Neg_CCend_Dry_Last,

            "Epsilon_CDend_Dry_Last":Epsilon_CDend_Dry_Last,
            "j_Neg_Int_CDend_Dry_Last":j_Neg_Int_CDend_Dry_Last,
            "Eta_Elely_CDend_Dry_Last":Eta_Elely_CDend_Dry_Last,
            "c_Elely_CDend_Dry_Last":c_Elely_CDend_Dry_Last,
            "Eta_Neg_rec_CDend_Dry_Last":Eta_Neg_rec_CDend_Dry_Last,
            "c_s_Neg_Surf_CDend_Dry_Last":c_s_Neg_Surf_CDend_Dry_Last,
            "Rough_Neg_CDend_Dry_Last":Rough_Neg_CDend_Dry_Last,
            "t_SEIcr_Neg_CDend_Dry_Last":t_SEIcr_Neg_CDend_Dry_Last,
        }
        savemat(BasicPath + Target+ str(count_i) + '/' + str(count_i)+ '-StructDara_for_Mat.mat',mdic)   
            

        
# Run model without electrolyte NonDry out
def Run_model_wo_Dry_out(CyclePack, DatePack_scan_i,  BasicPath, Target, book_name_xlsx ):
    V_max = 4.2;        V_min = 2.5;   Temper_RPT = 25; 
    (Total_Cycles, Cycle_bt_RPT, Update_Cycles ) = CyclePack;
    Small_Loop =  int(Cycle_bt_RPT/Update_Cycles);   SaveTimes = int(Total_Cycles/Cycle_bt_RPT);   RPT_Cycles = 1; 
    #index 
    ModelTimer = pb.Timer()
    # unpack data:
    [index_xlsx,
    Ratio_excess_i,
    cs_Neg_Init_i,
    Diff_SEI_i,
    R_SEI_i,
    Bulk_Sol_Con_i,
    D_Li_inSEI_i,
    c_Li_inte_ref_i,
    Diff_EC_i,      # 
    k_SEI_i,        #
    LAMcr_prop_i,   #
    Crack_rate_i,   #
    Couple_SEI_LiP_i,
    k_LiP_i,
    Temper_i,
    mesh_par
    ] = DatePack_scan_i;
    count_i = int(index_xlsx);
    #
    Experiment_Long   = pb.Experiment(  [(f"Discharge at 1 C until {V_min} V", f"Charge at 0.3 C until {V_max} V", f"Hold at {V_max} V until C/100"),  ] * Update_Cycles  )  
    Experiment_RPT    = pb.Experiment( [ (f"Discharge at 0.1C until {V_min} V",  "Rest for 6 hours",  f"Charge at 0.1C until {V_max} V" ) ] * RPT_Cycles ) 
    Experiment_Breakin= pb.Experiment( [ ( f"Discharge at 0.1C until {V_min} V",  "Rest for 6 hours",  f"Charge at 0.1C until {V_max} V" )  ] *2 )
    
    sheet_name_xlsx = 'Results';
    ################ Important: index definition #################################
    index = list(np.arange(1,SaveTimes+1)*(Small_Loop+RPT_Cycles)-1); #index.insert(0,0)  # get all last ageing cycle before RPT, plus the first RPT test
    index2= list(np.arange(0,SaveTimes+1)*(Small_Loop+RPT_Cycles));         # get all RPT test
    cycles = np.arange(0,SaveTimes+1)*Cycle_bt_RPT; 
    cycles3 = list(np.arange(1,SaveTimes+1)*(Cycle_bt_RPT));
    try:             
        for i in range(0,1):    ######################   update parameter and run first RPT
            ChemistryChen=pb.parameter_sets.OKane2022   
            ChemistryChen["electrolyte"] = "lipf6_Valoen2005"
            Para_0=pb.ParameterValues(chemistry=ChemistryChen)
            # DFN parameter
            Para_0.update({"Upper voltage cut-off [V]": 4.21})
            Para_0.update({"Lower voltage cut-off [V]": 2.49})
            Para_0.update({"Negative electrode diffusivity [m2.s-1]": 2e-13})
            Para_0.update({"Positive electrode diffusivity [m2.s-1]": 1e-13})
            Para_0.update(
                {'Initial concentration in negative electrode [mol.m-3]':cs_Neg_Init_i }
                , check_already_exists=False)
            # Solvent consumption sub-model
            Para_0.update(
                {'Initial electrolyte excessive amount ratio':Ratio_excess_i}
                , check_already_exists=False)   # 
            Para_0.update(
                {'Current solvent concentration in the reservoir [mol.m-3]':Bulk_Sol_Con_i}
                , check_already_exists=False)     
            Para_0.update(
                {'Current electrolyte concentration in the reservoir [mol.m-3]':1000.0}
                , check_already_exists=False)           
            Para_0.update(
                {'Ratio of Li-ion concentration change in electrolyte consider solvent consumption':  
                1.0 }, check_already_exists=False)
            # general SEI
            Para_0.update({'SEI resistivity [Ohm.m]':R_SEI_i}) ;
            # solvent-diffusion limited
            Para_0.update({'Outer SEI solvent diffusivity [m2.s-1]':Diff_SEI_i});
            Para_0.update({'Bulk solvent concentration [mol.m-3]':Bulk_Sol_Con_i});
            # interstitial-diffusion limited
            Para_0.update({'Inner SEI lithium interstitial diffusivity [m2.s-1]':D_Li_inSEI_i})       
            Para_0.update({'Lithium interstitial reference concentration [mol.m-3]':c_Li_inte_ref_i}) 
            # ec-reaction limited
            Para_0.update({'EC diffusivity [m2.s-1]':Diff_EC_i}) 
            Para_0.update({'SEI kinetic rate constant [m.s-1]':k_SEI_i}) 
            Para_0.update({'EC initial concentration in electrolyte [mol.m-3]':Bulk_Sol_Con_i}) 
            # LiP and coupling with SEI:
            Para_0.update({"Dead lithium decay constant [s-1]": Couple_SEI_LiP_i})   
            Para_0.update({'Lithium plating kinetic rate constant [m.s-1]':k_LiP_i})  
            # Crack model
            Para_0.update({"Negative electrode LAM constant propotional term": LAMcr_prop_i})
            Para_0.update({"Positive electrode LAM constant propotional term": LAMcr_prop_i})  
            # make it simple for now, but may want to have T dependency in the future
            Para_0.update({"Negative electrode cracking rate": Crack_rate_i})
            Para_0.update({"Positive electrode cracking rate": Crack_rate_i})
            Para_0.update({"Negative electrode volume change": 0.0})
            Para_0.update({"Positive electrode volume change": 0.0})

            Model_0 = pb.lithium_ion.DFN(     
                options={
                    "particle": "Fickian diffusion",          
                    "SEI":"solvent-diffusion limited",   
                    "SEI on cracks":"true",  
                    "SEI film resistance":"distributed",          
                    "SEI porosity change":"true",      
                    "particle mechanics":"swelling and cracking",  
                    "loss of active material":"stress-driven", 
                    "lithium plating":"partially reversible"      } ) #
            
            var = pb.standard_spatial_vars  
            var_pts = {var.x_n: 20,  var.x_s: 10,  var.x_p: 20,  var.r_n: int(mesh_par),  var.r_p: int(mesh_par),  }
            
            Sim_0    = pb.Simulation(
                Model_0,        experiment = Experiment_Breakin,
                parameter_values = Para_0,
                solver = pb.CasadiSolver(),
                var_pts=var_pts,) #mode="safe"
            Sol_0    = Sim_0.solve(calc_esoh=False)

        for i in range(0,1):    ######################   initialize for post-processing before running subsequent model   
            # for RPT:
            Cap_RPT_NonDry_All=[]; CapLoss_LiP_NonDry_All=[]; R_Local_ECM_NonDry_All =[];
            NegSOC_CDsta_NonDry_All =[];  NegSOC_CDend_NonDry_All =[];  PosSOC_CDsta_NonDry_All =[]; PosSOC_CDend_NonDry_All =[];
            CapLoss_SEI_NonDry_All=[];    CapLoss_SEIcr_NonDry_All=[];    Cap_Pos_NonDry_All=[];          Cap_Neg_NonDry_All=[];   
            t_SEIcr_Xavgtot_NonDry_All = []; Rough_Neg_Xavg_NonDry_All =[];  
            T_0                  =  Para_0['Initial temperature [K]']
            Porosity_Neg_0       =  Para_0['Negative electrode porosity']  
            Porosity_Pos_0       =  Para_0['Positive electrode porosity']  
            Porosity_Sep_0       =  Para_0['Separator porosity']  
            cs_Neg_Max           =  Para_0["Maximum concentration in negative electrode [mol.m-3]"];
            L_p                  =  Para_0["Positive electrode thickness [m]"]
            L_n                  =  Para_0["Negative electrode thickness [m]"]
            L_s                  =  Para_0["Separator thickness [m]"]
            L_y_0                =  Para_0["Initial Electrode width [m]"]
            L_z_0                =  Para_0["Initial Electrode height [m]"]
            Int_ElelyExces_Ratio =  Para_0["Initial electrolyte excessive amount ratio"] 
            Vol_Elely_Tot        =  ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  * L_y_0 * L_z_0 * Int_ElelyExces_Ratio # Set initial electrolyte amount [L] 
            Vol_Elely_JR         =  ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  * L_y_0 * L_z_0 
      

        for i in range(0,1):    #################### post process for the first RPT cycle    ############################
            Cap_RPT_NonDry_All.append    (Sol_0.cycles[-1].steps[0]["Discharge capacity [A.h]"].entries[-1] - Sol_0.cycles[-1].steps[0]["Discharge capacity [A.h]"].entries[0])
            CapLoss_LiP_NonDry_All.append(Sol_0.cycles[-1].steps[0]["Loss of capacity to lithium plating [A.h]"].entries[-1])
            CapLoss_SEI_NonDry_All.append(Sol_0.cycles[-1].steps[0]["Loss of capacity to SEI [A.h]"].entries[-1])
            CapLoss_SEIcr_NonDry_All.append(Sol_0.cycles[-1].steps[0]["Loss of capacity to SEI on cracks [A.h]"].entries[-1])
            t_SEIcr_Xavgtot_NonDry_All.append(Sol_0.cycles[-1].steps[0]["X-averaged total SEI on cracks thickness [m]"].entries[-1])
            Rough_Neg_Xavg_NonDry_All.append(Sol_0.cycles[-1].steps[0]["X-averaged negative electrode roughness ratio"].entries[-1])
            R_Local_ECM_NonDry_All.append(Sol_0.cycles[-1].steps[0]["Local ECM resistance [Ohm]"].entries[-1])
            NegSOC_CDsta_NonDry_All .append(Sol_0.cycles[-1].steps[0]["Negative electrode SOC"].entries[0]);  
            NegSOC_CDend_NonDry_All .append(Sol_0.cycles[-1].steps[0]["Negative electrode SOC"].entries[-1]);  
            PosSOC_CDsta_NonDry_All .append(Sol_0.cycles[-1].steps[0]["Positive electrode SOC"].entries[0]); 
            PosSOC_CDend_NonDry_All .append(Sol_0.cycles[-1].steps[0]["Positive electrode SOC"].entries[-1]);  

            Cap_Pos_NonDry_All.append    (Sol_0.cycles[-1].steps[0]["Positive electrode capacity [A.h]"].entries[-1])
            Cap_Neg_NonDry_All.append    (Sol_0.cycles[-1].steps[0]["Negative electrode capacity [A.h]"].entries[-1])      

            Epsilon_CCend_NonDry_First = Sol_0.cycles[-1].steps[-1]["Porosity"].entries[:,-1]
            j_Neg_Int_CCend_NonDry_First = Sol_0.cycles[-1].steps[-1]["Negative electrode interfacial current density [A.m-2]"].entries[:,-1] 
            Eta_Elely_CCend_NonDry_First = Sol_0.cycles[-1].steps[-1]["Electrolyte potential [V]"].entries[:,-1] 
            c_Elely_CCend_NonDry_First = Sol_0.cycles[-1].steps[-1]["Electrolyte concentration [mol.m-3]"].entries[:,-1] 
            Eta_Neg_rec_CCend_NonDry_First = Sol_0.cycles[-1].steps[-1]["Negative electrode reaction overpotential [V]"].entries[:,-1] 
            c_s_Neg_Surf_CCend_NonDry_First = Sol_0.cycles[-1].steps[-1]["Negative particle surface concentration [mol.m-3]"].entries[:,-1] 
            Rough_Neg_CCend_NonDry_First = Sol_0.cycles[-1].steps[-1]["Negative electrode roughness ratio"].entries[:,-1] 
            t_SEIcr_Neg_CCend_NonDry_First = Sol_0.cycles[-1].steps[-1]["Total SEI on cracks thickness [m]"].entries[:,-1] 
            
            Epsilon_CDend_NonDry_First = Sol_0.cycles[-1].steps[0]["Porosity"].entries[:,-1]
            j_Neg_Int_CDend_NonDry_First = Sol_0.cycles[-1].steps[0]["Negative electrode interfacial current density [A.m-2]"].entries[:,-1] 
            Eta_Elely_CDend_NonDry_First = Sol_0.cycles[-1].steps[0]["Electrolyte potential [V]"].entries[:,-1] 
            c_Elely_CDend_NonDry_First = Sol_0.cycles[-1].steps[0]["Electrolyte concentration [mol.m-3]"].entries[:,-1] 
            Eta_Neg_rec_CDend_NonDry_First = Sol_0.cycles[-1].steps[0]["Negative electrode reaction overpotential [V]"].entries[:,-1] 
            c_s_Neg_Surf_CDend_NonDry_First = Sol_0.cycles[-1].steps[0]["Negative particle surface concentration [mol.m-3]"].entries[:,-1] 
            Rough_Neg_CDend_NonDry_First = Sol_0.cycles[-1].steps[0]["Negative electrode roughness ratio"].entries[:,-1] 
            t_SEIcr_Neg_CDend_NonDry_First = Sol_0.cycles[-1].steps[0]["Total SEI on cracks thickness [m]"].entries[:,-1] 
        

        x_n = Sol_0.cycles[0].steps[0]["x_n [m]"].entries[:,-1]; x = Sol_0.cycles[0].steps[0]["x [m]"].entries[:,-1];   # Get location variable, can be any single solution

        # Para_All.append(Para_0);                                Model_All.append(Model_0);    Sol_All_i.append(Sol_0); 
        Model_NonDry_old = Model_0  ; Sol_NonDry_old = Sol_0    ;   del Model_0,Sol_0
        #Step3: Write a big loop to finish the long experiment,    
        Index_update_all     =[]; cycle_count =0; Index_update_all.append(cycle_count);
        k=0; 
        while k < SaveTimes:    # biggest loop, 
            i=0;     cap_i = 0;
            while i < Small_Loop:
                # 3rd: run model based on new parameter and last updated solution Model  , Sol, Ratio_CeLi, Para_update, ModelExperiment, SaveAs_Cycles
                Model_NonDry_i, Sol_NonDry_i   = Run_Model_Base_On_Last_Solution( 
                    Model_NonDry_old  , Sol_NonDry_old ,  
                    Para_0 ,Experiment_Long, Update_Cycles,Temper_i,mesh_par )
                Model_NonDry_old = Model_NonDry_i;      Sol_NonDry_old = Sol_NonDry_i;   
                del Model_NonDry_i,Sol_NonDry_i
                i += 1;   cycle_count +=  Update_Cycles; 
                Index_update_all.append(cycle_count);
            # run RPT
            Index_update_all.append(cycle_count);
            Model_NonDry_i, Sol_NonDry_i  = Run_Model_Base_On_Last_Solution_RPT( 
                Model_NonDry_old  , Sol_NonDry_old ,   
                Para_0,      Experiment_RPT, RPT_Cycles, Temper_RPT ,mesh_par )     
            for ii in range(0,1):     # add later RPT cycle 
                Cap_RPT_NonDry_All.append    (
                    Sol_NonDry_i.cycles[-1].steps[0]["Discharge capacity [A.h]"].entries[-1] 
                    - Sol_NonDry_i.cycles[-1].steps[0]["Discharge capacity [A.h]"].entries[0])
                CapLoss_LiP_NonDry_All.append(Sol_NonDry_i.cycles[-1].steps[0]["Loss of capacity to lithium plating [A.h]"].entries[-1])
                CapLoss_SEI_NonDry_All.append(Sol_NonDry_i.cycles[-1].steps[0]["Loss of capacity to SEI [A.h]"].entries[-1])
                CapLoss_SEIcr_NonDry_All.append(Sol_NonDry_i.cycles[-1].steps[0]["Loss of capacity to SEI on cracks [A.h]"].entries[-1])
                t_SEIcr_Xavgtot_NonDry_All.append(Sol_NonDry_i.cycles[-1].steps[0]["X-averaged total SEI on cracks thickness [m]"].entries[-1])
                Rough_Neg_Xavg_NonDry_All.append(Sol_NonDry_i.cycles[-1].steps[0]["X-averaged negative electrode roughness ratio"].entries[-1])
                R_Local_ECM_NonDry_All.append(Sol_NonDry_i.cycles[-1].steps[0]["Local ECM resistance [Ohm]"].entries[-1])
                NegSOC_CDsta_NonDry_All .append(Sol_NonDry_i.cycles[-1].steps[0]["Negative electrode SOC"].entries[0]);  
                NegSOC_CDend_NonDry_All .append(Sol_NonDry_i.cycles[-1].steps[0]["Negative electrode SOC"].entries[-1]);  
                PosSOC_CDsta_NonDry_All .append(Sol_NonDry_i.cycles[-1].steps[0]["Positive electrode SOC"].entries[0]); 
                PosSOC_CDend_NonDry_All .append(Sol_NonDry_i.cycles[-1].steps[0]["Positive electrode SOC"].entries[-1]);  
                Cap_Pos_NonDry_All.append    (Sol_NonDry_i.cycles[-1].steps[0]["Positive electrode capacity [A.h]"].entries[-1])
                Cap_Neg_NonDry_All.append    (Sol_NonDry_i.cycles[-1].steps[0]["Negative electrode capacity [A.h]"].entries[-1])      
            if k == SaveTimes-1:  # Last RPT cycle
                Epsilon_CCend_NonDry_Last =  Sol_NonDry_i.cycles[-1].steps[-1]["Porosity"].entries[:,-1]
                j_Neg_Int_CCend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[-1]["Negative electrode interfacial current density [A.m-2]"].entries[:,-1] 
                Eta_Elely_CCend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[-1]["Electrolyte potential [V]"].entries[:,-1] 
                c_Elely_CCend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[-1]["Electrolyte concentration [mol.m-3]"].entries[:,-1] 
                Eta_Neg_rec_CCend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[-1]["Negative electrode reaction overpotential [V]"].entries[:,-1] 
                c_s_Neg_Surf_CCend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[-1]["Negative particle surface concentration [mol.m-3]"].entries[:,-1] 
                Rough_Neg_CCend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[-1]["Negative electrode roughness ratio"].entries[:,-1] 
                t_SEIcr_Neg_CCend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[-1]["Total SEI on cracks thickness [m]"].entries[:,-1] 
                
                Epsilon_CDend_NonDry_Last =  Sol_NonDry_i.cycles[-1].steps[0]["Porosity"].entries[:,-1]
                j_Neg_Int_CDend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[0]["Negative electrode interfacial current density [A.m-2]"].entries[:,-1] 
                Eta_Elely_CDend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[0]["Electrolyte potential [V]"].entries[:,-1] 
                c_Elely_CDend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[0]["Electrolyte concentration [mol.m-3]"].entries[:,-1] 
                Eta_Neg_rec_CDend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[0]["Negative electrode reaction overpotential [V]"].entries[:,-1] 
                c_s_Neg_Surf_CDend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[0]["Negative particle surface concentration [mol.m-3]"].entries[:,-1] 
                Rough_Neg_CDend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[0]["Negative electrode roughness ratio"].entries[:,-1] 
                t_SEIcr_Neg_CDend_NonDry_Last = Sol_NonDry_i.cycles[-1].steps[0]["Total SEI on cracks thickness [m]"].entries[:,-1] 

            Model_NonDry_old = Model_NonDry_i  ;                 Sol_NonDry_old = Sol_NonDry_i    ;   del Model_NonDry_i,Sol_NonDry_i
            k += 1
    except:
        data = openpyxl.load_workbook(BasicPath + Target + book_name_xlsx)   
        str_error = traceback.format_exc()           
        table = data.get_sheet_by_name('Results')
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得列数
        values = [
            [count_i, "NonDry",cs_Neg_Init_i, Diff_SEI_i, R_SEI_i, 
            Bulk_Sol_Con_i,D_Li_inSEI_i, 
            c_Li_inte_ref_i,
            Diff_EC_i,      # 
            k_SEI_i,        #
            LAMcr_prop_i,   #
            Crack_rate_i,   #
            Couple_SEI_LiP_i,
            k_LiP_i,Temper_i,mesh_par,
            "nan","nan",
            "nan","nan", 
            "nan","nan","nan",str_error],
            ]
        for i in range(1, len(values)+1):
            for j in range(1, len(values[i-1])+1):
                table.cell(nrows+i, j).value = values[i-1][j-1]     
        data.save(BasicPath + Target + book_name_xlsx)
        print("Fail in {}".format(ModelTimer.time()))
    else:
        # Finish everything, try to write into excel with real results or 
        data = openpyxl.load_workbook(BasicPath + Target+ book_name_xlsx)         
        table = data.get_sheet_by_name('Results')
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得列数
        Cap_RPT_NonDry_Loss = Cap_RPT_NonDry_All[0] - Cap_RPT_NonDry_All[-1];
        values = [
            [count_i, "NonDry",
            cs_Neg_Init_i, Diff_SEI_i, R_SEI_i, 
            Bulk_Sol_Con_i,D_Li_inSEI_i, 
            c_Li_inte_ref_i,
            Diff_EC_i,      # 
            k_SEI_i,        #
            LAMcr_prop_i,   #
            Crack_rate_i,   #
            Couple_SEI_LiP_i,
            k_LiP_i,Temper_i,mesh_par,
            Cap_RPT_NonDry_Loss,CapLoss_LiP_NonDry_All[-1],
            CapLoss_SEI_NonDry_All[-1],CapLoss_SEIcr_NonDry_All[-1],
            "nan", "nan","nan"],
            ]
        for i in range(1, len(values)+1):
            for j in range(1, len(values[i-1])+1):
                table.cell(nrows+i, j).value = values[i-1][j-1]     
        data.save(BasicPath + Target+ book_name_xlsx)
        print("Succeed in {}".format(ModelTimer.time()))
        print('This is the ', count_i, ' scan')
        
        # Newly add (220517): save plots, not just a single line in excel file:
        if not os.path.exists(BasicPath + Target + str(count_i)):
            os.mkdir(BasicPath + Target + str(count_i) );
        # Fig. 1 how much capacity is loss, how much is due to SEI and LiP?
        for mm in range(0,1):
            fs=17;Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(cycles, Cap_RPT_NonDry_All,     '-o', label="Scan=" + str(count_i) )
            axs[1].plot(cycles, CapLoss_LiP_NonDry_All,'-o', label="LiP - Scan=" + str(count_i) )
            axs[1].plot(cycles, CapLoss_SEI_NonDry_All ,'-o', label="SEI - Scan" + str(count_i) )
            axs[1].plot(cycles, CapLoss_SEIcr_NonDry_All ,'-o', label="sei-on-cracks - Scan" + str(count_i) )
            for i in range(0,Num_subplot):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[i].set_ylabel("Capacity [A.h]",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
            axs[0].set_title("Discharge capacity",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Capacity loss to LiP and SEI",   fontdict={'family':'DejaVu Sans','size':fs+1})
            plt.savefig(BasicPath + Target+ str(count_i)+"/Cap-LLI.png", dpi=100)

            fs=17;Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(cycles, t_SEIcr_Xavgtot_NonDry_All,     '-o', label="Scan=" + str(count_i) )
            axs[1].plot(cycles, Rough_Neg_Xavg_NonDry_All,'-o', label="Scan=" + str(count_i) )
            axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_ylabel("Roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs})
            for i in range(0,Num_subplot):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
            axs[0].set_title("X-avg tot Neg SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("X-avg Neg roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
            plt.savefig(BasicPath + Target+ str(count_i)+"/Cracks related.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[1].plot(cycles, Cap_Neg_NonDry_All[0]-Cap_Neg_NonDry_All,'-o',label="Scan=" + str(count_i))
            axs[1].plot(cycles, Cap_Pos_NonDry_All[0]-Cap_Pos_NonDry_All   ,'-^',label="Scan=" + str(count_i))
            axs[0].plot(cycles, R_Local_ECM_NonDry_All,                  '-o',label="Scan="+ str(count_i))
            for i in range(0,1):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
            axs[0].set_ylabel("ECM resistance [Ohm]",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[0].set_title("Local ECM resistance",   fontdict={'family':'DejaVu Sans','size':fs+1})
            for i in range(1,2):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[i].set_ylabel("Capacity [A.h]",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
            axs[1].set_title("LAM of Neg and Pos",   fontdict={'family':'DejaVu Sans','size':fs+1})
            plt.savefig(BasicPath + Target+ str(count_i)+"/LAM-IR.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(cycles, PosSOC_CDsta_NonDry_All,'-o',label="Start" )
            axs[0].plot(cycles, PosSOC_CDend_NonDry_All,'-^',label="End" )
            axs[1].plot(cycles, NegSOC_CDsta_NonDry_All,'-o',label="Start" )
            axs[1].plot(cycles, NegSOC_CDend_NonDry_All,'-^',label="End" )
            for i in range(0,2):
                axs[i].set_xlabel("Cycle numbers",   fontdict={'family':'DejaVu Sans','size':fs})
                axs[i].set_ylabel("SOC",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)     
            axs[0].set_title("Neg SOC range (Dis)",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Pos SOC range (Dis)",   fontdict={'family':'DejaVu Sans','size':fs+1})
            plt.savefig(BasicPath + Target+ str(count_i)+"/SOC_RPT_dis.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(x/np.max(x), Epsilon_CDend_NonDry_First,'-o',label="First")
            axs[0].plot(x/np.max(x), Epsilon_CDend_NonDry_Last,'-^',label="Last"  )
            axs[1].plot(x_n/np.max(x), Eta_Neg_rec_CDend_NonDry_First,'-o',label="First" )
            axs[1].plot(x_n/np.max(x), Eta_Neg_rec_CDend_NonDry_Last,'-^',label="Last" )
            axs[0].set_xlabel("Dimensionless Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_xlabel("Dimensionless Neg thickness",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[0].set_title("Porosity",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Neg electrode reaction overpotential",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[0].set_ylabel("Porosity",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_ylabel("Overpotential [V]",   fontdict={'family':'DejaVu Sans','size':fs})
            for i in range(0,2):
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
            plt.savefig(BasicPath + Target+ str(count_i)+"/Por Neg_S_eta.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(x_n/np.max(x), t_SEIcr_Neg_CDend_NonDry_First,'-o',label="First")
            axs[0].plot(x_n/np.max(x), t_SEIcr_Neg_CDend_NonDry_Last,'-^',label="Last"  )
            axs[1].plot(x_n/np.max(x), Rough_Neg_CDend_NonDry_First,'-o',label="First" )
            axs[1].plot(x_n/np.max(x), Rough_Neg_CDend_NonDry_Last,'-^',label="Last" )
            axs[0].set_title("Tot Neg SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Neg roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_ylabel("Roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs})
            for i in range(0,2):
                axs[i].set_xlabel("Dimensionless Neg thickness",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
            plt.savefig(BasicPath + Target+ str(count_i)+"/Cracks related spatial.png", dpi=100)

            Num_subplot = 2;
            fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
            axs[0].plot(x/np.max(x), c_Elely_CDend_NonDry_First,'-o',label="First")
            axs[0].plot(x/np.max(x), c_Elely_CDend_NonDry_Last,'-^',label="Last"  )
            axs[1].plot(x/np.max(x), Eta_Elely_CDend_NonDry_First,'-o',label="First" )
            axs[1].plot(x/np.max(x), Eta_Elely_CDend_NonDry_Last,'-^',label="Last" )
            axs[0].set_title("Electrolyte concentration",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[1].set_title("Electrolyte potential",   fontdict={'family':'DejaVu Sans','size':fs+1})
            axs[0].set_ylabel("Concentration [mol.m-3]",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[1].set_ylabel("Potential [V]",   fontdict={'family':'DejaVu Sans','size':fs})
            for i in range(0,2):
                axs[i].set_xlabel("Dimensionless Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
                labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
                axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
                axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
            plt.savefig(BasicPath + Target+ str(count_i)+"/Electrolyte concentration and potential.png", dpi=100)
        
          
        # Newly add (220706): save data
        if not os.path.exists(BasicPath + Target + str(count_i)):
            os.mkdir(BasicPath + Target + str(count_i) );
        mdic = {
            "x":x,
            "x_n":x_n,
            "cycles": cycles,
            "Cap_RPT_NonDry_All": Cap_RPT_NonDry_All,
            "Real_SaveAs_Cycles": CapLoss_LiP_NonDry_All,
            "CapLoss_SEI_NonDry_All": CapLoss_SEI_NonDry_All,
            "CapLoss_SEIcr_NonDry_All": CapLoss_SEIcr_NonDry_All,
            "t_SEIcr_Xavgtot_NonDry_All":t_SEIcr_Xavgtot_NonDry_All,
            "Rough_Neg_Xavg_NonDry_All":Rough_Neg_Xavg_NonDry_All,
            "SaveTimes": SaveTimes,
            "Cap_Neg_NonDry_All": Cap_Neg_NonDry_All,
            "Cap_Pos_NonDry_All": Cap_Pos_NonDry_All,
            "R_Local_ECM_NonDry_All": R_Local_ECM_NonDry_All,
            "NegSOC_CDsta_NonDry_All"  : NegSOC_CDsta_NonDry_All,
            "NegSOC_CDend_NonDry_All"  : NegSOC_CDend_NonDry_All,
            "PosSOC_CDsta_NonDry_All"  : PosSOC_CDsta_NonDry_All,
            "PosSOC_CDend_NonDry_All"  : PosSOC_CDend_NonDry_All,

            "Epsilon_CCend_NonDry_First":Epsilon_CCend_NonDry_First,
            "j_Neg_Int_CCend_NonDry_First":j_Neg_Int_CCend_NonDry_First,
            "Eta_Elely_CCend_NonDry_First":Eta_Elely_CCend_NonDry_First,
            "c_Elely_CCend_NonDry_First":c_Elely_CCend_NonDry_First,
            "Eta_Neg_rec_CCend_NonDry_First":Eta_Neg_rec_CCend_NonDry_First,
            "c_s_Neg_Surf_CCend_NonDry_First":c_s_Neg_Surf_CCend_NonDry_First,
            "Rough_Neg_CCend_NonDry_First":Rough_Neg_CCend_NonDry_First,  
            "t_SEIcr_Neg_CCend_NonDry_First":t_SEIcr_Neg_CCend_NonDry_First,

            "Epsilon_CDend_NonDry_First":Epsilon_CDend_NonDry_First,
            "j_Neg_Int_CDend_NonDry_First":j_Neg_Int_CDend_NonDry_First,
            "Eta_Elely_CDend_NonDry_First":Eta_Elely_CDend_NonDry_First,
            "c_Elely_CDend_NonDry_First":c_Elely_CDend_NonDry_First,
            "Eta_Neg_rec_CDend_NonDry_First":Eta_Neg_rec_CDend_NonDry_First,
            "c_s_Neg_Surf_CDend_NonDry_First":c_s_Neg_Surf_CDend_NonDry_First,
            "Rough_Neg_CDend_NonDry_First":Rough_Neg_CDend_NonDry_First,
            "t_SEIcr_Neg_CDend_NonDry_First":t_SEIcr_Neg_CDend_NonDry_First,

            "Epsilon_CCend_NonDry_Last":Epsilon_CCend_NonDry_Last,
            "j_Neg_Int_CCend_NonDry_Last":j_Neg_Int_CCend_NonDry_Last,
            "Eta_Elely_CCend_NonDry_Last":Eta_Elely_CCend_NonDry_Last,
            "c_Elely_CCend_NonDry_Last":c_Elely_CCend_NonDry_Last,
            "Eta_Neg_rec_CCend_NonDry_Last":Eta_Neg_rec_CCend_NonDry_Last,
            "c_s_Neg_Surf_CCend_NonDry_Last":c_s_Neg_Surf_CCend_NonDry_Last,
            "Rough_Neg_CCend_NonDry_Last":Rough_Neg_CCend_NonDry_Last,
            "t_SEIcr_Neg_CCend_NonDry_Last":t_SEIcr_Neg_CCend_NonDry_Last,

            "Epsilon_CDend_NonDry_Last":Epsilon_CDend_NonDry_Last,
            "j_Neg_Int_CDend_NonDry_Last":j_Neg_Int_CDend_NonDry_Last,
            "Eta_Elely_CDend_NonDry_Last":Eta_Elely_CDend_NonDry_Last,
            "c_Elely_CDend_NonDry_Last":c_Elely_CDend_NonDry_Last,
            "Eta_Neg_rec_CDend_NonDry_Last":Eta_Neg_rec_CDend_NonDry_Last,
            "c_s_Neg_Surf_CDend_NonDry_Last":c_s_Neg_Surf_CDend_NonDry_Last,
            "Rough_Neg_CDend_NonDry_Last":Rough_Neg_CDend_NonDry_Last,
            "t_SEIcr_Neg_CDend_NonDry_Last":t_SEIcr_Neg_CDend_NonDry_Last,
            
        }
        savemat(BasicPath + Target+ str(count_i) + '/' + str(count_i)+ '-StructDara_for_Mat.mat',mdic)   
        
       

                
        
        
  
        
        







































































