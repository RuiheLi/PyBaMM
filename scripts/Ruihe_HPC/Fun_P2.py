
import pybamm as pb;import pandas as pd   ;import numpy as np;import os;
import io; import logging; import csv
import matplotlib.pyplot as plt;import os;#import imageio;import timeit
from scipy.io import savemat,loadmat;from pybamm import constants,exp,sqrt;
import matplotlib as mpl; 
from multiprocessing import Queue, Process, set_start_method
from queue import Empty
fs=17; 
font = {'family' : 'DejaVu Sans','size': fs}
mpl.rc('font', **font)
import openpyxl
import traceback
import random;import time, signal


def graphite_LGM50_electrolyte_exchange_current_density_Chen2020_6e_9(
    c_e, c_s_surf, c_s_max, T
):
    m_ref = 6e-14  # (A/m2)(m3/mol)**1.5 - includes ref concentrations
    E_r = 35000
    arrhenius = pb.exp(E_r / pb.constants.R * (1 / 298.15 - 1 / T))

    return (
        m_ref * arrhenius * c_e**0.5 * c_s_surf**0.5 * (c_s_max - c_s_surf) ** 0.5
    )

def graphite_LGM50_electrolyte_exchange_current_density_Chen2020_6e_5(
    c_e, c_s_surf, c_s_max, T
):
    m_ref =1e-13  # (A/m2)(m3/mol)**1.5 - includes ref concentrations
    E_r = 35000
    arrhenius = pb.exp(E_r / pb.constants.R * (1 / 298.15 - 1 / T))

    return (
        m_ref * arrhenius * c_e**0.5 * c_s_surf**0.5 * (c_s_max - c_s_surf) ** 0.5
    )
def graphite_LGM50_electrolyte_exchange_current_density_Chen2020(
    c_e, c_s_surf, c_s_max, T
):
    m_ref = 6.48e-7  # (A/m2)(m3/mol)**1.5 - includes ref concentrations
    E_r = 35000
    arrhenius = pb.exp(E_r / pb.constants.R * (1 / 298.15 - 1 / T))

    return (
        m_ref * arrhenius * c_e**0.5 * c_s_surf**0.5 * (c_s_max - c_s_surf) ** 0.5
    )


###################################################################
#############    From Patrick from Github        ##################
###################################################################
class TimeoutFunc(object):
    def __init__(self, func, timeout=None, timeout_val=None):
        assert callable(func), 'Positional argument 1 must be a callable method'
        self.func = func
        self.timeout = timeout
        self.timeout_val = timeout_val

    def _queued_f(self, *args, queue, **kwargs):
        Result_List = self.func(*args, **kwargs)
        try:
            queue.put(Result_List)
        except BrokenPipeError as exc:
            pass

    def __call__(self, *args, **kwargs):
        q = Queue(1)
        p = Process(target=self._queued_f, args=args,
            kwargs={**kwargs, 'queue': q})
        p.daemon = True
        p.start()
        try:
            Result_List = q.get(timeout=self.timeout)
        except Empty as exc: 
            # todo: need to incooperate other cases such as pb.expression_tree.exceptions.ModelError,
            #      pb.expression_tree.exceptions.SolverError
            p.terminate()
            Call_ref = RioCallback() 
            Result_List = [
                self.timeout_val,
                self.timeout_val,
                Call_ref]
        return Result_List  
###################################################################
#############    New functions from P3 - 221113        ############
###################################################################
# DEFINE my callback:
class RioCallback(pb.callbacks.Callback):
    def __init__(self, logfile=None):
        self.logfile = logfile
        self.success  = True
        if logfile is None:
            # Use pybamm's logger, which prints to command line
            self.logger = pb.logger
        else:
            # Use a custom logger, this will have its own level so set it to the same
            # level as the pybamm logger (users can override this)
            self.logger = pb.get_new_logger(__name__, logfile)
            self.logger.setLevel(pb.logger.level)
    
    def on_experiment_error(self, logs):
        self.success  = False
        self.logs = logs
    def on_experiment_infeasible(self, logs):
        self.success  = False
        self.logs = logs
# define to kill too long runs - Jianbo Huang kindly writes this
class TimeoutError(Exception):
	pass
def handle_signal(signal_num, frame):
	raise TimeoutError


# Define a function to calculate concentration change, whether electrolyte being squeezed out or added in
def Cal_new_con_Update(Sol,Para):   # subscript r means the reservoir
    # Note: c_EC_r is the initial EC  concentraiton in the reservoir; 
    #       c_e_r  is the initial Li+ concentraiton in the reservoir;
    #############################################################################################################################  
    ###############################           Step-1 Prepare parameters:        #################################################
    #############################################################################################################################  
    L_p   =Para["Positive electrode thickness [m]"]
    L_n   =Para["Negative electrode thickness [m]"]
    L_s   =Para["Separator thickness [m]"]
    L_y   =Para["Electrode width [m]"]   # Also update to change A_cc and therefore Q
    L_z   =Para["Electrode height [m]"]
    c_EC_r_old=Para["Current solvent concentration in the reservoir [mol.m-3]"]   # add two new parameters for paper 2
    c_e_r_old =Para["Current electrolyte concentration in the reservoir [mol.m-3]"]
    # Initial EC concentration in JR
    c_EC_JR_old =Para["Bulk solvent concentration [mol.m-3]"]  # Be careful when multiplying with pore volume to get amount in mole. Because of electrolyte dry out, that will give more amount than real.   
    # LLI due to electrode,Ratio of EC and lithium is 1:1 -> EC amount consumed is LLINegSEI[-1]
    LLINegSEI = (
        Sol["Loss of lithium to SEI [mol]"].entries[-1] 
        - Sol["Loss of lithium to SEI [mol]"].entries[0] )
    LLINegSEIcr = (
        Sol["Loss of lithium to SEI on cracks [mol]"].entries[-1]
        - 
        Sol["Loss of lithium to SEI on cracks [mol]"].entries[0]
        )
    LLINegDeadLiP = (
        Sol["Loss of lithium to dead lithium plating [mol]"].entries[-1] 
        - Sol["Loss of lithium to dead lithium plating [mol]"].entries[0])
    LLINegLiP = (
        Sol["Loss of lithium to lithium plating [mol]"].entries[-1] 
        - Sol["Loss of lithium to lithium plating [mol]"].entries[0])
    cLi_Xavg  = Sol["X-averaged electrolyte concentration [mol.m-3]"].entries[-1] 
    # Pore volume change with time:
    PoreVolNeg_0 = Sol["X-averaged negative electrode porosity"].entries[0]*L_n*L_y*L_z;
    PoreVolSep_0 = Sol["X-averaged separator porosity"].entries[0]*L_s*L_y*L_z;
    PoreVolPos_0 = Sol["X-averaged positive electrode porosity"].entries[0]*L_p*L_y*L_z;
    PoreVolNeg_1 = Sol["X-averaged negative electrode porosity"].entries[-1]*L_n*L_y*L_z;
    PoreVolSep_1 = Sol["X-averaged separator porosity"].entries[-1]*L_s*L_y*L_z;
    PoreVolPos_1 = Sol["X-averaged positive electrode porosity"].entries[-1]*L_p*L_y*L_z;
    #############################################################################################################################  
    ##### Step-2 Determine How much electrolyte is added (from the reservoir to JR) or squeezed out (from JR to reservoir) ######
    #######################   and finish electrolyte mixing     #################################################################  
    Vol_Elely_Tot_old = Para["Current total electrolyte volume in whole cell [m3]"] 
    Vol_Elely_JR_old  = Para["Current total electrolyte volume in jelly roll [m3]"] 
    if Vol_Elely_Tot_old - Vol_Elely_JR_old < 0:
        print('Model error! Electrolyte in JR is larger than in the cell!')
    Vol_Pore_tot_old  = PoreVolNeg_0 + PoreVolSep_0 + PoreVolPos_0    # pore volume at start time of the run
    Vol_Pore_tot_new  = PoreVolNeg_1 + PoreVolSep_1 + PoreVolPos_1    # pore volume at end   time of the run, intrinsic variable 
    Vol_Pore_decrease = Vol_Elely_JR_old  - Vol_Pore_tot_new # WHY Vol_Elely_JR_old not Vol_Pore_tot_old here? Because even the last state of the last solution (n-1) and the first state of the next solution (n) can be slightly different! 
    # EC:lithium:SEI=2:2:1     for SEI=(CH2OCO2Li)2, but because of too many changes are required, change to 2:1:1 for now
    # update 230703: Simon insist we should do: EC:lithium:SEI  =  2:2:1 and change V_SEI accordingly 
    # Because inner and outer SEI partial molar volume is the same, just set one for whole SEI
    VmolSEI   = Para["Outer SEI partial molar volume [m3.mol-1]"] # 9.8e-5,
    VmolLiP   = Para["Lithium metal partial molar volume [m3.mol-1]"] # 1.3e-05
    VmolEC    = Para["EC partial molar volume [m3.mol-1]"]
    #################   KEY EQUATION FOR DRY-OUT MODEL                   #################
    # UPDATE 230525: assume the formation of dead lithium doesn’t consumed EC
    Vol_EC_consumed  =  ( LLINegSEI + LLINegSEIcr   ) * 1 * VmolEC    # Mark: either with 2 or not, will decide how fast electrolyte being consumed!
    Vol_Elely_need   = Vol_EC_consumed - Vol_Pore_decrease
    Vol_SEILiP_increase = 0.5*(
        (LLINegSEI+LLINegSEIcr) * VmolSEI 
        #+ LLINegLiP * VmolLiP
        )    #  volume increase due to SEI+total LiP 
    #################   KEY EQUATION FOR DRY-OUT MODEL                   #################

    Test_V = (Vol_SEILiP_increase - Vol_Pore_decrease)/Vol_Pore_decrease*100  #  This value should always be zero, but now not, which becomes the source of error!
    Test_V2= (Vol_Pore_tot_old - Vol_Elely_JR_old) / Vol_Elely_JR_old * 100; # Porosity errors due to first time step
    
    # Start from here, there are serveral variables to be determined:
    #   1) Vol_Elely_add, or Vol_Elely_squeezed; depends on conditions. easier for 'squeezed' condition
    #   2) Vol_Elely_Tot_new, should always equals to Vol_Elely_Tot_old -  Vol_EC_consumed;
    #   3) Vol_Elely_JR_new, for 'add' condition: see old code; for 'squeezed' condition, equals to pore volume in JR
    #   4) Ratio_Dryout, for 'add' condition: see old code; for 'squeezed' condition, equals to Vol_Elely_Tot_new/Vol_Pore_tot_new 
    #   5) Ratio_CeEC_JR and Ratio_CeLi_JR: for 'add' condition: see old code; for 'squeezed' condition, equals to 1 (unchanged)
    #   6) c_e_r_new and c_EC_r_new: for 'add' condition: equals to old ones (unchanged); for 'squeezed' condition, need to carefully calculate     
    #   7) Width_new: for 'add' condition: see old code; for 'squeezed' condition, equals to L_y (unchanged)
    Vol_Elely_Tot_new = Vol_Elely_Tot_old - Vol_EC_consumed;

    
    if Vol_Elely_need < 0:
        print('Electrolyte is being squeezed out, check plated lithium (active and dead)')
        Vol_Elely_squeezed = - Vol_Elely_need;   # Make Vol_Elely_squeezed>0 for simplicity 
        Vol_Elely_add = 0.0;
        Vol_Elely_JR_new = Vol_Pore_tot_new; 
        Ratio_Dryout = Vol_Elely_Tot_new / Vol_Elely_JR_new
        Ratio_CeEC_JR = 1.0;    # the concentrations in JR don't need to change
        Ratio_CeLi_JR = 1.0; 
        SqueezedLiMol   =  Vol_Elely_squeezed*cLi_Xavg;
        SqueezedECMol   =  Vol_Elely_squeezed*c_EC_JR_old;
        Vol_Elely_reservoir_old = Vol_Elely_Tot_old - Vol_Elely_JR_old; 
        LiMol_reservoir_new = Vol_Elely_reservoir_old*c_e_r_old  + SqueezedLiMol;
        ECMol_reservoir_new = Vol_Elely_reservoir_old*c_EC_r_old + SqueezedECMol;
        c_e_r_new= LiMol_reservoir_new / (Vol_Elely_reservoir_old+Vol_Elely_squeezed);
        c_EC_r_new= ECMol_reservoir_new / (Vol_Elely_reservoir_old+Vol_Elely_squeezed);
        Width_new = L_y; 
    else:   # this means Vol_Elely_need >= 0, therefore the folliwing script should works for Vol_Elely_need=0 as well!
        # Important: Calculate added electrolyte based on excessive electrolyte, can be: 1) added as required; 2) added some, but less than required; 3) added 0 
        Vol_Elely_squeezed = 0;  
        if Vol_Elely_Tot_old > Vol_Elely_JR_old:                             # This means Vol_Elely_JR_old = Vol_Pore_tot_old ()
            if Vol_Elely_Tot_old-Vol_Elely_JR_old >= Vol_Elely_need:         # 1) added as required
                Vol_Elely_add     = Vol_Elely_need;  
                Vol_Elely_JR_new  = Vol_Pore_tot_new;  # also equals to 'Vol_Pore_tot_new', or Vol_Pore_tot_old - Vol_Pore_decrease
                Ratio_Dryout = 1.0;
            else:                                                            # 2) added some, but less than required;                                                         
                Vol_Elely_add     = Vol_Elely_Tot_old - Vol_Elely_JR_old;   
                Vol_Elely_JR_new  = Vol_Elely_Tot_new;                       # This means Vol_Elely_JR_new <= Vol_Pore_tot_new
                Ratio_Dryout = Vol_Elely_JR_new/Vol_Pore_tot_new;
        else:                                                                # 3) added 0 
            Vol_Elely_add = 0;
            Vol_Elely_JR_new  = Vol_Elely_Tot_new; 
            Ratio_Dryout = Vol_Elely_JR_new/Vol_Pore_tot_new;

        # Next: start mix electrolyte based on previous defined equation
        # Lithium amount in liquid phase, at initial time point
        TotLi_Elely_JR_Old = Sol["Total lithium in electrolyte [mol]"].entries[-1]; # remember this is only in JR
        # Added lithium and EC amount in the added lithium electrolyte: - 
        AddLiMol   =  Vol_Elely_add*c_e_r_old
        AddECMol   =  Vol_Elely_add*c_EC_r_old
        # Total amount of Li and EC in current electrolyte:
        # Remember Li has two parts, initial and added; EC has three parts, initial, consumed and added
        TotLi_Elely_JR_New   = TotLi_Elely_JR_Old + AddLiMol
        TotECMol_JR   = Vol_Elely_JR_old*c_EC_JR_old - LLINegSEI + AddECMol  # EC:lithium:SEI=2:2:1     for SEI=(CH2OCO2Li)2
        Ratio_CeEC_JR  = TotECMol_JR    /   Vol_Elely_JR_new   / c_EC_JR_old
        Ratio_CeLi_JR  = TotLi_Elely_JR_New    /   TotLi_Elely_JR_Old   /  Ratio_Dryout # Mark, change on 21-11-19
        c_e_r_new  = c_e_r_old;
        c_EC_r_new = c_EC_r_old;
        Width_new   = Ratio_Dryout * L_y;
        
    # Collect the above parameter in Data_Pack to shorten the code  
    Data_Pack   = [
        Vol_EC_consumed, 
        Vol_Elely_need, 
        Test_V, 
        Test_V2, 
        Vol_Elely_add, 
        Vol_Elely_Tot_new, 
        Vol_Elely_JR_new, 
        Vol_Pore_tot_new, 
        Vol_Pore_decrease, 
        c_e_r_new, c_EC_r_new,
        Ratio_Dryout, Ratio_CeEC_JR, 
        Ratio_CeLi_JR,
        Width_new, 
        ]  # 16 in total
    #print('Loss of lithium to SEI', LLINegSEI, 'mol') 
    #print('Loss of lithium to dead lithium plating', LLINegDeadLiP, 'mol') 
    #############################################################################################################################  
    ###################       Step-4 Update parameters here        ##############################################################
    #############################################################################################################################
    Para.update(   
        {'Bulk solvent concentration [mol.m-3]':  
         c_EC_JR_old * Ratio_CeEC_JR  })
    Para.update(
        {'EC initial concentration in electrolyte [mol.m-3]':
         c_EC_JR_old * Ratio_CeEC_JR },) 
    Para.update(   
        {'Ratio of Li-ion concentration change in electrolyte consider solvent consumption':  
        Ratio_CeLi_JR }, check_already_exists=False) 
    Para.update(   
        {'Current total electrolyte volume in whole cell [m3]':  Vol_Elely_Tot_new  }, 
        check_already_exists=False)
    Para.update(   
        {'Current total electrolyte volume in jelly roll [m3]':  Vol_Elely_JR_new  }, 
        check_already_exists=False)
    Para.update(   
        {'Ratio of electrolyte dry out in jelly roll':Ratio_Dryout}, 
        check_already_exists=False)
    Para.update(   {'Electrode width [m]':Width_new})    
    Para.update(   
        {'Current solvent concentration in the reservoir [mol.m-3]':c_EC_r_new}, 
        check_already_exists=False)     
    Para.update(   
        {'Current electrolyte concentration in the reservoir [mol.m-3]':c_e_r_new}, 
        check_already_exists=False)             
    return Data_Pack,Para

# Define a function to calculate based on previous solution
def Run_Model_Base_On_Last_Solution( 
    Model  , Sol , Para_update, ModelExperiment, 
    Update_Cycles,Temper_i ,mesh_list,submesh_strech):
    # Use Sulzer's method: inplace = false
    # Important line: define new model based on previous solution
    Ratio_CeLi = Para_update[
        "Ratio of Li-ion concentration change in electrolyte consider solvent consumption"]
    dict_short = {}; 
    list_short = []
    # update 220808 to satisfy random model option:
    for var, equation in Model.initial_conditions.items():
        #print(var._name)
        list_short.append(var._name)
    # delete Porosity times concentration and Electrolyte potential then add back
    list_short.remove("Porosity times concentration [mol.m-3]");
    list_short.remove("Electrolyte potential [V]");
    list_short.extend(
        ("Negative electrode porosity times concentration [mol.m-3]",
        "Separator porosity times concentration [mol.m-3]",
        "Positive electrode porosity times concentration [mol.m-3]",   

        "Negative electrolyte potential [V]",
        "Separator electrolyte potential [V]",
        "Positive electrolyte potential [V]",))
    for list_short_i in list_short:
        dict_short.update( { list_short_i : Sol.last_state[list_short_i].data  }  )
    dict_short["Negative electrode porosity times concentration [mol.m-3]"] = (
        dict_short["Negative electrode porosity times concentration [mol.m-3]"] * Ratio_CeLi )# important: update sol here!
    dict_short["Separator porosity times concentration [mol.m-3]"] = (
        dict_short["Separator porosity times concentration [mol.m-3]"] * Ratio_CeLi )# important: update sol here!
    dict_short["Positive electrode porosity times concentration [mol.m-3]"] = (
        dict_short["Positive electrode porosity times concentration [mol.m-3]"] * Ratio_CeLi )# important: update sol here!
    Model_new = Model.set_initial_conditions_from(dict_short, inplace=False)
    Para_update.update(   {'Ambient temperature [K]':Temper_i });   # run model at 45 degree C
    
    var = pb.standard_spatial_vars  
    var_pts = {
        var.x_n: int(mesh_list[0]),  
        var.x_s: int(mesh_list[1]),  
        var.x_p: int(mesh_list[2]),  
        var.r_n: int(mesh_list[3]),  
        var.r_p: int(mesh_list[4]),  
        }
    submesh_types = Model.default_submesh_types
    if submesh_strech == "nan":
            pass
    else:
        particle_mesh = pb.MeshGenerator(
            pb.Exponential1DSubMesh, 
            submesh_params={"side": "right", "stretch": int(submesh_strech)})
        submesh_types["negative particle"] = particle_mesh
        submesh_types["positive particle"] = particle_mesh 
    Call_Age = RioCallback()  # define callback
    Simnew = pb.Simulation(
        Model_new,
        experiment = ModelExperiment, 
        parameter_values=Para_update, 
        solver = pb.CasadiSolver(),
        var_pts = var_pts,
        submesh_types=submesh_types
    )
    try:
        Sol_new = Simnew.solve(
            calc_esoh=False,
            save_at_cycles = Update_Cycles,
            callbacks=Call_Age)
    except (
        pb.expression_tree.exceptions.ModelError,
        pb.expression_tree.exceptions.SolverError
        ) as e:
        Sol_new = "Model error or solver error"
    else:
        # add 230221 - update 230317 try to access Throughput capacity more than once
        i_try = 0
        while i_try<3:
            try:
                getSth2 = Sol_new['Throughput capacity [A.h]'].entries[-1]
            except:
                i_try += 1
                print(f"Fail to read Throughput capacity for the {i_try}th time")
            else:
                break
        i_try = 0
        while i_try<3:
            try:
                getSth = Sol['Throughput capacity [A.h]'].entries[-1]
            except:
                i_try += 1
                print(f"Fail to read Throughput capacity for the {i_try}th time")
            else:
                break
        Sol_new['Throughput capacity [A.h]'].entries += getSth
    # print("Solved this model in {}".format(ModelTimer.time()))
    Result_list = [Model_new, Sol_new,Call_Age]
    return Result_list

def Run_Model_Base_On_Last_Solution_RPT( 
    Model  , Sol,  Para_update, 
    ModelExperiment ,Update_Cycles, Temper_i,mesh_list,submesh_strech):
    # Use Sulzer's method: inplace = false
    Ratio_CeLi = Para_update["Ratio of Li-ion concentration change in electrolyte consider solvent consumption"]
    # print("Model is now using average EC Concentration of:",Para_update['Bulk solvent concentration [mol.m-3]'])
    # print("Ratio of electrolyte dry out in jelly roll is:",Para_update['Ratio of electrolyte dry out in jelly roll'])
    # print("Model is now using an electrode width of:",Para_update['Electrode width [m]'])
    # Important line: define new model based on previous solution
    dict_short = {}; 
    list_short = []
    # update 220808 to satisfy random model option:
    for var, equation in Model.initial_conditions.items():
        #print(var._name)
        list_short.append(var._name)
    # delete Porosity times concentration and Electrolyte potential then add back
    list_short.remove("Porosity times concentration [mol.m-3]");
    list_short.remove("Electrolyte potential [V]");
    list_short.extend(
        ("Negative electrode porosity times concentration [mol.m-3]",
        "Separator porosity times concentration [mol.m-3]",
        "Positive electrode porosity times concentration [mol.m-3]",   

        "Negative electrolyte potential [V]",
        "Separator electrolyte potential [V]",
        "Positive electrolyte potential [V]",))
    for list_short_i in list_short:
        dict_short.update( { list_short_i : Sol.last_state[list_short_i].data  }  )
    dict_short["Negative electrode porosity times concentration [mol.m-3]"] = (
        dict_short["Negative electrode porosity times concentration [mol.m-3]"] * Ratio_CeLi )# important: update sol here!
    dict_short["Separator porosity times concentration [mol.m-3]"] = (
        dict_short["Separator porosity times concentration [mol.m-3]"] * Ratio_CeLi )# important: update sol here!
    dict_short["Positive electrode porosity times concentration [mol.m-3]"] = (
        dict_short["Positive electrode porosity times concentration [mol.m-3]"] * Ratio_CeLi )# important: update sol here!
    Model_new = Model.set_initial_conditions_from(dict_short, inplace=False)
    Para_update.update(   {'Ambient temperature [K]':Temper_i });
    
    var = pb.standard_spatial_vars  
    var_pts = {
        var.x_n: int(mesh_list[0]),  
        var.x_s: int(mesh_list[1]),  
        var.x_p: int(mesh_list[2]),  
        var.r_n: int(mesh_list[3]),  
        var.r_p: int(mesh_list[4]),  }
    submesh_types = Model.default_submesh_types
    if submesh_strech == "nan":
        pass
    else:
        particle_mesh = pb.MeshGenerator(
            pb.Exponential1DSubMesh, 
            submesh_params={"side": "right", "stretch": int(submesh_strech)})
        submesh_types["negative particle"] = particle_mesh
        submesh_types["positive particle"] = particle_mesh 
    Simnew = pb.Simulation(
        Model_new,
        experiment = ModelExperiment, 
        parameter_values=Para_update, 
        solver = pb.CasadiSolver(),
        var_pts = var_pts,
        submesh_types=submesh_types
    )
    Call_RPT = RioCallback()  # define callback
    try:
        Sol_new = Simnew.solve(
            calc_esoh=False,
            save_at_cycles = Update_Cycles,
            callbacks=Call_RPT) 
    except (
        pb.expression_tree.exceptions.ModelError,
        pb.expression_tree.exceptions.SolverError
        ) as e:
        Sol_new = "Model error or solver error"
    else:
        # add 230221 - update 230317 try to access Throughput capacity more than once
        i_try = 0
        while i_try<3:
            try:
                getSth2 = Sol_new['Throughput capacity [A.h]'].entries[-1]
            except:
                i_try += 1
                print(f"Fail to read Throughput capacity for the {i_try}th time")
            else:
                break
        i_try = 0
        while i_try<3:
            try:
                getSth = Sol['Throughput capacity [A.h]'].entries[-1]
            except:
                i_try += 1
                print(f"Fail to read Throughput capacity for the {i_try}th time")
            else:
                break
        Sol_new['Throughput capacity [A.h]'].entries += getSth
    # print("Solved this model in {}".format(ModelTimer.time()))
    Result_List_RPT = [Model_new, Sol_new,Call_RPT]
    return Result_List_RPT

def write_excel_xlsx(path, sheet_name, value):
    import numpy as np
    index = len(value)
    workbook = openpyxl.Workbook()  # 新建工作簿（默认有一个sheet？）
    sheet = workbook.active  # 获得当前活跃的工作页，默认为第一个工作页
    sheet.title = sheet_name  # 给sheet页的title赋值
    for i in range(0, index):
        value_i = value[i]
        if isinstance(value_i, int):
            value_i = [value_i,]
        for j in range(0, len(value_i)):
            sheet.cell(row=i + 1, column=j + 1, value=str(value_i[j]))  # 行，列，值 这里是从1开始计数的
    workbook.save(path)  # 一定要保存
    print("Successfully create a excel file")

def recursive_scan(mylist,kvs, key_list, acc):
    # 递归终止条件
    if len(key_list) == 0:
        mylist.append(acc.copy())   # copy.deepcopy(acc) 如果value是个list，就要deep copy了
        return mylist
    # 继续递归
    k = key_list[0]
    for v in kvs[k]:
        acc[k] = v
        # print(v)
        recursive_scan(mylist,kvs, key_list[1:], acc)

# this function is to initialize the para with a known dict
def Para_init(Para_dict):
    Para_dict_used = Para_dict.copy();
    Para_0=pb.ParameterValues(Para_dict_used["Para_Set"]  )
    Para_dict_used.pop("Para_Set")
    import ast,json

    if Para_dict_used.__contains__("Scan No"):
        Para_dict_used.pop("Scan No")
    if Para_dict_used.__contains__("Exp No."):
        Para_dict_used.pop("Exp No.")

    if Para_dict_used.__contains__("Total ageing cycles"):
        Total_Cycles = Para_dict_used["Total ageing cycles"]  
        Para_dict_used.pop("Total ageing cycles")
    else:
        Total_Cycles = None
    if Para_dict_used.__contains__("Ageing cycles between RPT"):
        Cycle_bt_RPT = Para_dict_used["Ageing cycles between RPT"]  
        Para_dict_used.pop("Ageing cycles between RPT")
    else:
        Cycle_bt_RPT = None
    if Para_dict_used.__contains__("Update cycles for ageing"):
        Update_Cycles = Para_dict_used["Update cycles for ageing"]  
        Para_dict_used.pop("Update cycles for ageing")
    else:
        Update_Cycles = None
    if Para_dict_used.__contains__("Cycles within RPT"):
        RPT_Cycles = Para_dict_used["Cycles within RPT"]  
        Para_dict_used.pop("Cycles within RPT")
    else:
        RPT_Cycles = None
    if Para_dict_used.__contains__("Ageing temperature"):
        Temper_i = Para_dict_used["Ageing temperature"]  + 273.15 # update: change to K
        Para_dict_used.pop("Ageing temperature")
    else:
        Temper_i = None
    if Para_dict_used.__contains__("RPT temperature"):
        Temper_RPT = Para_dict_used["RPT temperature"] + 273.15 # update: change to K 
        Para_dict_used.pop("RPT temperature")
    else:
        Temper_RPT = None
    if Para_dict_used.__contains__("Mesh list"):
        mesh_list = Para_dict_used["Mesh list"]
        mesh_list = json.loads(mesh_list)  
        Para_dict_used.pop("Mesh list")
    if Para_dict_used.__contains__("Exponential mesh stretch"):
        submesh_strech = Para_dict_used["Exponential mesh stretch"]  
        Para_dict_used.pop("Exponential mesh stretch")
    else:
        submesh_strech = "nan";
    if Para_dict_used.__contains__("Model option"):
        model_options = Para_dict_used["Model option"] 
        # careful: need to change str to dict 
        model_options = ast.literal_eval(model_options)
        Para_dict_used.pop("Model option")

    if Para_dict_used.__contains__("Initial Neg SOC"):
        c_Neg1SOC_in = (
            Para_0["Maximum concentration in negative electrode [mol.m-3]"]
            *Para_dict_used["Initial Neg SOC"]  )
        Para_0.update(
            {"Initial concentration in negative electrode [mol.m-3]":
            c_Neg1SOC_in})
        Para_dict_used.pop("Initial Neg SOC")
    if Para_dict_used.__contains__("Initial Pos SOC"):    
        c_Pos1SOC_in = (
            Para_0["Maximum concentration in positive electrode [mol.m-3]"]
            *Para_dict_used["Initial Pos SOC"]  )
        Para_0.update(
            {"Initial concentration in positive electrode [mol.m-3]":
            c_Pos1SOC_in})
        Para_dict_used.pop("Initial Pos SOC")
    # 'Outer SEI partial molar volume [m3.mol-1]'
    if Para_dict_used.__contains__(
        "Outer SEI partial molar volume [m3.mol-1]"): 
        Para_0.update(
            {"Outer SEI partial molar volume [m3.mol-1]":
            Para_dict_used["Outer SEI partial molar volume [m3.mol-1]"]})
        Para_0.update(
            {"Inner SEI partial molar volume [m3.mol-1]":
            Para_dict_used["Outer SEI partial molar volume [m3.mol-1]"]})
        
        Para_dict_used.pop("Outer SEI partial molar volume [m3.mol-1]")


    CyclePack = [ 
        Total_Cycles,Cycle_bt_RPT,Update_Cycles,RPT_Cycles,
        Temper_i,Temper_RPT,mesh_list,submesh_strech,model_options];
    # Mark Ruihe - updated 230222 - from P3
    for key, value in Para_dict_used.items():
        # risk: will update parameter that doesn't exist, 
        # so need to make sure the name is right 
        if isinstance(value, str):
            Para_0.update({key: eval(value)})
            #Para_dict_used.pop(key)
        else:
            Para_0.update({key: value},check_already_exists=False)
    # update ambient temperature with "ageing temperature":
    Para_0.update(   {'Ambient temperature [K]':Temper_i });

    return CyclePack,Para_0

# Add 220808 - to simplify the post-processing
def GetSol_dict (my_dict, keys_all, Sol, 
    cycle_no,step_CD, step_CC , step_RE, step_CV ):

    [keys_loc,keys_tim,keys_cyc]  = keys_all;
    # get time_based variables:
    if len(keys_tim): 
        for key in keys_tim:
            step_no = eval("step_{}".format(key[0:2]))
            if key[3:] == "Time [h]":
                my_dict[key].append  (  (
                    Sol.cycles[cycle_no].steps[step_no][key[3:]].entries
                    -
                    Sol.cycles[cycle_no].steps[step_no][key[3:]].entries[0]).tolist()  )
            else:
                my_dict[key].append(  (
                    Sol.cycles[cycle_no].steps[step_no][key[3:]].entries).tolist()  )
    # get cycle_step_based variables: # isn't an array
    if len(keys_cyc): 
        for key in keys_cyc:
            if key in ["Discharge capacity [A.h]"]:
                step_no = step_CD;
                my_dict[key].append  (
                    Sol.cycles[cycle_no].steps[step_no][key].entries[-1]
                    - 
                    Sol.cycles[cycle_no].steps[step_no][key].entries[0])
            elif key in ["Throughput capacity [A.h]"]: # 
                i_try = 0
                while i_try<3:
                    try:
                        getSth = Sol[key].entries[-1]
                    except:
                        i_try += 1
                        print(f"Fail to read Throughput capacity for the {i_try}th time")
                    else:
                        break
                my_dict[key].append(abs(getSth))
            elif key[0:5] in ["CDend","CCend","CVend","REend",
                "CDsta","CCsta","CVsta","REsta",]:
                step_no = eval("step_{}".format(key[0:2]))
                if key[2:5] == "sta":
                    my_dict[key].append  (
                        Sol.cycles[cycle_no].steps[step_no][key[6:]].entries[0])
                elif key[2:5] == "end":
                    my_dict[key].append  (
                        Sol.cycles[cycle_no].steps[step_no][key[6:]].entries[-1])
                
    # get location_based variables:
    if len(keys_loc): 
        for key in keys_loc:
            if key in ["x_n [m]","x [m]","x_s [m]","x_p [m]"]:
                #print("These variables only add once")
                if not len(my_dict[key]):   # special: add only once
                    my_dict[key] = (Sol[key].entries[:,-1]).tolist()
            elif key[0:5] in ["CDend","CCend","CVend","REend",
                            "CDsta","CCsta","CVsta","REsta",]:      
                #print("These variables add multiple times")
                step_no = eval("step_{}".format(key[0:2]))
                if key[2:5] == "sta":
                    my_dict[key].append  ((
                        Sol.cycles[cycle_no].steps[step_no][key[6:]].entries[:,0]).tolist() )
                elif key[2:5] == "end":
                    my_dict[key].append ( (
                        Sol.cycles[cycle_no].steps[step_no][key[6:]].entries[:,-1]).tolist()  )
    return my_dict                              

############## Get initial cap ############## 
# Input: just parameter set
# Output: 0% and 100% SOC of neg/pos; cap at last RPT cycle
# Simply run RPT cycle for 3 times
def Get_initial_cap(Para_dict_i,Neg1SOC_in,Pos1SOC_in,):
    # Para_dict_i contains all global settings
    CyclePack,Para_0 = Para_init(Para_dict_i)
    [
        Total_Cycles,Cycle_bt_RPT,Update_Cycles,
        RPT_Cycles,Temper_i,Temper_RPT,
        mesh_list,submesh_strech,
        model_options] = CyclePack;

    # define experiment
    V_max = 4.2;        
    V_min = 2.5; 
    exp_RPT_text = [ (
        f"Discharge at 0.1C until {V_min} V",  
        "Rest for 1 hours",  
        f"Charge at 0.1C until {V_max} V" ) ] # (5 minute period)
    Experiment_RPT    = pb.Experiment( exp_RPT_text * RPT_Cycles     ) 

    import ast
    model_option_dict = ast.literal_eval(model_options)
    Model_0 = pb.lithium_ion.DFN(options=model_option_dict)

    var = pb.standard_spatial_vars  
    var_pts = {
        var.x_n: int(mesh_list[0]),  
        var.x_s: int(mesh_list[1]),  
        var.x_p: int(mesh_list[2]),  
        var.r_n: int(mesh_list[3]),  
        var.r_p: int(mesh_list[4]),  }       
    submesh_types = Model_0.default_submesh_types
    if submesh_strech == "nan":
        pass
    else:
        particle_mesh = pb.MeshGenerator(
            pb.Exponential1DSubMesh, 
            submesh_params={"side": "right", "stretch": submesh_strech})
        submesh_types["negative particle"] = particle_mesh
        submesh_types["positive particle"] = particle_mesh 
    
    c_Neg1SOC_in = (
        Para_0["Maximum concentration in negative electrode [mol.m-3]"]
        *Neg1SOC_in)
    c_Pos1SOC_in = (
        Para_0["Maximum concentration in positive electrode [mol.m-3]"]
        *Pos1SOC_in)
    Para_0.update(
        {"Initial concentration in negative electrode [mol.m-3]":
        c_Neg1SOC_in})
    Para_0.update(
        {"Initial concentration in positive electrode [mol.m-3]":
        c_Pos1SOC_in})
    Sim_0    = pb.Simulation(
            Model_0,        experiment = Experiment_RPT,
            parameter_values = Para_0,
            solver = pb.CasadiSolver(),
            var_pts=var_pts,
            submesh_types=submesh_types) #mode="safe"
    Sol_0    = Sim_0.solve(calc_esoh=False)

    Cap = [];  Neg1SOC=[]; Pos1SOC = [];

    for i in range(0,RPT_Cycles):
        Cap.append(
            Sol_0.cycles[i].steps[0]["Discharge capacity [A.h]"].entries[-1] - 
            Sol_0.cycles[i].steps[0]["Discharge capacity [A.h]"].entries[0]) 
        Neg1SOC.append(
            Sol_0.cycles[i].steps[0]["Negative electrode stoichiometry"].entries[0]) 
        Pos1SOC.append(
            Sol_0.cycles[i].steps[0]["Positive electrode stoichiometry"].entries[0]) 
        
    return Sol_0, Cap, Neg1SOC, Pos1SOC

def Get_initial_cap2(Para_dict_i):
    # Para_dict_i contains all global settings
    CyclePack,Para_0 = Para_init(Para_dict_i)
    [
        Total_Cycles,Cycle_bt_RPT,Update_Cycles,
        RPT_Cycles,Temper_i,Temper_RPT,
        mesh_list,submesh_strech,
        model_options] = CyclePack;

    # define experiment
    V_max = 4.2;        
    V_min = 2.5; 
    exp_RPT_text = [ (
        f"Discharge at 0.1C until {V_min} V (30 minute period)",  
        "Rest for 1 hours (5 minute period)",  
        f"Charge at 0.1C until {V_max} V (30 minute period)" ) ] # 
    Experiment_RPT    = pb.Experiment( exp_RPT_text * RPT_Cycles     ) 
    import ast
    model_option_dict = ast.literal_eval(model_options)
    Model_0 = pb.lithium_ion.DFN(options=model_option_dict)

    var = pb.standard_spatial_vars  
    var_pts = {
        var.x_n: int(mesh_list[0]),  
        var.x_s: int(mesh_list[1]),  
        var.x_p: int(mesh_list[2]),  
        var.r_n: int(mesh_list[3]),  
        var.r_p: int(mesh_list[4]),  }       
    submesh_types = Model_0.default_submesh_types
    if submesh_strech == "nan":
        pass
    else:
        particle_mesh = pb.MeshGenerator(
            pb.Exponential1DSubMesh, 
            submesh_params={"side": "right", "stretch": submesh_strech})
        submesh_types["negative particle"] = particle_mesh
        submesh_types["positive particle"] = particle_mesh 

    Sim_0    = pb.Simulation(
            Model_0,        experiment = Experiment_RPT,
            parameter_values = Para_0,
            solver = pb.CasadiSolver(),
            var_pts=var_pts,
            submesh_types=submesh_types) #mode="safe"
    Sol_0    = Sim_0.solve(calc_esoh=False)

    Cap = [];  Neg1SOC=[]; Pos1SOC = [];

    for i in range(0,RPT_Cycles):
        Cap.append(
            Sol_0.cycles[i].steps[0]["Discharge capacity [A.h]"].entries[-1] - 
            Sol_0.cycles[i].steps[0]["Discharge capacity [A.h]"].entries[0]) 
        Neg1SOC.append(
            Sol_0.cycles[i].steps[0]["Negative electrode stoichiometry"].entries[0]) 
        Pos1SOC.append(
            Sol_0.cycles[i].steps[0]["Positive electrode stoichiometry"].entries[0]) 
        
    return Sol_0,Cap, Neg1SOC, Pos1SOC

# define the model and run break-in cycle - 
# input parameter: model_options, Experiment_Breakin, Para_0, mesh_list, submesh_strech
# output: Sol_0 , Model_0, Call_Breakin
def Run_Breakin(
    model_options, Experiment_Breakin, 
    Para_0, mesh_list, submesh_strech):

    Model_0 = pb.lithium_ion.DFN(options=model_options)
    
    # update 220926 - add diffusivity and conductivity as variables:
    c_e = Model_0.variables["Electrolyte concentration [mol.m-3]"]
    T = Model_0.variables["Cell temperature [K]"]
    D_e = Para_0["Electrolyte diffusivity [m2.s-1]"]
    sigma_e = Para_0["Electrolyte conductivity [S.m-1]"]
    Model_0.variables["Electrolyte diffusivity [m2.s-1]"] = D_e(c_e, T)
    Model_0.variables["Electrolyte conductivity [S.m-1]"] = sigma_e(c_e, T)
    var = pb.standard_spatial_vars  
    #print("Get here?-2")
    var_pts = {
        var.x_n: int(mesh_list[0]),  
        var.x_s: int(mesh_list[1]),  
        var.x_p: int(mesh_list[2]),  
        var.r_n: int(mesh_list[3]),  
        var.r_p: int(mesh_list[4]),  }       
    submesh_types = Model_0.default_submesh_types
    if submesh_strech == "nan":
        pass
    else:
        particle_mesh = pb.MeshGenerator(
            pb.Exponential1DSubMesh, 
            submesh_params={"side": "right", "stretch": submesh_strech})
        submesh_types["negative particle"] = particle_mesh
        submesh_types["positive particle"] = particle_mesh 
    Sim_0    = pb.Simulation(
        Model_0,        experiment = Experiment_Breakin,
        parameter_values = Para_0,
        solver = pb.CasadiSolver(),
        var_pts=var_pts,
        submesh_types=submesh_types) #mode="safe"
    Call_Breakin = RioCallback()
    try:
        #print("Get here?-1")
        Sol_0    = Sim_0.solve(calc_esoh=False,callbacks=Call_Breakin)
    except (
        pb.expression_tree.exceptions.ModelError,
        pb.expression_tree.exceptions.SolverError
        ) as e:
        Sol_0 = "Model error or solver error"
        error_logs = str(e)
    else:
        error_logs = "nan"
    Result_list_breakin = [Model_0,Sol_0,Call_Breakin,error_logs]

    return Result_list_breakin

# Input: Para_0
# Output: mdic_dry, Para_0
def Initialize_mdic_dry(Para_0,Int_ElelyExces_Ratio):
    mdic_dry = {
        "CeEC_All": [],
        "c_EC_r_new_All": [],
        "c_e_r_new_All": [],
        "Ratio_CeEC_All":[],
        "Ratio_CeLi_All":[],
        "Ratio_Dryout_All":[],

        "Vol_Elely_Tot_All": [],
        "Vol_Elely_JR_All":[],
        "Vol_Pore_tot_All": [],        
        "Vol_EC_consumed_All":[],
        "Vol_Elely_need_All":[],
        "Width_all":[],
        "Vol_Elely_add_All":[],
        "Vol_Pore_decrease_All":[],
        "Test_V_All":[],
        "Test_V2_All":[],
    }
    T_0                  =  Para_0['Initial temperature [K]']
    Porosity_Neg_0       =  Para_0['Negative electrode porosity']  
    Porosity_Pos_0       =  Para_0['Positive electrode porosity']  
    Porosity_Sep_0       =  Para_0['Separator porosity']  
    cs_Neg_Max           =  Para_0["Maximum concentration in negative electrode [mol.m-3]"];
    L_p                  =  Para_0["Positive electrode thickness [m]"]
    L_n                  =  Para_0["Negative electrode thickness [m]"]
    L_s                  =  Para_0["Separator thickness [m]"]
    L_y                  =  Para_0["Electrode width [m]"]
    Para_0.update({'Initial Electrode width [m]':L_y}, check_already_exists=False)
    L_y_0                =  Para_0["Initial Electrode width [m]"]
    L_z                  =  Para_0["Electrode height [m]"]
    Para_0.update({'Initial Electrode height [m]':L_z}, check_already_exists=False)
    L_z_0                =  Para_0["Initial Electrode height [m]"]
    
    Vol_Elely_Tot        = (
        ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  
        * L_y_0 * L_z_0 * Int_ElelyExces_Ratio
     ) # Set initial electrolyte amount [L] 
    Vol_Elely_JR         =  (
        ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  
        * L_y_0 * L_z_0 )
    Vol_Pore_tot         =  (
        ( L_n*Porosity_Neg_0 +  L_p*Porosity_Pos_0  +  L_s*Porosity_Sep_0  )  
        * L_y_0 * L_z_0 )
    Ratio_CeEC           =  1.0; Ratio_CeLi =  1.0  ;Ratio_Dryout         =  1.0
    Vol_EC_consumed      =  0;Vol_Elely_need       =  0;Vol_Elely_add        =  0
    Vol_Pore_decrease    =  0;Test_V2 = 0; Test_V = 0;
    print('Initial electrolyte amount is ', Vol_Elely_Tot*1e6, 'mL') 
    Para_0.update(
        {'Current total electrolyte volume in jelly roll [m3]':
        Vol_Elely_JR}, check_already_exists=False)
    Para_0.update(
        {'Current total electrolyte volume in whole cell [m3]':
        Vol_Elely_Tot}, check_already_exists=False)   
    
    mdic_dry["Vol_Elely_Tot_All"].append(Vol_Elely_Tot*1e6);            
    mdic_dry["Vol_Elely_JR_All"].append(Vol_Elely_JR*1e6);     
    mdic_dry["Vol_Pore_tot_All"].append(Vol_Pore_tot*1e6);           
    mdic_dry["Ratio_CeEC_All"].append(Ratio_CeEC);                      
    mdic_dry["Ratio_CeLi_All"].append(Ratio_CeLi);             
    mdic_dry["Ratio_Dryout_All"].append(Ratio_Dryout);
    mdic_dry["Vol_EC_consumed_All"].append(Vol_EC_consumed*1e6);        
    mdic_dry["Vol_Elely_need_All"].append(Vol_Elely_need*1e6);     
    mdic_dry["Width_all"].append(L_y_0);
    mdic_dry["Vol_Elely_add_All"].append(Vol_Elely_add*1e6);            
    mdic_dry["Vol_Pore_decrease_All"].append(Vol_Pore_decrease*1e6);
    mdic_dry["Test_V_All"].append(Test_V*1e6); 
    mdic_dry["Test_V2_All"].append(Test_V2*1e6); 
    mdic_dry["c_e_r_new_All"].append(Para_0["Initial concentration in electrolyte [mol.m-3]"]); 
    mdic_dry["c_EC_r_new_All"].append(Para_0['EC initial concentration in electrolyte [mol.m-3]'])

    return mdic_dry,Para_0
    
def Update_mdic_dry(Data_Pack,mdic_dry):
    [
        Vol_EC_consumed, 
        Vol_Elely_need, 
        Test_V, 
        Test_V2, 
        Vol_Elely_add, 
        Vol_Elely_Tot_new, 
        Vol_Elely_JR_new, 
        Vol_Pore_tot_new, 
        Vol_Pore_decrease, 
        c_e_r_new, c_EC_r_new,
        Ratio_Dryout, Ratio_CeEC_JR, 
        Ratio_CeLi_JR,
        Width_new, ]= Data_Pack;
    mdic_dry["Vol_Elely_Tot_All"].append(Vol_Elely_Tot_new*1e6);            
    mdic_dry["Vol_Elely_JR_All"].append(Vol_Elely_JR_new*1e6);     
    mdic_dry["Vol_Pore_tot_All"].append(Vol_Pore_tot_new*1e6);           
    mdic_dry["Ratio_CeEC_All"].append(Ratio_CeEC_JR);                      
    mdic_dry["Ratio_CeLi_All"].append(Ratio_CeLi_JR);             
    mdic_dry["Ratio_Dryout_All"].append(Ratio_Dryout);
    mdic_dry["Vol_EC_consumed_All"].append(Vol_EC_consumed*1e6);        
    mdic_dry["Vol_Elely_need_All"].append(Vol_Elely_need*1e6);     
    mdic_dry["Width_all"].append(Width_new);
    mdic_dry["Vol_Elely_add_All"].append(Vol_Elely_add*1e6);            
    mdic_dry["Vol_Pore_decrease_All"].append(Vol_Pore_decrease*1e6);
    mdic_dry["Test_V_All"].append(Test_V*1e6); 
    mdic_dry["Test_V2_All"].append(Test_V2*1e6); 
    mdic_dry["c_e_r_new_All"].append(c_e_r_new); 
    mdic_dry["c_EC_r_new_All"].append(c_EC_r_new)

    return mdic_dry

def Get_Values_Excel(
    Pass_Fail,
    mpe_tot,punish,
    # Start_shape,Middle_shape,End_shape, # add 230526
    model_options,my_dict_RPT,mdic_dry,
    DryOut,Scan_i,Para_dict_i,str_exp_AGE_text,
    str_exp_RPT_text,str_error_AGE_final,
    str_error_RPT,details):

    if model_options.__contains__("SEI on cracks"):
        LossCap_seioncrack = my_dict_RPT["CDend Loss of capacity to SEI on cracks [A.h]"][-1]
    else:
        LossCap_seioncrack = "nan"
    if model_options.__contains__("lithium plating"):
        LossCap_LiP = my_dict_RPT["CDend Loss of capacity to lithium plating [A.h]"][-1]
    else:
        LossCap_LiP = "nan"
    if model_options.__contains__("SEI"):
        LossCap_SEI = my_dict_RPT["CDend Loss of capacity to SEI [A.h]"][-1]
    else:
        LossCap_SEI = "nan"
    if DryOut == "On":
        Vol_Elely_Tot_All_final = mdic_dry["Vol_Elely_Tot_All"][-1];
        Vol_Elely_JR_All_final  = mdic_dry["Vol_Elely_JR_All"][-1];
        Width_all_final         = mdic_dry["Width_all"][-1];
    else:
        Vol_Elely_Tot_All_final = "nan"
        Vol_Elely_JR_All_final  = "nan"
        Width_all_final         = "nan"
    value_list_temp = list(Para_dict_i.values())
    values_para = []
    for value_list_temp_i in value_list_temp:
        values_para.append(str(value_list_temp_i))
    # sequence: scan no, exp, pass or fail, mpe, dry-out, 
    value_Pre = [
        str(Scan_i),Pass_Fail,
        mpe_tot,punish,
        # Start_shape,Middle_shape,End_shape, 
        DryOut,]
    values_pos = [
        str_exp_AGE_text,
        str_exp_RPT_text,
        str(my_dict_RPT["Discharge capacity [A.h]"][0] 
        - 
        my_dict_RPT["Discharge capacity [A.h]"][-1]),

        str(LossCap_LiP),
        str(LossCap_SEI),
        str(LossCap_seioncrack),

        str(my_dict_RPT["CDend Negative electrode capacity [A.h]"][0] 
        - 
        my_dict_RPT["CDend Negative electrode capacity [A.h]"][-1]),

        str(my_dict_RPT["CDend Positive electrode capacity [A.h]"][0] 
        - 
        my_dict_RPT["CDend Positive electrode capacity [A.h]"][-1]),

        str(Vol_Elely_Tot_All_final), 
        str(Vol_Elely_JR_All_final),
        str(Width_all_final),
        str_error_AGE_final, # Mark Ruihe
        str_error_RPT,details   # Mark Ruihe
        ]
    values = [*value_Pre,*values_para,*values_pos]
    return values        

# Function to read exp
def Read_Exp(BasicPath,Exp_Any_Cell,Exp_Path,Exp_head,Exp_Any_Temp,i):
    Exp_Any_AllData  = {}
    for cell in Exp_Any_Cell:
        Exp_Any_AllData[cell] = {} # one cell
        # For extracted directly measured capacity, resistance, etc.
        Exp_Any_AllData[cell]["Extract Data"] = pd.read_csv(
            BasicPath+Exp_Path[i]+
            f"{Exp_head[i]} - cell {cell} ({Exp_Any_Temp[cell]}degC) - Extracted Data.csv", 
            index_col=0)
        Exp_Any_AllData[cell]["Extract Data"].loc[     # update 230617- fill this with avg T
            0, 'Age set average temperature (degC)'] = float(Exp_Any_Temp[cell])
        # Read for DMA results, further a dictionary
        Exp_Any_AllData[cell]["DMA"] = {}
        Exp_Any_AllData[cell]["DMA"]["Cap_Offset"]=pd.read_csv(
            BasicPath+Exp_Path[i]+ "DMA Output/" + f"cell {cell}/" + 
            f"{Exp_head[i]} - Capacity and offset data from OCV-fitting for cell {cell}.csv", 
            index_col=0)
        Exp_Any_AllData[cell]["DMA"]["LLI_LAM"]=pd.read_csv(
            BasicPath+Exp_Path[i]+ "DMA Output/" + f"cell {cell}/" + 
            f"{Exp_head[i]} - DM data from OCV-fitting for cell {cell}.csv", 
            index_col=0)
        Exp_Any_AllData[cell]["DMA"]["Fit_SOC"]=pd.read_csv(
            BasicPath+Exp_Path[i]+ "DMA Output/" + f"cell {cell}/" + 
            f"{Exp_head[i]} - fitting parameters from OCV-fitting for cell {cell}.csv", 
            index_col=0)
        Exp_Any_AllData[cell]["DMA"]["RMSE"]=pd.read_csv(
            BasicPath+Exp_Path[i]+ "DMA Output/" + f"cell {cell}/" + 
            f"{Exp_head[i]} - RMSE data from OCV-fitting for cell {cell}.csv", 
            index_col=0)
    print("Finish reading Experiment!")
    return Exp_Any_AllData
# judge ageing shape: - NOT Ready yet!
def check_concave_convex(x_values, y_values):
    # Check if the series has at least three points
    if len(x_values) < 3 or len(y_values) < 3:
        return ["None","None""None"]

    # Calculate the slopes between adjacent points based on "x" and "y" values
    slopes = []
    for i in range(1, len(y_values)):
        slope = (y_values[i] - y_values[i - 1]) / (x_values[i] - x_values[i - 1])
        slopes.append(slope)

    # Determine the indices for the start, middle, and end groups
    start_index = 0
    middle_index = len(y_values) // 2 - 1
    end_index = len(y_values) - 3

    # Check if the start group is concave, convex, or linear
    start_group = slopes[:3]
    start_avg = sum(start_group[:2]) / 2
    start_point = start_group[1]
    # TODO we still need to consider a proper index to judge linear
    if abs(start_avg - start_point) <= 0.005 * start_point:
        start_shape = "Linear"
    elif start_avg <= start_point:
        start_shape = "Sub"
    else:
        start_shape = "Super"

    # Check if the middle group is concave, convex, or linear
    middle_group = slopes[middle_index:middle_index + 3]
    middle_avg = sum(middle_group[:2]) / 2
    middle_point = middle_group[1]
    if abs(middle_avg - middle_point) <= 0.005 * middle_point:
        middle_shape = "Linear"
    elif middle_avg <= middle_point:
        middle_shape = "Sub"
    else:
        middle_shape = "Super"

    # Check if the end group is concave, convex, or linear
    end_group = slopes[end_index:]
    end_avg = sum(end_group[:2]) / 2
    end_point = end_group[1]
    if abs(end_avg - end_point) <= 0.005 * end_point:
        end_shape = "Linear"
    elif end_avg <= end_point:
        end_shape = "Sub"
    else:
        end_shape = "Super"

    # Return the shape results
    shape_results = [start_shape, middle_shape, end_shape]
    return shape_results

# plot inside the function:
def Plot_Cyc_RPT_4(
        my_dict_RPT, XY_Exp, Plot_Exp,  
        Scan_i,Temper_i,model_options,
        BasicPath, Target,fs,dpi):
    
    Num_subplot = 5;
    fig, axs = plt.subplots(2,1, figsize=(5,7.8),tight_layout=True)
    axs[0].plot(
        my_dict_RPT['Throughput capacity [kA.h]'], 
        my_dict_RPT['CDend SOH [%]'],     
        '-o', label="Scan=" + str(Scan_i) )
    axs[1].plot(
        my_dict_RPT["Throughput capacity [kA.h]"][1:], 
        np.array(my_dict_RPT["avg_Age_T"]),     '-o', ) 
    # Plot Charge Throughput (A.h) vs SOH
    color_exp     = [0, 0, 0, 0.3]; marker_exp     = "v";
    color_exp_Avg = [0, 0, 0, 0.7]; marker_exp_Avg = "s";
    if Plot_Exp == True:
        # Update 230518: Plot Experiment Average - at 1 expeirment and 1 temperature
        [X_exp,Y_exp] = XY_Exp
        axs[0].plot(
            X_exp,Y_exp,color=color_exp_Avg,
            marker=marker_exp_Avg,label=f"Exp") 
    axs[0].set_ylabel("SOH %")
    axs[1].set_ylabel(r"Avg age T [$^\circ$C]")
    axs[0].set_xlabel("Charge Throughput (kA.h)")
    for i in range(0,2):
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); 
        [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1);del labels
    axs[0].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
    fig.suptitle(
        f"Scan_{Scan_i}-{str(int(Temper_i- 273.15))}"
        +r"$^\circ$C - Summary", fontsize=fs+2)
    plt.savefig(
        BasicPath + Target+    "Plots/" +  
        f"0_Scan_{Scan_i}-{str(int(Temper_i- 273.15))}degC Summary.png", dpi=dpi)
    plt.close()  # close the figure to save RAM

    if model_options.__contains__("SEI on cracks"):
        """ Num_subplot = 2;
        fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
        axs[0].plot(my_dict_RPT['Throughput capacity [kA.h]'], my_dict_RPT["CDend X-averaged total SEI on cracks thickness [m]"],     '-o', label="Scan=" + str(Scan_i) )
        axs[1].plot(my_dict_RPT['Throughput capacity [kA.h]'], my_dict_RPT["CDend X-averaged negative electrode roughness ratio"],'-o', label="Scan=" + str(Scan_i) )
        axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[1].set_ylabel("Roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs})
        for i in range(0,Num_subplot):
            axs[i].set_xlabel("Charge Throughput (kA.h)",   fontdict={'family':'DejaVu Sans','size':fs})
            labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
            axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
            axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
        axs[0].set_title("X-avg tot Neg SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
        axs[1].set_title("X-avg Neg roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
        plt.savefig(BasicPath + Target+"Plots/" +
            f"Scan_{Scan_i}-{str(int(Temper_i- 273.15))}degC - Cracks related_Scan.png", dpi=dpi)
        plt.close()  # close the figure to save RAM """

        Num_subplot = 2;
        fig, axs = plt.subplots(1,Num_subplot, figsize=(12,4.8),tight_layout=True)
        axs[0].plot(my_dict_RPT['Throughput capacity [kA.h]'], 
            my_dict_RPT["CDend Negative electrode capacity [A.h]"][0]
            -
            my_dict_RPT["CDend Negative electrode capacity [A.h]"],'-o',label="Neg Scan=" + str(Scan_i))
        axs[1].plot(my_dict_RPT['Throughput capacity [kA.h]'], 
            my_dict_RPT["CDend Positive electrode capacity [A.h]"][0]
            -
            my_dict_RPT["CDend Positive electrode capacity [A.h]"],'-^',label="Pos Scan=" + str(Scan_i))
        """ axs[0].plot(
            my_dict_RPT['Throughput capacity [kA.h]'], 
            my_dict_RPT["CDend X-averaged total SEI on cracks thickness [m]"],                  
            '-o',label="Scan="+ str(Scan_i)) """
        for i in range(0,2):
            axs[i].set_xlabel("Charge Throughput (kA.h)",   fontdict={'family':'DejaVu Sans','size':fs})
            labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
            axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
            axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
        #axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
        #axs[0].set_title("CDend X-avg tot SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
        for i in range(0,2):
            axs[i].set_xlabel("Charge Throughput (kA.h)",   fontdict={'family':'DejaVu Sans','size':fs})
            axs[i].set_ylabel("Capacity [A.h]",   fontdict={'family':'DejaVu Sans','size':fs})
            labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
            axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
            axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)
            axs[i].set_title("LAM of Neg and Pos",   fontdict={'family':'DejaVu Sans','size':fs+1})
        plt.savefig(BasicPath + Target+"Plots/" +
            f"Scan_{Scan_i}-{str(int(Temper_i- 273.15))}degC LAM-IR.png", dpi=dpi)
        plt.close()  # close the figure to save RAM
    Num_subplot = 2;
    fig, axs = plt.subplots(1,Num_subplot, figsize=(8,3.2),tight_layout=True)
    axs[0].plot(my_dict_RPT['Throughput capacity [kA.h]'], my_dict_RPT["CDsta Positive electrode stoichiometry"] ,'-o',label="Start" )
    axs[0].plot(my_dict_RPT['Throughput capacity [kA.h]'], my_dict_RPT["CDend Positive electrode stoichiometry"] ,'-^',label="End" )
    axs[1].plot(my_dict_RPT['Throughput capacity [kA.h]'], my_dict_RPT["CDsta Negative electrode stoichiometry"],'-o',label="Start" )
    axs[1].plot(my_dict_RPT['Throughput capacity [kA.h]'], my_dict_RPT["CDend Negative electrode stoichiometry"],'-^',label="End" )
    for i in range(0,2):
        axs[i].set_xlabel("Charge Throughput (kA.h)",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[i].set_ylabel("Stoichiometry",   fontdict={'family':'DejaVu Sans','size':fs})
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].ticklabel_format(style='sci', axis='x', scilimits=(-1e-2,1e-2))
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
        axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)     
    axs[0].set_title("Neg Sto. range (Dis)",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[1].set_title("Pos Sto. range (Dis)",   fontdict={'family':'DejaVu Sans','size':fs+1})
    plt.savefig(BasicPath + Target+"Plots/"+
        f"Scan_{Scan_i}-{str(int(Temper_i- 273.15))}degC SOC_RPT_dis.png", dpi=dpi) 
    plt.close()  # close the figure to save RAM

    return

def Plot_Loc_AGE_4(my_dict_AGE,Scan_i,Temper_i,model_options,BasicPath, Target,fs,dpi):
    Num_subplot = 2;
    fig, axs = plt.subplots(1,Num_subplot, figsize=(8,3.2),tight_layout=True)
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Porosity"][0],'-o',label="First")
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Porosity"][-1],'-^',label="Last"  )
    axs[1].plot(
        my_dict_AGE["x_n [m]"],
        my_dict_AGE["CDend Negative electrode reaction overpotential [V]"][0],'-o',label="First" )
    axs[1].plot(
        my_dict_AGE["x_n [m]"],
        my_dict_AGE["CDend Negative electrode reaction overpotential [V]"][-1],'-^',label="Last" )

    axs[0].set_xlabel("Dimensional Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[1].set_xlabel("Dimensional Neg thickness",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[0].set_title("Porosity",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[1].set_title("Neg electrode reaction overpotential",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[0].set_ylabel("Porosity",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[1].set_ylabel("Overpotential [V]",   fontdict={'family':'DejaVu Sans','size':fs})
    for i in range(0,2):
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].ticklabel_format(style='sci', axis='x', scilimits=(-1e-2,1e-2))
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
        axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
    plt.savefig(BasicPath + Target+"Plots/" +
        f"Scan_{Scan_i}-{str(int(Temper_i- 273.15))}degC Por Neg_S_eta.png", dpi=dpi) 
    plt.close()  # close the figure to save RAM

    """ update: disable cracking related things just because when sei-on-crack is false it will easily give errors; 
    if model_options.__contains__("SEI on cracks"):
        Num_subplot = 2;
        fig, axs = plt.subplots(1,Num_subplot, figsize=(8,3.2),tight_layout=True)
        axs[0].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Negative electrode roughness ratio"][0],'-o',label="First")
        axs[0].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Negative electrode roughness ratio"][-1],'-^',label="Last"  )
        axs[1].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Total SEI on cracks thickness [m]"][0],'-o',label="First" )
        axs[1].plot(my_dict_AGE["x_n [m]"], my_dict_AGE["CDend Total SEI on cracks thickness [m]"][-1],'-^',label="Last" )
        axs[0].set_title("Tot Neg SEI on cracks thickness",   fontdict={'family':'DejaVu Sans','size':fs+1})
        axs[1].set_title("Neg roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
        axs[0].set_ylabel("SEI on cracks thickness [m]",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[1].set_ylabel("Roughness ratio",   fontdict={'family':'DejaVu Sans','size':fs})
        for i in range(0,2):
            axs[i].set_xlabel("Dimensional Neg thickness",   fontdict={'family':'DejaVu Sans','size':fs})
            labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
            axs[i].ticklabel_format(style='sci', axis='x', scilimits=(-1e-2,1e-2))
            axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
            axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
        plt.savefig(BasicPath + Target+"Plots/" +
            f"Scan_{Scan_i}-{str(int(Temper_i- 273.15))}degC Cracks related spatial.png", dpi=dpi) 
        plt.close()  # close the figure to save RAM """
    Num_subplot = 2;
    fig, axs = plt.subplots(1,Num_subplot, figsize=(8,3.2),tight_layout=True)
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte concentration [mol.m-3]"][0],'-o',label="First")
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte concentration [mol.m-3]"][-1],'-^',label="Last"  )
    axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte potential [V]"][0],'-o',label="First" )
    axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte potential [V]"][-1],'-^',label="Last" )
    axs[0].set_title("Electrolyte concentration",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[1].set_title("Electrolyte potential",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[0].set_ylabel("Concentration [mol.m-3]",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[1].set_ylabel("Potential [V]",   fontdict={'family':'DejaVu Sans','size':fs})
    for i in range(0,2):
        axs[i].set_xlabel("Dimensional Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].ticklabel_format(style='sci', axis='x', scilimits=(-1e-2,1e-2))
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
        axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
    plt.savefig(BasicPath + Target+"Plots/" +
            f"Scan_{Scan_i}-{str(int(Temper_i- 273.15))}degC Electrolyte concentration and potential.png", dpi=dpi)
    plt.close()  # close the figure to save RAM
    Num_subplot = 2;
    fig, axs = plt.subplots(1,Num_subplot, figsize=(8,3.2),tight_layout=True)
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte diffusivity [m2.s-1]"][0],'-o',label="First")
    axs[0].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte diffusivity [m2.s-1]"][-1],'-^',label="Last"  )
    axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte conductivity [S.m-1]"][0],'-o',label="First" )
    axs[1].plot(my_dict_AGE["x [m]"], my_dict_AGE["CDend Electrolyte conductivity [S.m-1]"][-1],'-^',label="Last" )
    axs[0].set_title("Electrolyte diffusivity",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[1].set_title("Electrolyte conductivity",   fontdict={'family':'DejaVu Sans','size':fs+1})
    axs[0].set_ylabel("Diffusivity [m2.s-1]",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[1].set_ylabel("Conductivity [S.m-1]",   fontdict={'family':'DejaVu Sans','size':fs})
    for i in range(0,2):
        axs[i].set_xlabel("Dimensional Cell thickness",   fontdict={'family':'DejaVu Sans','size':fs})
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].ticklabel_format(style='sci', axis='x', scilimits=(-1e-2,1e-2))
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
        axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
    plt.savefig(BasicPath + Target+"Plots/" +
            f"Scan_{Scan_i}-{str(int(Temper_i- 273.15))}degC Electrolyte diffusivity and conductivity.png", dpi=dpi)
    plt.close()  # close the figure to save RAM
    return

def Plot_Dryout(
    Cyc_Update_Index,mdic_dry,ce_EC_0,Temper_i,
    Scan_i,BasicPath, Target,fs,dpi):
    
    CeEC_All =np.full(np.size(mdic_dry["Ratio_CeEC_All"]),ce_EC_0); 
    for i in range(1,np.size(mdic_dry["Ratio_CeEC_All"])):
        for k in range(0,i):
            CeEC_All[i] *= mdic_dry["Ratio_CeEC_All"][k]
    mdic_dry["CeEC_All"]= CeEC_All.tolist()

    Num_subplot = 3;
    fig, axs = plt.subplots(1,Num_subplot, figsize=(12,3.2),tight_layout=True)
    axs[0].plot(Cyc_Update_Index, mdic_dry["Vol_EC_consumed_All"],'-o',label="EC consumed")
    axs[0].plot(Cyc_Update_Index, mdic_dry["Vol_Elely_need_All"],'-.',label="Elely needed")
    axs[0].plot(Cyc_Update_Index, mdic_dry["Vol_Elely_add_All"],'-s',label="Elely added")
    axs[0].plot(Cyc_Update_Index, mdic_dry["Vol_Pore_decrease_All"],'--',label="Pore decreased")
    axs[1].plot(Cyc_Update_Index, mdic_dry["Vol_Elely_Tot_All"],     '-o', label="Total electrolyte in cell" )
    axs[1].plot(Cyc_Update_Index, mdic_dry["Vol_Elely_JR_All"],     '--', label="Total electrolyte in JR" )
    axs[1].plot(Cyc_Update_Index, mdic_dry["Vol_Pore_tot_All"],     '-s', label="Total pore in JR" )
    axs[2].plot(Cyc_Update_Index, mdic_dry["Ratio_Dryout_All"],     '-s', label="Dry out ratio" )
    axs[2].set_ylabel("Ratio",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[2].set_xlabel("Cycle number",   fontdict={'family':'DejaVu Sans','size':fs})
    axs[2].set_title("Dry out ratio",   fontdict={'family':'DejaVu Sans','size':fs+1})
    for i in range(0,2):
        axs[i].set_xlabel("Cycle number",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[i].set_ylabel("Volume [mL]",   fontdict={'family':'DejaVu Sans','size':fs})
        axs[i].set_title("Volume",   fontdict={'family':'DejaVu Sans','size':fs+1})
        labels = axs[i].get_xticklabels() + axs[i].get_yticklabels(); [label.set_fontname('DejaVu Sans') for label in labels]
        axs[i].tick_params(labelcolor='k', labelsize=fs, width=1) ;  del labels;
        axs[i].legend(prop={'family':'DejaVu Sans','size':fs-2},loc='best',frameon=False)    
    plt.savefig(BasicPath + Target+"Plots/" +
        f"Scan_{Scan_i}-{str(int(Temper_i- 273.15))}degC Volume_total.png", 
        dpi=dpi)
    plt.close()  # close the figure to save RAM
    return

# update 230312: add a function to get the discharge capacity and resistance
def Get_0p1s_R0(sol_RPT,Index,cap_full):
    Res_0p1s = []; SOC = [100,];
    for i,index in enumerate(Index):
        cycle = sol_RPT.cycles[index]
        Res_0p1s.append(   (
            np.mean(cycle.steps[1]["Terminal voltage [V]"].entries[-10:-1])
            - cycle.steps[2]["Terminal voltage [V]"].entries[0]
        ) / cycle.steps[2]["Current [A]"].entries[0] * 1000)
        if i > 0:
            Dis_Cap = abs(
                cycle.steps[2]["Discharge capacity [A.h]"].entries[0] 
                - cycle.steps[2]["Discharge capacity [A.h]"].entries[-1] )
            SOC.append(SOC[-1]-Dis_Cap/cap_full*100)
    return np.mean(Res_0p1s),Res_0p1s,SOC

# update 230517 add a function to get R_50%SOC from C/2 discharge
def Get_R_from_0P5C_CD(step_0P5C_CD,cap_full):
    # print("Total data points: ",len(step_0P5C_CD["Time [h]"].entries))
    Dis_Cap = abs(
        step_0P5C_CD["Discharge capacity [A.h]"].entries[0] 
        - step_0P5C_CD["Discharge capacity [A.h]"].entries )
    SOC_0p5C = (1-Dis_Cap/cap_full)*100
    #print(SOC_0p5C)
    V_ohmic = (
    step_0P5C_CD['Battery open-circuit voltage [V]'].entries 
    - step_0P5C_CD["Terminal voltage [V]"].entries
    + step_0P5C_CD["Battery particle concentration overpotential [V]"].entries 
    + step_0P5C_CD["X-averaged battery concentration overpotential [V]" ].entries
    )
    # print("Applied current [A]:",step_0P5C_CD["Current [A]"].entries[0])
    Res_0p5C = V_ohmic/step_0P5C_CD["Current [A]"].entries[0] * 1e3
    Res_0p5C_50SOC = np.interp(50,np.flip(SOC_0p5C),np.flip(Res_0p5C),)
    SOC_0p5C = SOC_0p5C.tolist()
    Res_0p5C = Res_0p5C.tolist()
    Res_0p5C_50SOC = Res_0p5C_50SOC.tolist()
    return SOC_0p5C,Res_0p5C,Res_0p5C_50SOC   #,Rohmic_CD_2

# Update 23-05-18
# function to get cell average index from several cells in one T and one Exp
def Get_Cell_Mean_1T_1Exp(Exp_Any_AllData,Exp_temp_i_cell):
    # Get the cell with smallest X "Charge Throughput (A.h)" - to interpolate
    X_1_last = [];    X_5_last = [];
    for cell in Exp_temp_i_cell:
        df = Exp_Any_AllData[cell]["Extract Data"]
        xtemp = np.array(df["Charge Throughput (A.h)"])/1e3
        X_1_last.append(xtemp[-1])
        index_Res = df[df['0.1s Resistance (Ohms)'].le(10)].index
        xtemp2 = np.array(df["Charge Throughput (A.h)"][index_Res])/1e3
        X_5_last.append(xtemp2[-1])
    index_X1 = np.argmin(X_1_last) # find index of the smallest value
    index_X5 = np.argmin(X_5_last)
    # Get the interpolate ones
    Y_1_st = []; Y_2_st = []; Y_3_st = []; Y_4_st = []; Y_5_st = [];Y_6_st = [];
    df_t1 = Exp_Any_AllData[Exp_temp_i_cell[index_X1]]["Extract Data"]
    X_1_st = np.array(df_t1["Charge Throughput (A.h)"])/1e3
    df_t5 = Exp_Any_AllData[Exp_temp_i_cell[index_X5]]["Extract Data"]
    index_Res_5 = df_t5[df_t5['0.1s Resistance (Ohms)'].le(10)].index
    X_5_st = np.array(df_t5["Charge Throughput (A.h)"][index_Res_5])/1e3
    for cell in Exp_temp_i_cell:
        df = Exp_Any_AllData[cell]["Extract Data"]
        chThr_temp = np.array(df["Charge Throughput (A.h)"])/1e3
        df_DMA = Exp_Any_AllData[cell]["DMA"]["LLI_LAM"]
        Y_1_st.append( list(np.interp(X_1_st,chThr_temp,
                        np.array(df_DMA["SoH"])*100)) )
        Y_2_st.append( list(np.interp(X_1_st,chThr_temp,
                        np.array(df_DMA["LLI"])*100)) )
        Y_3_st.append( list(np.interp(X_1_st,chThr_temp,
                        np.array(df_DMA["LAM NE_tot"])*100)) )
        Y_4_st.append( list(np.interp(X_1_st,chThr_temp,
                        np.array(df_DMA["LAM PE"])*100)) )
        Y_6_st.append(list(
            np.interp(X_1_st, chThr_temp, 
            df["Age set average temperature (degC)"].astype(float))))

        index_Res = df[df['0.1s Resistance (Ohms)'].le(10)].index
        Y_5_st.append( list(np.interp(
            X_5_st,np.array(df["Charge Throughput (A.h)"][index_Res])/1e3,
            np.array(df["0.1s Resistance (Ohms)"][index_Res])*1e3, )) )
    # Get the mean values of the interpolate ones:
    def Get_Mean(X_1_st,Y_1_st):
        Y_1_st_avg = []
        for i in range(len(X_1_st)):
            sum=0
            for j in range(len(Y_1_st)):
                sum += Y_1_st[j][i]
            Y_1_st_avg.append(sum/len(Y_1_st))
        return Y_1_st_avg   
    Y_1_st_avg = Get_Mean(X_1_st,Y_1_st)
    Y_2_st_avg = Get_Mean(X_1_st,Y_2_st)
    Y_3_st_avg = Get_Mean(X_1_st,Y_3_st)
    Y_4_st_avg = Get_Mean(X_1_st,Y_4_st)
    Y_5_st_avg = Get_Mean(X_5_st,Y_5_st)
    Y_6_st_avg = Get_Mean(X_1_st,Y_6_st)
    XY_pack = [
        X_1_st,X_5_st,Y_1_st_avg,Y_2_st_avg,
        Y_3_st_avg,Y_4_st_avg,Y_5_st_avg,Y_6_st_avg]
    return XY_pack

# Update 23-05-18 Compare MPE - code created by ChatGPT
def mean_percentage_error(A, B):
    if 0 in A:
        indices = np.where(A == 0)[0]
        A = np.delete(A, indices)
        B = np.delete(B, indices)
        print(A,B)
    errors = np.abs(A - B) / np.abs(A)
    mpe = np.mean(errors) * 100
    return mpe

# Update 23-05-18
# compare modelling result with modelling:
# idea: do interpolation following TODO this is where weighting works
# initial:: X_1_st,X_5_st,Y_1_st_avg,Y_2_st_avg,Y_3_st_avg,Y_4_st_avg,Y_5_st_avg
# to compare: my_dict_RPT
def Compare_Exp_Model(
        my_dict_RPT, XY_Exp, Scan_i,
        Temper_i,BasicPath, Target,fs,dpi, PlotCheck):
    
    [X_exp,Y_exp] = XY_Exp
    mX_1 = my_dict_RPT['Throughput capacity [kA.h]']
    if mX_1[-1] > X_exp[-1]:
        punish = 1; 
        mX_1_st = X_exp   # do interpolation on modelling result
        mY_1_st = np.interp(mX_1_st,my_dict_RPT['Throughput capacity [kA.h]'], 
            my_dict_RPT['CDend SOH [%]'])
        # experiment result remain unchanged
        Y_1_st_avgm = Y_exp
    else:                # do interpolation on expeirment results
        punish = X_exp[-1] / mX_1[-1]  # punishment error, add when simulation end early
        mX_1_st = mX_1 #  standard for experiment following modelling
        Y_1_st_avgm = np.interp(mX_1_st,X_exp,Y_exp)
        mY_1_st = my_dict_RPT['CDend SOH [%]']
    # Now we can calculate MPE! mean_percentage_error
    mpe_1 = np.sum(abs(np.array(Y_1_st_avgm)-np.array(mY_1_st)))/len(Y_1_st_avgm) # SOH [%]
    mpe_tot = (1.0*mpe_1 +(punish>1.0)* punish * 2 )
    # plot and check:
    if PlotCheck == True:
        fig, axs = plt.subplots( figsize=(6,4),tight_layout=True)
        axs.plot(mX_1_st, mY_1_st,'-o', label="Model" )
        axs.plot(mX_1_st, Y_1_st_avgm,'-o', label="Exp"  )
        axs.legend()
        axs.set_ylabel("SOH %")
        axs.set_xlabel("Charge Throughput (kA.h)")
        fig.suptitle(
            f"Scan {str(Scan_i)}-{str(int(Temper_i-273.15))}"
            +r"$^\circ$C - Calculate Error", fontsize=fs+2)
        plt.savefig(BasicPath + Target
            +"Plots/"+ f"Scan {str(Scan_i)}-{str(int(Temper_i-273.15))}degC"
            +r"- Calculate Error.png", dpi=dpi)
        plt.close()  # close the figure to save RAM
    mpe_all = np.around([mpe_tot,punish],2)
    return mpe_all

def Compare_Exp_Model_1Cyc_Dch(
    my_dict_RPT, XY_Exp, Scan_i,
    Temper_i,BasicPath, Target,fs,dpi, PlotCheck):
    
    [X_exp,Y_exp] = XY_Exp
    mX_1 = my_dict_RPT["CD Time [h]"][0]
    mY_1 = my_dict_RPT["CD Terminal voltage [V]"][0]
    if mX_1[-1] > X_exp[-1]:
        punish = 1; 
        mX_1_st = X_exp   # do interpolation on modelling result
        mY_1_st = np.interp(mX_1_st,mX_1,mY_1)
        # experiment result remain unchanged
        Y_1_st_avgm = Y_exp
    else:                # do interpolation on expeirment results
        punish = X_exp[-1] / mX_1[-1]  # punishment error, add when simulation end early
        mX_1_st = mX_1 #  standard for experiment following modelling
        Y_1_st_avgm = np.interp(mX_1_st,X_exp,Y_exp)
        mY_1_st = mY_1
    # Now we can calculate MPE! mean_percentage_error
    mpe_1 = np.sum(abs(np.array(Y_1_st_avgm)-np.array(mY_1_st)))/len(Y_1_st_avgm) # SOH [%]
    mpe_tot = (1.0*mpe_1 +(punish>1.0)* punish * 2 )
    # plot and check:
    fig, axs = plt.subplots( figsize=(6,4),tight_layout=True)
    axs.plot(mX_1, mY_1,'-', label="Model" )
    axs.plot(X_exp,Y_exp,'-', label="Exp"  )
    axs.legend()
    axs.set_ylabel("Time [h]")
    axs.set_xlabel("Voltage [V]")
    fig.suptitle(
        f"Scan {str(Scan_i)}-{str(int(Temper_i))}"
        +r"$^\circ$C - Original", fontsize=fs+2)
    plt.savefig(BasicPath + Target
        +"Plots/"+ f"Scan {str(Scan_i)}-{str(int(Temper_i))}degC"
        +r"- Compare Original.png", dpi=dpi)
    plt.close()  # close the figure to save RAM

    if PlotCheck == True:
        fig, axs = plt.subplots( figsize=(6,4),tight_layout=True)
        axs.plot(mX_1_st, mY_1_st,'-', label="Model" )
        axs.plot(mX_1_st, Y_1_st_avgm,'-', label="Exp"  )
        axs.legend()
        axs.set_ylabel("Time [h]")
        axs.set_xlabel("Voltage [V]")
        fig.suptitle(
            f"Scan {str(Scan_i)}-{str(int(Temper_i))}"
            +r"$^\circ$C - Calculate Error", fontsize=fs+2)
        plt.savefig(BasicPath + Target
            +"Plots/"+ f"Scan {str(Scan_i)}-{str(int(Temper_i))}degC"
            +r"- Calculate Error.png", dpi=dpi)
        plt.close()  # close the figure to save RAM
    
    mpe_all = np.around([mpe_tot,punish],2)
    return mpe_all

# read scan files:
def load_combinations_from_csv(Para_file):
    dataframe = pd.read_csv(Para_file)
    parameter_names = dataframe.columns.tolist()
    combinations = dataframe.values.tolist()
    return parameter_names, combinations

def RunOneCyc(Para_dict_i,BasicPath, XY_Exp, 
    purpose,    Exp_pack, keys_all,dpi,fs,
    flag_RunOneCyc,   Plot_Exp,Timeout,Return_Sol,
    Check_Small_Time):

    Target  = f'/{purpose}/'
    ModelTimer = pb.Timer()
    if Check_Small_Time == True:
        SmallTimer = pb.Timer()
    
    # update 230731 - catch log information
    log_buffer = io.StringIO()
    # Set up the custom logger to capture the log messages
    custom_logger = logging.getLogger('pybamm')
    custom_handler = logging.StreamHandler(log_buffer)
    custom_logger.addHandler(custom_handler)
    custom_logger.setLevel(logging.WARNING) 
    # Set up a separate logger for your own print statements
    own_logger = logging.getLogger('my_own_logger')
    own_logger.setLevel(logging.INFO)  # Set the log level for your own print statements
    own_handler = logging.StreamHandler()  # Print to the console
    own_logger.addHandler(own_handler)
    
    # Your print statements within the Run_model function
    own_logger.info("Your own print statement")
    ##########################################################
    ##############    Part-1: Initialization    ##############
    ##########################################################
    font = {'family' : 'DejaVu Sans','size'   : fs}
    mpl.rc('font', **font)

    # new to define: exp_text_list,  Path_pack
    # exp_index_pack , 
    # define here:
    index_i   = Para_dict_i["Scan No"]  
    Scan_i = int(index_i)
    own_logger.info('Start Now! Scan %d.' % Scan_i)  
    Temp_K = Para_dict_i["Ageing temperature"]  # for one cycle this is the only temperature 
    
    CyclePack,Para_0 = Para_init(Para_dict_i)
    [ 
        Total_Cycles,_,_,_,
        _,_,mesh_list,submesh_strech,model_options]= CyclePack
    [
        exp_1Cyc_text,_,
        step_1Cyc_CD,step_1Cyc_CC,step_1Cyc_CV,
        _ ,_,_,
        cycle_no,book_name_xlsx,
        ]  = Exp_pack
    [keys_all_RPT,_] = keys_all
    str_exp_1Cyc_text  = str(exp_1Cyc_text)
    Experiment_OneCyc   = pb.Experiment( exp_1Cyc_text * Total_Cycles  )  
    # initialize my_dict for outputs
    my_dict_1Cyc = {}
    for keys in keys_all_RPT:
        for key in keys:
            my_dict_1Cyc[key]=[];
    
    if Check_Small_Time == True:
        own_logger.info(f'Scan {Scan_i}: Spent {SmallTimer.time()} on Initialization')
        SmallTimer.reset()
    
    ##########################################################
    ##############    Part-2: Run model         ##############
    ##########################################################
    ##########################################################
    Timeout_text = 'I timed out'
    ##########    2-1: Define model and run the only cycle
    try:
        #print(model_options)
        #print(type(model_options))
        Timelimit = int(3600*3)
        # the following turns on for HPC only!
        if Timeout == True:
            timeout_RPT = TimeoutFunc(
                Run_Breakin, 
                timeout=Timelimit, 
                timeout_val=Timeout_text)
            Result_list  = timeout_RPT(
                model_options, Experiment_OneCyc, 
                Para_0, mesh_list, submesh_strech)
        else:
            Result_list  = Run_Breakin(
                model_options, Experiment_OneCyc, 
                Para_0, mesh_list, submesh_strech)
        [Model_0,Sol_0,Call_1Cyc,error_logs] = Result_list
        if Return_Sol == True:
            pass
        if Call_1Cyc.success == False:
            Sol_0 = "nan"
            call_logs = str(Call_1Cyc.logs)
            str_error_1Cyc = "Experiment error or infeasible"
            own_logger.info("Fail due to Experiment error or infeasible")
            1/0
        if Sol_0 == Timeout_text: # to do: distinguish different failure cases
            call_logs = "Fail due to Timeout"
            str_error_1Cyc = "Timeout"
            own_logger.info("Fail due to Timeout")
            1/0
        if Sol_0 == "Model error or solver error":
            call_logs = error_logs
            own_logger.info("Fail due to Model error or solver error")
            1/0
    except ZeroDivisionError as e:
        str_error_1Cyc = str(e)
        Sol_0 = "nan"
        if Check_Small_Time == True:
            own_logger.info(f"Scan {Scan_i}: Fail break-in cycle within {SmallTimer.time()}, need to exit the whole scan now due to {str_error_1Cyc} but do not know how!")
            SmallTimer.reset()
        else:
            own_logger.info(f"Scan {Scan_i}: Fail break-in cycle, need to exit the "
            f"whole scan now due to {str_error_1Cyc} but do not know how!")
        Flag_1Cyc = False 
    else:
        str_error_1Cyc  = "nan";  call_logs = "nan";
        if Check_Small_Time == True:    
            own_logger.info(f"Scan {Scan_i}: Finish break-in cycle within {SmallTimer.time()}")
            SmallTimer.reset()
        else:
            own_logger.info(f"Scan {Scan_i}: Finish break-in cycle")
        # post-process for the only cycle
        my_dict_1Cyc = GetSol_dict (my_dict_1Cyc,keys_all_RPT, Sol_0, 
            0, step_1Cyc_CD, step_1Cyc_CC,_ , step_1Cyc_CV   )
        Flag_1Cyc = True
        if Check_Small_Time == True:    
            own_logger.info(f"Scan {Scan_i}: Finish post-process for break-in cycle within {SmallTimer.time()}")
            SmallTimer.reset()
        else:
            own_logger.info(f"Scan {Scan_i}: Finish post-process for break-in cycle")
    ##########################################################
    ##############   Part-3: Post-prosessing    ##############
    ##########################################################
    if Flag_1Cyc == False: 
        # Get the log messages from the buffer and store it in a variable
        log_messages = log_buffer.getvalue()
        log_messages = log_messages.strip() # # Remove leading and trailing spaces from the log message
        # Remove the custom handler to prevent duplicate logs
        custom_logger.removeHandler(custom_handler)

        value_list_temp = list(Para_dict_i.values())
        values_para = []
        for value_list_temp_i in value_list_temp:
            values_para.append(str(value_list_temp_i))
        # sequence: scan no, exp, pass or fail, mpe, dry-out, 
        mpe_all = ["nan","nan",]
        Pass_Fail = "Die"
        value_Pre = [str(Scan_i),Pass_Fail,*mpe_all,"No Dry out for One Cycle"]
        values_pos =[
            str_exp_1Cyc_text,
            "No RPT for One Cyc",
            "nan","nan",
            "nan","nan", 
            "nan","nan",
            "nan","nan","nan",
            str_error_1Cyc,log_messages]
        values = [*value_Pre,*values_para,*values_pos]
        values = [values,]
        own_logger.info(str_error_1Cyc)
        own_logger.info("Fail in {}".format(ModelTimer.time())) 
        # Round_No = f"Case_{Scan_i}_Exp_{index_exp}_{Temp_K}oC"
        book_name_xlsx_seperate =   str(Scan_i)+ '_' + book_name_xlsx;
        sheet_name_xlsx =  str(Scan_i);
        write_excel_xlsx(
            BasicPath + Target + "Excel/" +book_name_xlsx_seperate, 
            sheet_name_xlsx, values)
        my_dict_1Cyc["Error %"] = "nan"
        my_dict_1Cyc["Error ageT %"] = "nan"
        my_dict_1Cyc["punish"] = "nan"
    else: # main post-processing 
        mpe_all = Compare_Exp_Model_1Cyc_Dch( my_dict_1Cyc, XY_Exp, Scan_i,
            Temp_K,BasicPath, Target,fs,dpi, PlotCheck=True)
        [mpe_tot,punish] = mpe_all
        # set pass or fail TODO figure out how much should be appropriate:
        if mpe_tot < 3:
            Pass_Fail = "Pass"
        else:
            Pass_Fail = "Fail"
        my_dict_1Cyc["Error %"] = mpe_tot
        my_dict_1Cyc["punish"]  = punish
        ##########################################################
        #########      3-1: Plot cycle,location, Dryout related 
        if len(my_dict_1Cyc["CDend Porosity"])>1:
            Plot_Loc_AGE_4(
                my_dict_1Cyc,Scan_i,Temp_K+273.15,
                model_options,BasicPath, Target,fs,dpi)
        if Check_Small_Time == True:    
            own_logger.info(f"Scan {Scan_i}: Finish all plots within {SmallTimer.time()}")
            SmallTimer.reset()
        else:
            pass
        ##########################################################
        ##########################################################
        #########      3-3: Save summary to excel 
        # Get the log messages from the buffer and store it in a variable
        log_messages = log_buffer.getvalue()
        log_messages = log_messages.strip() # # Remove leading and trailing spaces from the log message
        # Remove the custom handler to prevent duplicate logs
        custom_logger.removeHandler(custom_handler)

        values=Get_Values_Excel(
            Pass_Fail,
            mpe_tot,punish,
            model_options,my_dict_1Cyc,"nan",
            "nan",Scan_i,Para_dict_i,"nan",
            str_exp_1Cyc_text,"nan",
            "nan",log_messages)
        values = [values,]
        book_name_xlsx_seperate =   str(Scan_i)+ '_' + book_name_xlsx;
        sheet_name_xlsx =  str(Scan_i);
        write_excel_xlsx(
            BasicPath + Target + "Excel/" + book_name_xlsx_seperate, 
            sheet_name_xlsx, values)
        #########      3-2: Save data as .mat or .json
        import json
        try:
            with open(
                # 'data.json', 
                BasicPath + Target+"Mats/" + str(Scan_i)+ '-StructData.json',
                'w'
                ) as json_file:
                    json.dump(my_dict_1Cyc, json_file) 
        except:
            print(f"Scan {Scan_i}: Encounter problems when saving json file!")
        else: 
            print(f"Scan {Scan_i}: Successfully save json file!")
        if Check_Small_Time == True:    
            print(f"Scan {Scan_i}: Try saving within {SmallTimer.time()}")
            SmallTimer.reset()
        else:
            pass
        own_logger.info("Succeed doing something in {}".format(ModelTimer.time()))
        own_logger.info(f'This is the end of No. {Scan_i} scan')

    Sol_RPT = [];  
    return my_dict_1Cyc,Sol_RPT,Sol_0,Call_1Cyc

def Run_model(
    Para_dict_i,BasicPath, XY_Exp, 
    purpose,    Exp_pack, keys_all,dpi,fs,
    flag_RunOneCyc,   Plot_Exp,Timeout,Return_Sol,
    Check_Small_Time ):

    
    if flag_RunOneCyc:
        midc_merge,Sol_RPT,Sol_AGE,Call_1Cyc = RunOneCyc(Para_dict_i,BasicPath, XY_Exp, 
            purpose,    Exp_pack, keys_all,dpi,fs,
            flag_RunOneCyc,   Plot_Exp,Timeout,Return_Sol,
            Check_Small_Time) # run only short cycle
    else:
        midc_merge,Sol_RPT,Sol_AGE,Call_1Cyc = Run_Age(
            Para_dict_i,BasicPath, XY_Exp, 
            purpose,    Exp_pack, keys_all,dpi,fs,
            Plot_Exp,Timeout,Return_Sol,
            Check_Small_Time )
    

    return midc_merge,Sol_RPT,Sol_AGE,Call_1Cyc

def Run_Age(
    Para_dict_i,BasicPath, XY_Exp, 
    purpose,    Exp_pack, keys_all,dpi,fs,
    Plot_Exp,Timeout,Return_Sol,
    Check_Small_Time ): # true or false to plot,timeout,return,check small time

    ##########################################################
    ##############    Part-0: Log of the scripts    ##########
    ##########################################################
    # add 221205: if Timeout=='True', use Patrick's version, disable pool
    #             else, use pool to accelerate 
    # add Return_Sol, on HPC, always set to False, as it is useless, 
    # add 230221: do sol_new['Throughput capacity [A.h]'].entries += sol_old['Throughput capacity [A.h]'].entries 
    #             and for "Throughput energy [W.h]", when use Model.set_initial_conditions_from
    #             this is done inside the two functions Run_Model_Base_On_Last_Solution(_RPT)
    # change 230621: add ability to scan one parameter set at different temperature 
    #                and at Exp-2,3,5
    # 230723: change for HPC paper
    ModelTimer = pb.Timer()
    Call_Age = "nan"
    if Check_Small_Time == True:
        SmallTimer = pb.Timer()

    # update 230731 - catch log information
    log_buffer = io.StringIO()
    # Set up the custom logger to capture the log messages
    custom_logger = logging.getLogger('pybamm')
    custom_handler = logging.StreamHandler(log_buffer)
    custom_logger.addHandler(custom_handler)
    custom_logger.setLevel(logging.WARNING) 
    # Set up a separate logger for your own print statements
    own_logger = logging.getLogger('my_own_logger')
    own_logger.setLevel(logging.INFO)  # Set the log level for your own print statements
    own_handler = logging.StreamHandler()  # Print to the console
    own_logger.addHandler(own_handler)
    
    # Your print statements within the Run_model function
    own_logger.info("Your own print statement")

    ##########################################################
    ##############    Part-1: Initialization    ##############
    ##########################################################
    font = {'family' : 'DejaVu Sans','size'   : fs}
    mpl.rc('font', **font)

    # new to define: exp_text_list,  Path_pack
    # exp_index_pack , 
    # define here:
    index_i   = Para_dict_i["Scan No"]  
    Scan_i = int(index_i)
    own_logger.info('Start Now! Scan %d.' % Scan_i)  
    Temp_K = Para_dict_i["Ageing temperature"]  
    Round_No = f"Case_{Scan_i}_{Temp_K}oC"  # index to identify different rounds of running 
    # Load cycling requirement
    [
        exp_AGE_text,exp_RPT_text,
        step_AGE_CD,step_AGE_CC,step_AGE_CV,
        step_0p1C_CD ,step_0p1C_CC,step_0p1C_RE,
        cycle_no,book_name_xlsx,
        ]  = Exp_pack

    # set up experiment
    Target  = f'/{purpose}/'
    Sol_RPT = [];  Sol_AGE = [];
    # pb.set_logging_level('INFO') # show more information!
    # set_start_method('fork') # from Patrick

    # Un-pack data:
    CyclePack,Para_0 = Para_init(Para_dict_i) # initialize the parameter
    [ 
        Total_Cycles,Cycle_bt_RPT,Update_Cycles,RPT_Cycles,
        Temper_i,Temper_RPT,mesh_list,submesh_strech,model_options]= CyclePack
    [keys_all_RPT,keys_all_AGE] = keys_all
    str_exp_AGE_text  = str(exp_AGE_text)
    str_exp_RPT_text  = str(exp_RPT_text)

    # define experiment
    Experiment_Long   = pb.Experiment( exp_AGE_text * Update_Cycles  )  
    # update 24-04-2023: delete GITT
    Experiment_RPT    = pb.Experiment( exp_RPT_text*1 ) 
    Experiment_Breakin= Experiment_RPT

    #####  index definition ######################
    Small_Loop =  int(Cycle_bt_RPT/Update_Cycles);   
    SaveTimes  =  int(Total_Cycles/Cycle_bt_RPT);   

    # initialize my_dict for outputs
    my_dict_RPT = {}
    for keys in keys_all_RPT:
        for key in keys:
            my_dict_RPT[key]=[];
    my_dict_AGE = {}; 
    for keys in keys_all_AGE:
        for key in keys:
            my_dict_AGE[key]=[]
    my_dict_RPT["Cycle_RPT"] = []
    my_dict_AGE["Cycle_AGE"] = []
    my_dict_RPT["avg_Age_T"] = [] # Update add 230617 
    Cyc_Update_Index     =[]
    call_logs = "nan"

    # update 220924: merge DryOut and Int_ElelyExces_Ratio
    temp_Int_ElelyExces_Ratio =  Para_0["Initial electrolyte excessive amount ratio"] 
    ce_EC_0 = Para_0['EC initial concentration in electrolyte [mol.m-3]'] # used to calculate ce_EC_All
    if temp_Int_ElelyExces_Ratio < 1:
        Int_ElelyExces_Ratio = -1;
        DryOut = "Off";
    else:
        Int_ElelyExces_Ratio = temp_Int_ElelyExces_Ratio;
        DryOut = "On";
    own_logger.info(f"Scan {Scan_i}: DryOut = {DryOut}")
    if DryOut == "On":  
        mdic_dry,Para_0 = Initialize_mdic_dry(Para_0,Int_ElelyExces_Ratio)
    else:
        mdic_dry ={}
    if Check_Small_Time == True:
        own_logger.info(f'Scan {Scan_i}: Spent {SmallTimer.time()} on Initialization')
        SmallTimer.reset()
    ##########################################################
    ##############    Part-2: Run model         ##############
    ##########################################################
    ##########################################################
    Timeout_text = 'I timed out'
    ##########    2-1: Define model and run break-in cycle
    try:  
        Timelimit = int(3600*3)
        # the following turns on for HPC only!
        if Timeout == True:
            timeout_RPT = TimeoutFunc(
                Run_Breakin, 
                timeout=Timelimit, 
                timeout_val=Timeout_text)
            Result_list_breakin  = timeout_RPT(
                model_options, Experiment_Breakin, 
                Para_0, mesh_list, submesh_strech)
        else:
            Result_list_breakin  = Run_Breakin(
                model_options, Experiment_Breakin, 
                Para_0, mesh_list, submesh_strech)
        [Model_0,Sol_0,Call_Breakin,error_logs] = Result_list_breakin
        if Return_Sol == True:
            Sol_RPT.append(Sol_0)
        if Call_Breakin.success == False:
            Sol_0 = "nan"
            call_logs = str(Call_1Cyc.logs)
            own_logger.info("Fail due to Experiment error or infeasible")
            1/0
        if Sol_0 == Timeout_text: # to do: distinguish different failure cases
            own_logger.info("Fail due to Timeout")
            1/0
        if Sol_0 == "Model error or solver error":
            call_logs = error_logs
            own_logger.info("Fail due to Model error or solver error")
            1/0
    except ZeroDivisionError as e:
        str_error_Breakin = str(e)
        if Check_Small_Time == True:
            own_logger.info(f"Scan {Scan_i}: Fail break-in cycle within {SmallTimer.time()}, need to exit the whole scan now due to {str_error_Breakin} but do not know how!")
            SmallTimer.reset()
        else:
            own_logger.info(f"Scan {Scan_i}: Fail break-in cycle, need to exit the whole scan now due to {str_error_Breakin} but do not know how!")
        Flag_Breakin = False 
    else:
        str_error_Breakin = "nan"
        if Check_Small_Time == True:    
            own_logger.info(f"Scan {Scan_i}: Finish break-in cycle within {SmallTimer.time()}")
            SmallTimer.reset()
        else:
            own_logger.info(f"Scan {Scan_i}: Finish break-in cycle")
        # post-process for break-in cycle - 0.1C only
        my_dict_RPT = GetSol_dict (my_dict_RPT,keys_all_RPT, Sol_0, 
            0, step_0p1C_CD, step_0p1C_CC,step_0p1C_RE , step_AGE_CV   )

        cycle_count =0
        my_dict_RPT["Cycle_RPT"].append(cycle_count)
        Cyc_Update_Index.append(cycle_count)
        Flag_Breakin = True
        if Check_Small_Time == True:    
            own_logger.info(f"Scan {Scan_i}: Finish post-process for break-in cycle within {SmallTimer.time()}")
            SmallTimer.reset()
        else:
            own_logger.info(f"Scan {Scan_i}: Finish post-process for break-in cycle")
        
    Flag_AGE = True; str_error_AGE_final = "Empty";   str_error_RPT = "Empty";
    #############################################################
    #######   2-2: Write a big loop to finish the long experiment    
    if Flag_Breakin == True: 
        k=0
        # Para_All.append(Para_0);Model_All.append(Model_0);Sol_All_i.append(Sol_0); 
        Para_0_Dry_old = Para_0;     Model_Dry_old = Model_0  ; Sol_Dry_old = Sol_0;   del Model_0,Sol_0
        while k < SaveTimes:    
            i=0  
            avg_Age_T = []  
            while i < Small_Loop:
                if DryOut == "On":
                    Data_Pack,Paraupdate   = Cal_new_con_Update (  Sol_Dry_old,   Para_0_Dry_old )
                if DryOut == "Off":
                    Paraupdate = Para_0
                # Run aging cycle:
                try:
                    Timelimit = int(60*60*2)
                    if Timeout == True:
                        timeout_AGE = TimeoutFunc(
                            Run_Model_Base_On_Last_Solution, 
                            timeout=Timelimit, 
                            timeout_val=Timeout_text)
                        Result_list_AGE = timeout_AGE( 
                            Model_Dry_old  , Sol_Dry_old , Paraupdate ,Experiment_Long, 
                            Update_Cycles,Temper_i,mesh_list,submesh_strech )
                    else:
                        Result_list_AGE = Run_Model_Base_On_Last_Solution( 
                            Model_Dry_old  , Sol_Dry_old , Paraupdate ,Experiment_Long, 
                            Update_Cycles,Temper_i,mesh_list,submesh_strech )
                    [Model_Dry_i, Sol_Dry_i , Call_Age ] = Result_list_AGE
                    if Return_Sol == True:
                        Sol_AGE.append(Sol_Dry_i)
                    #print(f"Temperature for ageing is now: {Temper_i}")  
                    if Call_Age.success == False:
                        own_logger.info("Fail due to Experiment error or infeasible")
                        str_error_AGE = "Experiment error or infeasible"
                        1/0
                    if Sol_Dry_i == Timeout_text: # fail due to timeout
                        own_logger.info("Fail due to Timeout")
                        str_error_AGE = "Timeout"
                        1/0
                    if Sol_Dry_i == "Model error or solver error":
                        own_logger.info("Fail due to Model error or solver error")
                        str_error_AGE = "Model error or solver error"
                        1/0
                except ZeroDivisionError as e: # ageing cycle fails
                    if Check_Small_Time == True:    
                        own_logger.info(f"Scan {Scan_i}: Fail during No.{Cyc_Update_Index[-1]} ageing cycles within {SmallTimer.time()} due to {str_error_AGE}")
                        SmallTimer.reset()
                    else:
                        own_logger.info(f"Scan {Scan_i}: Fail during No.{Cyc_Update_Index[-1]} ageing cycles due to {str_error_AGE}")
                    Flag_AGE = False
                    str_error_AGE_final = str_error_AGE
                    break
                else:                           # ageing cycle SUCCEED
                    Para_0_Dry_old = Paraupdate; Model_Dry_old = Model_Dry_i; Sol_Dry_old = Sol_Dry_i;   
                    del Paraupdate,Model_Dry_i,Sol_Dry_i
                    
                    if Check_Small_Time == True:    
                        own_logger.info(f"Scan {Scan_i}: Finish for No.{Cyc_Update_Index[-1]} ageing cycles within {SmallTimer.time()}")
                        SmallTimer.reset()
                    else:
                        own_logger.info(f"Scan {Scan_i}: Finish for No.{Cyc_Update_Index[-1]} ageing cycles")

                    # post-process for first ageing cycle and every -1 ageing cycle
                    if k==0 and i==0:    
                        my_dict_AGE = GetSol_dict (my_dict_AGE,keys_all_AGE, Sol_Dry_old, 
                            0, step_AGE_CD , step_AGE_CC , step_0p1C_RE, step_AGE_CV   )     
                        my_dict_AGE["Cycle_AGE"].append(1)
                    my_dict_AGE = GetSol_dict (my_dict_AGE,keys_all_AGE, Sol_Dry_old, 
                        cycle_no, step_AGE_CD , step_AGE_CC , step_0p1C_RE, step_AGE_CV   )    
                    cycle_count +=  Update_Cycles; 
                    avg_Age_T.append(np.mean(
                        Sol_Dry_old["Volume-averaged cell temperature [C]"].entries))
                    my_dict_AGE["Cycle_AGE"].append(cycle_count)           
                    Cyc_Update_Index.append(cycle_count)
                    
                    if DryOut == "On":
                        mdic_dry = Update_mdic_dry(Data_Pack,mdic_dry)
                    
                    if Check_Small_Time == True:    
                        own_logger.info(f"Scan {Scan_i}: Finish post-process for No.{Cyc_Update_Index[-1]} ageing cycles within {SmallTimer.time()}")
                        SmallTimer.reset()
                    else:
                        pass
                    i += 1;   ##################### Finish small loop and add 1 to i 
            
            # run RPT, and also update parameters (otherwise will have problems)
            if DryOut == "On":
                Data_Pack , Paraupdate  = Cal_new_con_Update (  Sol_Dry_old,   Para_0_Dry_old   )
            if DryOut == "Off":
                Paraupdate = Para_0     
            try:
                Timelimit = int(60*60*2)
                if Timeout == True:
                    timeout_RPT = TimeoutFunc(
                        Run_Model_Base_On_Last_Solution_RPT, 
                        timeout=Timelimit, 
                        timeout_val=Timeout_text)
                    Result_list_RPT = timeout_RPT(
                        Model_Dry_old  , Sol_Dry_old ,   
                        Paraupdate,      Experiment_RPT, RPT_Cycles, 
                        Temper_RPT ,mesh_list ,submesh_strech
                    )
                else:
                    Result_list_RPT = Run_Model_Base_On_Last_Solution_RPT(
                        Model_Dry_old  , Sol_Dry_old ,   
                        Paraupdate,      Experiment_RPT, RPT_Cycles, 
                        Temper_RPT ,mesh_list ,submesh_strech
                    )
                [Model_Dry_i, Sol_Dry_i,Call_RPT]  = Result_list_RPT
                if Return_Sol == True:
                    Sol_RPT.append(Sol_Dry_i)
                #print(f"Temperature for RPT is now: {Temper_RPT}")  
                if Call_RPT.success == False:
                    own_logger.info("Fail due to Experiment error or infeasible")
                    str_error_RPT = "Experiment error or infeasible"
                    1/0 
                if Sol_Dry_i == Timeout_text:
                    #own_logger.info("Fail due to Timeout")
                    str_error_RPT = "Timeout"
                    1/0
                if Sol_Dry_i == "Model error or solver error":
                    own_logger.info("Fail due to Model error or solver error")
                    str_error_RPT = "Model error or solver error"
                    1/0
            except ZeroDivisionError as e:
                if Check_Small_Time == True:    
                    own_logger.info(f"Scan {Scan_i}: Fail during No.{Cyc_Update_Index[-1]} RPT cycles within {SmallTimer.time()}, due to {str_error_RPT}")
                    SmallTimer.reset()
                else:
                    own_logger.info(f"Scan {Scan_i}: Fail during No.{Cyc_Update_Index[-1]} RPT cycles, due to {str_error_RPT}")
                break
            else:
                # post-process for RPT
                Cyc_Update_Index.append(cycle_count)
                if Check_Small_Time == True:    
                    own_logger.info(f"Scan {Scan_i}: Finish for No.{Cyc_Update_Index[-1]} RPT cycles within {SmallTimer.time()}")
                    SmallTimer.reset()
                else:
                    own_logger.info(f"Scan {Scan_i}: Finish for No.{Cyc_Update_Index[-1]} RPT cycles")
                my_dict_RPT = GetSol_dict (my_dict_RPT,keys_all_RPT, Sol_Dry_i, 
                    0,step_0p1C_CD, step_0p1C_CC,step_0p1C_RE , step_AGE_CV   )
                my_dict_RPT["Cycle_RPT"].append(cycle_count)
                my_dict_RPT["avg_Age_T"].append(np.mean(avg_Age_T))  # Make sure avg_Age_T and 
                if DryOut == "On":
                    mdic_dry = Update_mdic_dry(Data_Pack,mdic_dry)
                Para_0_Dry_old = Paraupdate;    Model_Dry_old = Model_Dry_i;
                Sol_Dry_old = Sol_Dry_i    ;   
                del Paraupdate,Model_Dry_i,Sol_Dry_i
                if Check_Small_Time == True:    
                    own_logger.info(f"Scan {Scan_i}: Finish post-process for No.{Cyc_Update_Index[-1]} RPT cycles within {SmallTimer.time()}")
                    SmallTimer.reset()
                else:
                    pass
                if Flag_AGE == False:
                    break
            k += 1 
    ############################################################# 
    #########   An extremely bad case: cannot even finish breakin
    if Flag_Breakin == False: 
        # Get the log messages from the buffer and store it in a variable
        log_messages = log_buffer.getvalue()
        log_messages = log_messages.strip() # # Remove leading and trailing spaces from the log message
        # Remove the custom handler to prevent duplicate logs
        custom_logger.removeHandler(custom_handler)
        
        value_list_temp = list(Para_dict_i.values())
        values_para = []
        for value_list_temp_i in value_list_temp:
            values_para.append(str(value_list_temp_i))
        # sequence: scan no, exp, pass or fail, mpe, dry-out, 
        mpe_all = ["nan","nan",
            "nan","nan", 
            "nan","nan","nan"]
        Pass_Fail = "Die"
        value_Pre = [str(Scan_i),Pass_Fail,*mpe_all,DryOut,]
        values_pos =[
            str_exp_AGE_text,
            str_exp_RPT_text,
            "nan","nan",
            "nan","nan", 
            "nan","nan",
            "nan","nan",
            "nan",str_error_Breakin,log_messages]
        values = [*value_Pre,*values_para,*values_pos]
        own_logger.info(str_error_Breakin)
        own_logger.info("Fail in {}".format(ModelTimer.time())) 
        # Round_No = f"Case_{Scan_i}_Exp_{index_exp}_{Temp_K}oC"
        book_name_xlsx_seperate =   str(Scan_i)+ '_' + book_name_xlsx;
        sheet_name_xlsx =  str(Scan_i);
        write_excel_xlsx(
            BasicPath + Target + "Excel/" +book_name_xlsx_seperate, 
            sheet_name_xlsx, values)
        my_dict_RPT["Error %"] = "nan"
        my_dict_RPT["Error SOH %"] = "nan"
        my_dict_RPT["Error LLI %"] = "nan"
        my_dict_RPT["Error LAM NE %"] = "nan"
        my_dict_RPT["Error LAM PE %"] = "nan"
        my_dict_RPT["Error Res %"] = "nan"
        my_dict_RPT["Error ageT %"] = "nan"
        my_dict_RPT["punish"] = "nan"
        
        midc_merge = {**my_dict_RPT, **my_dict_AGE,**mdic_dry}

        return midc_merge,Sol_RPT,Sol_AGE
    ##########################################################
    ##############   Part-3: Post-prosessing    ##############
    ##########################################################
    # Newly add (220517): save plots, not just a single line in excel file:     
    # Newly add (221114): make plotting as functions
    # Niall_data = loadmat( 'Extracted_all_cell.mat'
    else:
        #if not os.path.exists(BasicPath + Target + str(Scan_i)):
        #os.mkdir(BasicPath + Target + str(Scan_i) );
        # Update 230221 - Add model LLI, LAM manually 
        my_dict_RPT['Throughput capacity [kA.h]'] = (
            np.array(my_dict_RPT['Throughput capacity [A.h]'])/1e3).tolist()
        my_dict_RPT['CDend SOH [%]'] = ((
            np.array(my_dict_RPT["Discharge capacity [A.h]"])
            /my_dict_RPT["Discharge capacity [A.h]"][0])*100).tolist()
        
        if Check_Small_Time == True:    
            own_logger.info(f"Scan {Scan_i}: Getting extra variables within {SmallTimer.time()}")
            SmallTimer.reset()
        else:
            pass

        [X_exp,Y_exp] = XY_Exp  # interpolate for exp only
        mpe_all = Compare_Exp_Model( my_dict_RPT, XY_Exp, Scan_i,
            Temper_i,BasicPath, Target,fs,dpi, PlotCheck=True)
        [mpe_tot,punish] = mpe_all
        # set pass or fail TODO figure out how much should be appropriate:
        if mpe_tot < 3:
            Pass_Fail = "Pass"
        else:
            Pass_Fail = "Fail"
        my_dict_RPT["Error %"] = mpe_tot
        my_dict_RPT["punish"] = punish

        ##########################################################
        #########      3-1: Plot cycle,location, Dryout related 
        # update 23-05-25 there is a bug in Cyc_Update_Index, need to slide a bit:
        Cyc_Update_Index.insert(0,0); del Cyc_Update_Index[-1]

        Plot_Cyc_RPT_4(
            my_dict_RPT, XY_Exp, Plot_Exp,  
            Scan_i,Temper_i,model_options,
            BasicPath, Target,fs,dpi)
        if len(my_dict_AGE["CDend Porosity"])>1:
            Plot_Loc_AGE_4(
                my_dict_AGE,Scan_i,Temper_i,
                model_options,BasicPath, Target,fs,dpi)
        if DryOut == "On":
            Plot_Dryout(
                Cyc_Update_Index,mdic_dry,ce_EC_0,Temper_i,
                Scan_i,BasicPath, Target,fs,dpi)
        if Check_Small_Time == True:    
            own_logger.info(f"Scan {Scan_i}: Finish all plots within {SmallTimer.time()}")
            SmallTimer.reset()
        else:
            pass
        # update 23-05-26 judge ageing shape: contain 3 index
        """ [Start_shape,Middle_shape,End_shape] = check_concave_convex(
        my_dict_RPT['Throughput capacity [kA.h]'], 
        my_dict_RPT['CDend SOH [%]']) """
        ##########################################################
        ##########################################################
        #########      3-3: Save summary to excel 
        # Get the log messages from the buffer and store it in a variable
        log_messages = log_buffer.getvalue()
        log_messages = log_messages.strip() # # Remove leading and trailing spaces from the log message
        # Remove the custom handler to prevent duplicate logs
        custom_logger.removeHandler(custom_handler)

        values=Get_Values_Excel(
            Pass_Fail,
            mpe_tot,punish,
            model_options,my_dict_RPT,mdic_dry,
            DryOut,Scan_i,Para_dict_i,str_exp_AGE_text,
            str_exp_RPT_text,str_error_AGE_final,
            str_error_RPT,log_messages)
        values = [values,]
        book_name_xlsx_seperate =   str(Scan_i)+ '_' + book_name_xlsx;
        sheet_name_xlsx =  str(Scan_i);
        write_excel_xlsx(
            BasicPath + Target + "Excel/" + book_name_xlsx_seperate, 
            sheet_name_xlsx, values)
        # pick only the cyc - most concern
        Keys_cyc_mat =[
            'Throughput capacity [kA.h]',
            'CDend SOH [%]',
        ]
        my_dict_mat = {}
        for key in Keys_cyc_mat:
            my_dict_mat[key]=my_dict_RPT[key]
        #########      3-2: Save data as .mat or .json
        my_dict_RPT["Cyc_Update_Index"] = Cyc_Update_Index
        my_dict_RPT["SaveTimes"]    = SaveTimes
        midc_merge = {**my_dict_RPT, **my_dict_AGE,**mdic_dry}
        import json
        try:
            with open(
                # 'data.json', 
                BasicPath + Target+"Mats/" + str(Scan_i)+ '-StructData.json',
                'w'
                ) as json_file:
                    json.dump(midc_merge, json_file) 
        except:
            own_logger.info(f"Scan {Scan_i}: Encounter problems when saving json file!")
        else: 
            own_logger.info(f"Scan {Scan_i}: Successfully save json file!")
        try:
            savemat(
                BasicPath + Target+"Mats/" 
                + str(Scan_i)+ '-Ageing_summary_only.mat',
                my_dict_mat)  
        except:
            own_logger.info(f"Scan {Scan_i}: Encounter problems when saving mat file!")
        else: 
            own_logger.info(f"Scan {Scan_i}: Successfully save mat file!")

        if Check_Small_Time == True:    
            own_logger.info(f"Scan {Scan_i}: Try saving within {SmallTimer.time()}")
            SmallTimer.reset()
        else:
            pass
        own_logger.info("Succeed doing something in {}".format(ModelTimer.time()))
        own_logger.info(f'This is the end of No. {Scan_i} scan')
        return midc_merge,Sol_RPT,Sol_AGE,Call_Age

# Functions, will put into the .py file later
def save_rows_to_csv(file_path, rows, header):
    with open(file_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(header)  # Write parameter names as the header row
        writer.writerows(rows)
def generate_combinations(Bounds, Num_tot):
    lower_bounds = []
    upper_bounds = []
    for bound in Bounds:
        lower_bounds.append(bound[0])
        upper_bounds.append(bound[1])
    combinations = []
    for _ in range(Num_tot):
        combination = []
        for lower, upper in zip(lower_bounds, upper_bounds):
            value = random.uniform(lower, upper)
            combination.append(value)
        combinations.append(combination)
    return combinations

def save_combinations_to_csv(combinations, parameter_names, filename):
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(parameter_names)  # Write parameter names as the first row
        for combination in combinations:
            writer.writerow(combination)
# Pack input needs to be a list
def Get_Scan_files(
        BasicPath,Target_name,model_options,
        parameter_names,para_short_name,
        Pack,
        rows_per_file,Bundle):
    
    import itertools
    para_dict_Same = {
        "Cycles within RPT":1,
        "RPT temperature":25,
        "Mesh list":[5,5,5,60,20],  
        "Para_Set": "OKane2023",
        "Model option":model_options,
        "Current solvent concentration in the reservoir [mol.m-3]":4541.0,
        "Current electrolyte concentration in the reservoir [mol.m-3]":1000,
        "Ratio of Li-ion concentration change in " 
        "electrolyte consider solvent consumption":1.0,
        'EC initial concentration in electrolyte [mol.m-3]':4541.0,
        'Typical EC concentration in electrolyte [mol.m-3]':4541.0, 
        "Negative electrode number of cracks per unit area [m-2]": 3.18e15,
        "Initial inner SEI thickness [m]": 1.23625e-08,
        "Initial outer SEI thickness [m]": 1.23625e-08,
        "Negative electrode porosity": 0.222393,
        }
    unchange_key2 = list(para_dict_Same.keys())
    unchange_val2 = list(para_dict_Same.values())
    short_pack = [lst for lst in Pack if len(lst) > 1]
    selected_indices = [i for i, lst in enumerate(Pack) if len(lst) > 1]
    shortList_para_short_name = [para_short_name[i] for i in selected_indices]
    shortList_para_short_name.insert(0,"No")
    really_change_val =  [
        list(comb) for comb in itertools.product(*short_pack)]

    change_val =  [
        list(comb) for comb in itertools.product(*Pack)]
    combinations = [[i+1,*elem, *unchange_val2] for i,elem in enumerate(change_val)]
    comb_short   = [[i+1,*elem] for i,elem in enumerate(really_change_val)]
    parameter_names = [*parameter_names,*unchange_key2]
    print("Total cases number is",len(combinations))
    if Bundle:
        # Specify the total number of cases
        total_cases = len(combinations)
        # Specify the number of rows per CSV file, rows_per_file
        # Calculate the number of files needed
        num_files = (total_cases - 1) // rows_per_file + 1
        # Create the target folder
        folder_path = os.path.join(BasicPath, "Get_Random_sets", Target_name)
        os.makedirs(folder_path, exist_ok=True)
        # Write data to each CSV file
        for i in range(num_files):
            file_name = f"Bundle_{i+1}.csv"
            file_path = os.path.join(folder_path, file_name)
            start_row = i * rows_per_file
            end_row = min(start_row + rows_per_file, total_cases)
            rows = combinations[start_row:end_row]
            save_rows_to_csv(file_path, rows, parameter_names)
        filename = BasicPath+f"/Get_Random_sets/{Target_name}/"+f'{Target_name}.csv'
        filename_short = BasicPath+f"/Get_Random_sets/{Target_name}/"+f'{Target_name}_s.csv'
        save_combinations_to_csv(combinations, parameter_names, filename)
        save_combinations_to_csv(comb_short, shortList_para_short_name, filename_short)
        print(f"Combinations saved to '{Target_name}.csv' file.") 
        print(f"CSV files created in folder '{Target_name}'.")
    else:
        filename = BasicPath+"/Get_Random_sets/"+f'{Target_name}.csv'
        save_combinations_to_csv(combinations, parameter_names, filename)
        print(f"Combinations saved to '{Target_name}.csv' file.") 
    return len(combinations)
# pack input can not either be a list or a tuple
def get_list_from_tuple(d, num):
    if d[1] > d[0]:
        if d[1] > 100 * d[0]:
            result_list = (np.exp(np.linspace(np.log(d[0]), np.log(d[1]), num=num))).tolist()
        else:
            result_list = (np.linspace(d[0], d[1], num=num)).tolist()
    else:
        result_list = []
    return result_list
def Get_Scan_Orth_Latin(
        BasicPath,Target_name,model_options,
        parameter_names,para_short_name,
        Pack, num,
        rows_per_file,Bundle):
    
    import itertools; from pyDOE import lhs
    para_dict_Same = {
        "Cycles within RPT":1,
        "RPT temperature":25,
        "Mesh list":[5,5,5,60,20],  
        "Para_Set": "OKane2023",
        "Model option":model_options,
        "Current solvent concentration in the reservoir [mol.m-3]":4541.0,
        "Current electrolyte concentration in the reservoir [mol.m-3]":1000,
        "Ratio of Li-ion concentration change in " 
        "electrolyte consider solvent consumption":1.0,
        'EC initial concentration in electrolyte [mol.m-3]':4541.0,
        'Typical EC concentration in electrolyte [mol.m-3]':4541.0, 
        "Negative electrode number of cracks per unit area [m-2]": 3.18e15,
        "Initial inner SEI thickness [m]": 1.23625e-08,
        "Initial outer SEI thickness [m]": 1.23625e-08,
        "Negative electrode porosity": 0.222393,
        }
    unchange_key2 = list(para_dict_Same.keys())
    unchange_val2 = list(para_dict_Same.values())
    
    Pack_tuple = []; Pack_tuple_index = []
    Pack_list = [];  Pack_list_index  = []
    for i,item in enumerate(Pack):
        if isinstance(item, tuple):
            Pack_tuple.append(item)
            Pack_tuple_index.append(i)
        elif isinstance(item, list):
            Pack_list.append(item)
            Pack_list_index.append(i)
    com_tuple = []; comb_tu_list =[]
    if len(Pack_tuple) > 1:
        for tuple_i in Pack_tuple:
            com_tuple.append( get_list_from_tuple(tuple_i, num) )
        # apply Latin Hypercube:
        #print(com_tuple)
        samples = lhs(len(com_tuple), samples=num)
        for sample in samples:
            combination = []
            for i, candidate_list in enumerate(com_tuple):
                index = int(sample[i] * num)
                combination.append(candidate_list[index])
            comb_tu_list.append(combination)
    else:
        print("error! Pack_tuple must has 2 elements")
    # apply product sampling:
    comb_li_list = [list(comb) for comb in itertools.product(*Pack_list)]
    #print(comb_tu_list)
    #print(comb_li_list)
    Big_Comb = []
    for comb_tu in comb_tu_list:
        for comb_li in comb_li_list:
            big_comb = [0] * (len(comb_tu)+len(comb_li))
            for comb_tu_i,index in zip(comb_tu,Pack_tuple_index):
                big_comb[index] = comb_tu_i
            for comb_li_i,index in zip(comb_li,Pack_list_index):
                big_comb[index] = comb_li_i
            Big_Comb.append(big_comb)
    #print(Big_Comb)
    Big_Comb
    combinations = [[i+1,*elem, *unchange_val2] for i,elem in enumerate(Big_Comb)]

    parameter_names = [*parameter_names,*unchange_key2]
    print("Total cases number is",len(combinations))
    if Bundle:
        # Specify the total number of cases
        total_cases = len(combinations)
        # Specify the number of rows per CSV file, rows_per_file
        # Calculate the number of files needed
        num_files = (total_cases - 1) // rows_per_file + 1
        # Create the target folder
        folder_path = os.path.join(BasicPath, "Get_Random_sets", Target_name)
        os.makedirs(folder_path, exist_ok=True)
        # Write data to each CSV file
        for i in range(num_files):
            file_name = f"Bundle_{i+1}.csv"
            file_path = os.path.join(folder_path, file_name)
            start_row = i * rows_per_file
            end_row = min(start_row + rows_per_file, total_cases)
            rows = combinations[start_row:end_row]
            save_rows_to_csv(file_path, rows, parameter_names)
        filename = BasicPath+f"/Get_Random_sets/{Target_name}/"+f'{Target_name}.csv'
        filename_short = BasicPath+f"/Get_Random_sets/{Target_name}/"+f'{Target_name}_s.csv'
        save_combinations_to_csv(combinations, parameter_names, filename)
        # save_combinations_to_csv(comb_short, shortList_para_short_name, filename_short)
        print(f"Combinations saved to '{Target_name}.csv' file.") 
        print(f"CSV files created in folder '{Target_name}'.")
    else:
        filename = BasicPath+"/Get_Random_sets/"+f'{Target_name}.csv'
        save_combinations_to_csv(combinations, parameter_names, filename)
        print(f"Combinations saved to '{Target_name}.csv' file.") 
    return len(combinations)

def Get_Scan_General(
        BasicPath,Target_name,Para_dict, 
        num,rows_per_file,Bundle=True):

    import itertools; from pyDOE import lhs
    ParaDict_unchange ={}; ParaDict_change_List = {}; ParaDict_change_Tuple = {}
    for key, value in Para_dict.items():
        # depending on what values is:
        # it is a list, then it is changing
        if isinstance(value,list):
            if len(value) >1:
                #print(value)
                ParaDict_change_List[key] = value
            else:
                ParaDict_unchange[key] = value[0]
        # it is a list, so it also changes
        elif isinstance(value,tuple):
            #print(value)
            # create a list based on the tuple:
            ParaDict_change_Tuple[key] = get_list_from_tuple(value, num)
        else:
            ParaDict_unchange[key] = value
    if len(ParaDict_change_List) > 1:
        comb_li_list = [list(comb) for comb in itertools.product(*ParaDict_change_List.values())]
    elif len(ParaDict_change_List) == 0:
        comb_li_list = []
    else:
        comb_li_list = list(ParaDict_change_List.values())
        #print(comb_li_list)
        comb_li_list = comb_li_list [0]
    #comb_li_list
    
    comb_tu_list =[]
    if len(ParaDict_change_Tuple) > 1:
        samples = lhs(len(ParaDict_change_Tuple), samples=num)
        for sample in samples:
            combination = []
            for i, candidate_list in enumerate(ParaDict_change_Tuple.values()):
                index = int(sample[i] * num)
                combination.append(candidate_list[index])
            comb_tu_list.append(combination)
    elif len(ParaDict_change_Tuple) == 0:
        pass
    else:
        comb_tu_list = list(ParaDict_change_Tuple.values())
        comb_tu_list = comb_tu_list[0]
    #comb_tu_list
    Big_Comb = []
    if comb_li_list and comb_tu_list:  # Check if both lists are not empty
        for comb_li in comb_li_list:
            for comb_tu in comb_tu_list:
                if isinstance(comb_tu, list):
                    if isinstance(comb_li, list):
                        Big_Comb.append([*comb_li, *comb_tu])
                    else:
                        Big_Comb.append([comb_li, *comb_tu])
                else:
                    if isinstance(comb_li, list):
                        Big_Comb.append([*comb_li, comb_tu])
                    else:
                        Big_Comb.append([comb_li, comb_tu])
    else:
        if comb_li_list:  # Only comb_tu_list is empty
            for comb_li in comb_li_list:
                Big_Comb.append(comb_li)
        elif comb_tu_list:  # Only comb_li_list is empty
            for comb_tu in comb_tu_list:
                if isinstance(comb_tu, list):
                    Big_Comb.append(comb_tu)
                else:
                    Big_Comb.append([comb_tu])

    # Print or do something with Big_Comb
    print(Big_Comb)
    if len(Big_Comb) > 1:
        print(f"Total scan number is: {len(Big_Comb)}")
        comb_short = [[i+1,*elem, ] for i,elem in enumerate(Big_Comb)]
        short_name_List =  [
            "Scan No",
            *list(ParaDict_change_List.keys()),
            *list(ParaDict_change_Tuple.keys()),]
        combinations = [[i+1,*elem, *ParaDict_unchange.values()] for i,elem in enumerate(Big_Comb)]
        parameter_names = [
            "Scan No",
            *list(ParaDict_change_List.keys()),
            *list(ParaDict_change_Tuple.keys()),
            *list(ParaDict_unchange.keys())]
    else:
        print(f"Total scan number is: 1")
        combinations = [[1, *ParaDict_unchange.values()]]
        parameter_names = [
            "Scan No",
            *list(ParaDict_unchange.keys())]

    if Bundle:
        # Specify the total number of cases
        total_cases = len(combinations)
        # Specify the number of rows per CSV file, rows_per_file
        # Calculate the number of files needed
        num_files = (total_cases - 1) // rows_per_file + 1
        folder_path = os.path.join(BasicPath,  Target_name)
        os.makedirs(folder_path, exist_ok=True)
        # Write data to each CSV file
        for i in range(num_files):
            file_name = f"Bundle_{i+1}.csv"
            file_path = os.path.join(folder_path, file_name)
            start_row = i * rows_per_file
            end_row = min(start_row + rows_per_file, total_cases)
            rows = combinations[start_row:end_row]
            save_rows_to_csv(file_path, rows, parameter_names)
        filename = BasicPath+f"/{Target_name}/"+f'{Target_name}.csv'
        filename_short = BasicPath+f"/{Target_name}/"+f'{Target_name}_s.csv'
        save_combinations_to_csv(combinations, parameter_names, filename)
        if len(Big_Comb) > 1:
            save_combinations_to_csv(comb_short, short_name_List, filename_short)
        print(f"Combinations saved to '{Target_name}.csv' file.") 
        print(f"CSV files created in folder '{Target_name}'.")
    else:
        filename = BasicPath+"/"+f'{Target_name}.csv'
        filename_short = BasicPath+f"/"+f'{Target_name}_s.csv'
        save_combinations_to_csv(combinations, parameter_names, filename)
        if len(Big_Comb) > 1:
            save_combinations_to_csv(comb_short, short_name_List, filename_short)
        print(f"Combinations saved to '{Target_name}.csv' file.") 
    return parameter_names,combinations



import shutil
def Move_png(Root_Path,rows_per_file,Scan_end_end,purpose_i):
    Path_purpose = os.path.join(Root_Path, purpose_i)
    source_folders = []
    for i_bundle in range(int(Scan_end_end/rows_per_file)):
        Scan_start = (i_bundle)*rows_per_file+1;    
        Scan_end   = min(Scan_start + rows_per_file-1, Scan_end_end)    
        purpose = f"{purpose_i}_Case_{Scan_start}_{Scan_end}"
        source_folders.append(purpose)
        #print(purpose)

    # Create the Plot_Collect folder if it doesn't exist
    plot_collect_directory = os.path.join(Path_purpose, "Plot_Collect")
    os.makedirs(plot_collect_directory, exist_ok=True)

    # Move the .png files to the Plot_Collect folder
    for folder in source_folders:
        plots_directory = os.path.join(Path_purpose, folder, "Plots")
        if os.path.exists(plots_directory):
            for filename in os.listdir(plots_directory):
                if filename.startswith("0_") and filename.endswith(".png"):
                    source_file = os.path.join(plots_directory, filename)
                    destination_file = os.path.join(plot_collect_directory, filename)
                    shutil.move(source_file, destination_file)
    return 

def Copy_1st_row(source_file,target_file):
    # Load source workbook and target workbook
    source_workbook = openpyxl.load_workbook(source_file)
    target_workbook = openpyxl.load_workbook(target_file)
    # Assuming there's only one sheet in each workbook
    source_sheet = source_workbook.active
    target_sheet = target_workbook.active
    # Read the first row from the source workbook
    first_row_values = [cell.value for cell in source_sheet[1]]
    # Append the first row values to the target workbook
    target_sheet.append(first_row_values)
    # Save the changes to the target workbook
    target_workbook.save(target_file)
    # print("First row from", source_file, "has been copied and appended to the last row of", target_file)
    target_workbook.close()
    return 
def Collect_Excel(BasicPath,target_file,purpose_i,rows_per_file,Scan_end_end): 
    Scan_start_all = (
        np.arange(1,Scan_end_end+1,rows_per_file)).tolist()
    Scan_end_all = (
        np.arange(rows_per_file,Scan_end_end+rows_per_file,rows_per_file)
        ).tolist()
    for Scan_start,Scan_end in zip(Scan_start_all,Scan_end_all):
        Indexs =np.arange(Scan_start,Scan_end+1)
        book_name_xlsx = f'Summary_{purpose_i}_Case_{Scan_start}_{Scan_end}.xlsx';
        for k,index_i in enumerate(Indexs):
            #print(index_i)
            try:
                old_book = str(index_i) + '_' + book_name_xlsx
                source_file=BasicPath+f"/{purpose_i}_Case_{Scan_start}_{Scan_end}/"+"Excel/"+old_book
                Copy_1st_row(source_file,target_file)
            except:
                print(f"Something goes wrong for Scan {index_i}!")
            else:
                print(f"Successfuly write results for Scan {index_i}!") 
    return




























































